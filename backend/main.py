import hmac
import logging
import os
from datetime import date as date_type
from typing import List, Optional
from urllib.parse import quote as url_quote

log = logging.getLogger("main")

from fastapi import FastAPI, Depends, HTTPException, Query, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse, RedirectResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from starlette.middleware.base import BaseHTTPMiddleware

from database import init_db, get_db, engine, SessionLocal
from models import Produit, Pompe, Releve, Utilisateur, Role, Livraison, PrixVente, Employe, FichePaie, Depense, Achat, ParametreDepense, OTPCode, LoginSecurityEvent, SessionToken
from otp_service import (
    OTP_ENABLED, OTP_PENDING_COOKIE, OTP_PENDING_MAX_AGE,
    create_otp, send_otp_email, verify_otp, cleanup_expired_otps, _mask_email,
    create_admin_code, send_admin_code_email, verify_admin_code, EMAIL_USER,
)
from activity_log import (
    log_event,
    LOGIN_SUCCESS, LOGIN_FAILED, LOGOUT,
    OTP_SENT, OTP_VERIFIED, OTP_FAILED,
    ADMIN_CODE_REQUESTED, ADMIN_CODE_VERIFIED, ADMIN_CODE_FAILED,
)
from pos_routes import router as pos_router
from pos_analyse_routes import router as pos_analyse_router
from caisse_routes import router as caisse_router
from hotel_routes import router as hotel_router
from cuisine_routes import router as cuisine_router
from admin_routes import router as admin_router
from zelle_routes import router as zelle_router
from auth import (
    SESSION_COOKIE, hash_password, verify_password,
    hash_code_acces, verify_code_acces,
    create_session, get_session_user, delete_session,
    make_api_key, verify_api_key, revoke_api_key,
)
from password_reset import request_reset, verify_reset_token, consume_reset_token

app = FastAPI(title="Suivi des Meters - Station")

app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"]
)

PERIODES = ["Matin", "Apres-midi"]

# Bug 9 fix : constante au niveau module, plus de magic number dans la fonction
MAX_MODIFICATIONS_PAR_RELEVE = 2

# Clé API statique admin depuis .env (override de secours pour scripts/CI)
_ADMIN_API_KEY = os.getenv("ADMIN_API_KEY", "")

# Chemins accessibles sans être connecté
_PUBLIC_PATHS    = {"/login", "/api/login", "/api/otp/verify", "/api/otp/request-admin-code", "/api/otp/verify-admin-code",
                    "/api/auth/forgot-password", "/api/auth/reset-password", "/api/auth/reset-password/verify"}
_PUBLIC_PREFIXES = ("/docs", "/redoc", "/openapi.json", "/api/auth/oauth/")


class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        if path in _PUBLIC_PATHS or path.startswith(_PUBLIC_PREFIXES):
            return await call_next(request)

        token      = request.cookies.get(SESSION_COOKIE)
        api_key_hdr = request.headers.get("X-API-Key", "")

        db = SessionLocal()
        user = None
        try:
            user = get_session_user(db, token)
            if not user and api_key_hdr:
                # Clé statique admin depuis .env (timing-safe)
                if _ADMIN_API_KEY and hmac.compare_digest(api_key_hdr, _ADMIN_API_KEY):
                    user = db.query(Utilisateur).filter_by(username="admin", actif=True).first()
                else:
                    user = verify_api_key(db, api_key_hdr)
        finally:
            db.close()

        if not user:
            if path.startswith("/api/"):
                return JSONResponse({"detail": "Non authentifié"}, status_code=401)
            return RedirectResponse(url="/login")

        # Stocker les attributs scalaires dans un objet simple avant la fermeture de session
        class _UserProxy:
            __slots__ = ("id","role","role_id","actif","nom_complet","username","poste","email","api_key_hash")
        proxy = _UserProxy()
        proxy.id           = user.id
        proxy.role         = user.role
        proxy.role_id      = user.role_id
        proxy.actif        = user.actif
        proxy.nom_complet  = user.nom_complet
        proxy.username     = user.username
        proxy.poste        = user.poste
        proxy.email        = user.email
        proxy.api_key_hash = user.api_key_hash
        request.state.user = proxy
        return await call_next(request)


app.add_middleware(AuthMiddleware)

# Module POS bar/restaurant
app.include_router(pos_router)
app.include_router(pos_analyse_router)
app.include_router(caisse_router)
app.include_router(hotel_router)
app.include_router(cuisine_router)
app.include_router(admin_router)
app.include_router(zelle_router)


@app.on_event("startup")
def startup():
    init_db()
    # Migration douce — ajoute source_fond si la colonne n'existe pas encore
    from sqlalchemy import text as _text
    try:
        with engine.connect() as _c:
            _c.execute(_text("ALTER TABLE zelle_transactions ADD COLUMN IF NOT EXISTS source_fond VARCHAR(50)"))
            _c.commit()
    except Exception:
        pass
    from datetime import datetime, timezone as _tz
    from models import SessionToken as _ST
    with SessionLocal() as _db:
        # Purge des sessions expirées
        _db.query(_ST).filter(_ST.expires_at < datetime.now(_tz.utc)).delete()
        _db.commit()
        # Purge des OTP expirés ou utilisés
        try:
            n = cleanup_expired_otps(_db)
            if n:
                import logging; logging.getLogger("otp").info("Startup: %d OTP expirés supprimés", n)
        except Exception:
            pass  # Table pas encore créée (avant migration)


# ---------- Authentification ----------
_CODE_RE = __import__("re").compile(r"^\d{9}$")


class LoginIn(BaseModel):
    email:      str
    password:   str
    code_acces: str


@app.post("/api/login")
def login(data: LoginIn, request: Request, response: Response, db: Session = Depends(get_db)):
    ip    = request.client.host if request.client else None
    ua    = request.headers.get("user-agent", "")
    email = data.email.strip().lower()
    user  = db.query(Utilisateur).filter_by(email=email, actif=True).first()
    _err  = "Email, mot de passe ou code d'accès incorrect"
    if not user:
        log_event(db, LOGIN_FAILED, ip_address=ip, details={"email": email, "raison": "utilisateur_inconnu"})
        raise HTTPException(401, _err)
    if not verify_password(data.password, user.password_hash):
        log_event(db, LOGIN_FAILED, user_id=user.id, ip_address=ip, details={"raison": "mot_de_passe_incorrect"})
        raise HTTPException(401, _err)
    if not user.code_acces_hash or not verify_code_acces(data.code_acces, user.code_acces_hash):
        log_event(db, LOGIN_FAILED, user_id=user.id, ip_address=ip, details={"raison": "pin_incorrect"})
        raise HTTPException(401, _err)

    # ── Vérification en deux étapes (OTP) ────────────────────────────
    if OTP_ENABLED:
        if not user.email:
            raise HTTPException(400, "Aucun email associé à ce compte — contactez l'administrateur.")
        # Domaines internes/fictifs qui ne peuvent pas recevoir d'email
        _bad_tlds = {".local", ".localhost", ".internal", ".test", ".example", ".invalid", ".lan"}
        if any(user.email.lower().endswith(t) for t in _bad_tlds):
            raise HTTPException(503,
                f"L'adresse email de ce compte ({user.email}) ne peut pas recevoir de messages. "
                "Contactez l'administrateur pour mettre à jour l'email.")
        try:
            code, pending_token = create_otp(db, user.id)
            send_otp_email(user.nom_complet or user.username, user.email, code)
        except ValueError as e:
            raise HTTPException(429, str(e))
        except RuntimeError as e:
            raise HTTPException(503, str(e))

        log_event(db, OTP_SENT, user_id=user.id, ip_address=ip)
        response.set_cookie(
            OTP_PENDING_COOKIE, pending_token,
            httponly=True, samesite="lax", max_age=OTP_PENDING_MAX_AGE, path="/",
        )
        return {
            "otp_required": True,
            "email_hint":   _mask_email(user.email),
        }

    # ── Connexion directe (OTP désactivé) ────────────────────────────
    token = create_session(db, user.id, ip_address=ip, user_agent=ua)
    log_event(db, LOGIN_SUCCESS, user_id=user.id, ip_address=ip)
    response.set_cookie(
        SESSION_COOKIE, token,
        httponly=True, samesite="lax", max_age=7 * 24 * 3600, path="/",
    )
    raw_key = make_api_key(db, user.id)
    return {
        "id": user.id, "username": user.username,
        "nom_complet": user.nom_complet, "role": user.role,
        "api_key": raw_key,
    }


class OTPVerifyIn(BaseModel):
    code: str


@app.post("/api/otp/verify")
def otp_verify(data: OTPVerifyIn, request: Request, response: Response, db: Session = Depends(get_db)):
    """Étape 2 — vérifie le code OTP et ouvre la session si valide."""
    ip            = request.client.host if request.client else None
    ua            = request.headers.get("user-agent", "")
    pending_token = request.cookies.get(OTP_PENDING_COOKIE, "")
    try:
        user = verify_otp(db, pending_token, data.code)
    except ValueError as e:
        log_event(db, OTP_FAILED, ip_address=ip, details={"raison": str(e)})
        raise HTTPException(401, str(e))

    log_event(db, OTP_VERIFIED, user_id=user.id, ip_address=ip)
    response.delete_cookie(OTP_PENDING_COOKIE, path="/")

    token = create_session(db, user.id, ip_address=ip, user_agent=ua)
    log_event(db, LOGIN_SUCCESS, user_id=user.id, ip_address=ip)
    response.set_cookie(
        SESSION_COOKIE, token,
        httponly=True, samesite="lax", max_age=7 * 24 * 3600, path="/",
    )
    raw_key = make_api_key(db, user.id)
    return {
        "id": user.id, "username": user.username,
        "nom_complet": user.nom_complet, "role": user.role,
        "api_key": raw_key,
    }


@app.post("/api/otp/request-admin-code")
def otp_request_admin_code(request: Request, response: Response, db: Session = Depends(get_db)):
    """
    Génère un code 5 chiffres et l'envoie à l'email administrateur.
    L'employé doit contacter l'admin verbalement pour obtenir ce code.
    Prolonge le cookie pending_token à 60 min pour laisser le temps à l'admin de répondre.
    """
    ip            = request.client.host if request.client else None
    pending_token = request.cookies.get(OTP_PENDING_COOKIE, "")

    if not pending_token:
        raise HTTPException(400, "Session expirée — recommencez la connexion.")

    otp = db.query(OTPCode).filter(OTPCode.pending_token == pending_token).first()
    if not otp:
        raise HTTPException(400, "Session invalide — recommencez la connexion.")

    user = db.get(Utilisateur, otp.user_id)
    if not user or not user.actif:
        raise HTTPException(400, "Compte introuvable ou désactivé.")

    try:
        code = create_admin_code(db, user.id)
        send_admin_code_email(user.nom_complet or user.username, user.username, code)
    except RuntimeError as e:
        raise HTTPException(503, str(e))

    log_event(db, ADMIN_CODE_REQUESTED, user_id=user.id, ip_address=ip)

    # Prolonger le cookie à 60 min pour que l'employé ait le temps de contacter l'admin
    response.set_cookie(
        OTP_PENDING_COOKIE, pending_token,
        httponly=True, samesite="lax", max_age=3600, path="/",
    )
    return {
        "admin_code_sent": True,
        "admin_email_hint": _mask_email(EMAIL_USER) if EMAIL_USER else "l'administrateur",
    }


class AdminCodeVerifyIn(BaseModel):
    code: str


@app.post("/api/otp/verify-admin-code")
def otp_verify_admin_code(
    data: AdminCodeVerifyIn,
    request: Request,
    response: Response,
    db: Session = Depends(get_db),
):
    """Vérifie le code admin 5 chiffres et ouvre la session si valide."""
    ip            = request.client.host if request.client else None
    ua            = request.headers.get("user-agent", "")
    pending_token = request.cookies.get(OTP_PENDING_COOKIE, "")

    try:
        user = verify_admin_code(db, pending_token, data.code)
    except ValueError as e:
        log_event(db, ADMIN_CODE_FAILED, ip_address=ip, details={"raison": str(e)})
        raise HTTPException(401, str(e))

    log_event(db, ADMIN_CODE_VERIFIED, user_id=user.id, ip_address=ip)
    response.delete_cookie(OTP_PENDING_COOKIE, path="/")

    token = create_session(db, user.id, ip_address=ip, user_agent=ua)
    log_event(db, LOGIN_SUCCESS, user_id=user.id, ip_address=ip)
    response.set_cookie(
        SESSION_COOKIE, token,
        httponly=True, samesite="lax", max_age=7 * 24 * 3600, path="/",
    )
    raw_key = make_api_key(db, user.id)
    return {
        "id": user.id, "username": user.username,
        "nom_complet": user.nom_complet, "role": user.role,
        "api_key": raw_key,
    }


@app.post("/api/logout")
def logout(request: Request, response: Response, db: Session = Depends(get_db)):
    ip    = request.client.host if request.client else None
    token = request.cookies.get(SESSION_COOKIE)
    user  = get_session_user(db, token) if token else None
    delete_session(db, token)
    if user:
        log_event(db, LOGOUT, user_id=user.id, ip_address=ip)
    response.delete_cookie(SESSION_COOKIE, path="/")
    return {"ok": True}


class LoginMetaIn(BaseModel):
    photo_b64: Optional[str] = None
    latitude:  Optional[float] = None
    longitude: Optional[float] = None


def _ip_geoloc(ip: str | None):
    """Retourne (lat, lng) approximatif via ip-api.com ou (None, None) si indisponible."""
    if not ip or ip in ("127.0.0.1", "::1", "localhost"):
        return None, None
    # Ignorer les IP privées
    import ipaddress as _ip
    try:
        obj = _ip.ip_address(ip)
        if obj.is_private or obj.is_loopback:
            return None, None
    except ValueError:
        return None, None
    try:
        import requests as _req
        r = _req.get(f"http://ip-api.com/json/{ip}?fields=status,lat,lon",
                     timeout=2, headers={"User-Agent": "tracking-carburant/1.0"})
        d = r.json()
        if d.get("status") == "success":
            return float(d["lat"]), float(d["lon"])
    except Exception:
        pass
    return None, None


@app.post("/api/auth/login-meta")
def login_meta_save(data: LoginMetaIn, request: Request, db: Session = Depends(get_db)):
    """Enregistre la photo et la géolocalisation de connexion."""
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(401, "Non authentifié")
    token   = request.cookies.get(SESSION_COOKIE)
    session = db.query(SessionToken).filter_by(token=token).first() if token else None
    ip      = request.client.host if request.client else None

    lat, lng = data.latitude, data.longitude
    # Fallback : géolocalisation par IP si le navigateur n'a pas transmis de coordonnées
    if lat is None or lng is None:
        lat, lng = _ip_geoloc(ip)

    event = LoginSecurityEvent(
        user_id    = user.id,
        session_id = session.id if session else None,
        photo_b64  = data.photo_b64,
        latitude   = lat,
        longitude  = lng,
        ip_address = ip,
        user_agent = request.headers.get("user-agent", "")[:255],
    )
    db.add(event)
    db.commit()
    return {"ok": True}


@app.get("/api/me")
def me(request: Request, db: Session = Depends(get_db)):
    from models import Employe as _Employe
    state_user = request.state.user
    user = db.get(Utilisateur, state_user.id)
    if not user:
        raise HTTPException(401, "Session invalide")
    perms = None
    role_nom = user.poste
    if user.role_obj:
        perms = user.role_obj.permissions
        role_nom = user.role_obj.nom
    emp = db.query(_Employe).filter_by(utilisateur_id=user.id).first()

    # Auto-créer un enregistrement Employe pour les comptes caissière qui n'en ont pas.
    # Cela permet au système POS de filtrer leurs ventes par caissier_id.
    if not emp and (role_nom or user.poste or "").lower().startswith("caissier"):
        from datetime import date as _date
        parts = (user.nom_complet or user.username).split(" ", 1)
        emp = _Employe(
            prenom=parts[0],
            nom=parts[1] if len(parts) > 1 else parts[0],
            poste="Caissier",
            date_embauche=_date.today(),
            salaire_base=0,
            type_contrat="CDI",
            actif=True,
            utilisateur_id=user.id,
        )
        db.add(emp)
        db.commit()
        db.refresh(emp)

    return {
        "id":          user.id,
        "username":    user.username,
        "nom_complet": user.nom_complet,
        "email":       user.email,
        "role":        user.role,
        "poste":       user.poste,
        "role_id":     user.role_id,
        "role_nom":    role_nom,
        "permissions": perms,
        "est_admin":   user.role == "admin" or bool(perms and perms.get("admin", False)),
        "employe_id":  emp.id if emp else None,
        "employe_nom": (emp.nom + " " + emp.prenom) if emp else None,
    }


# ---------- Gestion des clés API ----------

@app.post("/api/auth/api-key")
def rotate_api_key(request: Request, db: Session = Depends(get_db)):
    """Génère une nouvelle clé API pour l'utilisateur connecté (révoque l'ancienne)."""
    user = request.state.user
    raw_key = make_api_key(db, user.id)
    return {"api_key": raw_key}


@app.delete("/api/auth/api-key")
def revoke_api_key_endpoint(request: Request, db: Session = Depends(get_db)):
    """Révoque la clé API de l'utilisateur connecté."""
    user = request.state.user
    revoke_api_key(db, user.id)
    return {"ok": True}


@app.get("/api/auth/me")
def auth_me(request: Request):
    """Retourne les infos de l'utilisateur authentifié (session ou clé API)."""
    user = request.state.user
    return {
        "id": user.id, "username": user.username,
        "nom_complet": user.nom_complet, "role": user.role,
        "has_api_key": bool(user.api_key_hash),
    }


class ChangePasswordIn(BaseModel):
    ancien_mot_de_passe: str
    nouveau_mot_de_passe: str


@app.post("/api/auth/change-password")
def change_password(data: ChangePasswordIn, request: Request, db: Session = Depends(get_db)):
    user = db.get(Utilisateur, request.state.user.id)
    # Comptes OAuth : le hash sentinel ne contient pas ":" → verify_password retourne False
    if user.oauth_provider:
        raise HTTPException(400,
            f"Ce compte est lié à {user.oauth_provider.capitalize()} — "
            "la connexion par mot de passe n'est pas activée pour ce compte.")
    if not verify_password(data.ancien_mot_de_passe, user.password_hash):
        raise HTTPException(400, "Mot de passe actuel incorrect")
    if len(data.nouveau_mot_de_passe) < 6:
        raise HTTPException(400, "Le nouveau mot de passe doit contenir au moins 6 caractères")
    user.password_hash = hash_password(data.nouveau_mot_de_passe)
    db.commit()
    return {"ok": True}


# ── Récupération de mot de passe (endpoints publics) ─────────────────────────

class ForgotPasswordIn(BaseModel):
    identifier: str   # email ou username


@app.post("/api/auth/forgot-password")
async def forgot_password(data: ForgotPasswordIn, request: Request, db: Session = Depends(get_db)):
    """
    Envoie un email de réinitialisation.
    Retourne toujours la même réponse (anti-énumération).
    """
    ip       = request.client.host if request.client else None
    base_url = str(request.base_url).rstrip("/")
    try:
        request_reset(data.identifier, db, ip, base_url)
    except ValueError as e:
        # Rate limit uniquement — on propage le message
        raise HTTPException(429, str(e))
    except RuntimeError as e:
        raise HTTPException(503, str(e))
    return {"ok": True, "message": "Si cet identifiant existe, un email a été envoyé."}


@app.get("/api/auth/reset-password/verify")
def verify_reset(token: str, db: Session = Depends(get_db)):
    """Vérifie un token sans le consommer (pré-validation à l'ouverture du formulaire)."""
    try:
        user = verify_reset_token(token, db)
        return {"ok": True, "username": user.username}
    except ValueError as e:
        raise HTTPException(400, str(e))


class ResetPasswordIn(BaseModel):
    token:        str
    new_password: str


@app.post("/api/auth/reset-password")
def reset_password(data: ResetPasswordIn, db: Session = Depends(get_db)):
    """Consomme le token et change le mot de passe. Révoque toutes les sessions."""
    try:
        consume_reset_token(data.token, data.new_password, db)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "message": "Mot de passe modifié. Toutes les sessions ont été révoquées."}


# ══════════════════════════════════════════════════════════════════════════════
# Dépendance : accès réservé aux administrateurs
# ══════════════════════════════════════════════════════════════════════════════

def require_admin(request: Request, db: Session = Depends(get_db)) -> Utilisateur:
    """Autorise si l'utilisateur a le rôle 'admin' ou permissions.admin == True."""
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(403, "Accès réservé aux administrateurs")
    if user.role == "admin":
        return user
    if user.role_id:
        u = db.get(Utilisateur, user.id)
        if u and u.role_obj and u.role_obj.permissions.get("admin", False):
            return u
    raise HTTPException(403, "Accès réservé aux administrateurs")


def require_pdg(request: Request) -> Utilisateur:
    """Autorise uniquement le rôle 'pdg'. Lève 403 sinon."""
    user = getattr(request.state, "user", None)
    if not user or user.role != "pdg":
        raise HTTPException(403, "Accès réservé au PDG")
    return user


# ══════════════════════════════════════════════════════════════════════════════
# Gestion des comptes employés (admin seulement)
# ══════════════════════════════════════════════════════════════════════════════

_USERNAME_RE = __import__("re").compile(r"^[a-zA-Z0-9._\-]{3,60}$")
_EMAIL_RE    = __import__("re").compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


class CreateUtilisateurIn(BaseModel):
    username:    str
    nom_complet: str
    password:    str
    code_acces:  str             # exactement 9 chiffres
    role:        str = "operateur"
    email:       str             # obligatoire — sert à la connexion
    poste:       Optional[str] = None   # poste de l'employé → contrôle d'accès UI


def _user_public(u: Utilisateur) -> dict:
    """Sérialise un Utilisateur sans exposer api_key_hash ni password_hash."""
    return {
        "id":            u.id,
        "username":      u.username,
        "nom_complet":   u.nom_complet,
        "role":          u.role,
        "poste":         u.poste,
        "actif":         u.actif,
        "has_api_key":   bool(u.api_key_hash),
        "email":         u.email,
        "oauth_provider":u.oauth_provider,
        "created_at":    u.created_at.isoformat() if u.created_at else None,
    }


def _count_active_admins(db: Session) -> int:
    return db.query(Utilisateur).filter_by(role="admin", actif=True).count()


@app.post("/api/auth/utilisateurs")
def creer_utilisateur(
    data: CreateUtilisateurIn,
    _admin: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Crée un compte employé. La clé API générée est retournée une seule fois."""
    if data.role not in ("admin", "operateur", "pdg"):
        raise HTTPException(400, "Rôle invalide — choisir 'admin', 'operateur' ou 'pdg'")
    if len(data.password) < 6:
        raise HTTPException(400, "Le mot de passe doit contenir au moins 6 caractères")
    if not _CODE_RE.match(data.code_acces):
        raise HTTPException(400, "Le code d'accès doit contenir exactement 9 chiffres")
    username = data.username.strip()
    if not username:
        raise HTTPException(400, "Le nom d'utilisateur est requis")
    if not _USERNAME_RE.match(username):
        raise HTTPException(400,
            "Identifiant invalide — 3 à 60 caractères, lettres/chiffres/points/tirets/underscores uniquement, pas d'espaces")
    if not data.nom_complet.strip():
        raise HTTPException(400, "Le nom complet est requis")
    email_clean = data.email.strip().lower()
    if not _EMAIL_RE.match(email_clean):
        raise HTTPException(400, f"Adresse email invalide : {data.email}")
    if db.query(Utilisateur).filter_by(username=username).first():
        raise HTTPException(409, f"L'identifiant '{username}' est déjà utilisé")
    if db.query(Utilisateur).filter_by(email=email_clean).first():
        raise HTTPException(409, "Cet email est déjà associé à un compte")

    u = Utilisateur(
        username=username,
        password_hash=hash_password(data.password),
        code_acces_hash=hash_code_acces(data.code_acces),
        nom_complet=data.nom_complet.strip(),
        role=data.role,
        poste=data.poste.strip() if data.poste else None,
        email=email_clean,
    )
    db.add(u)
    db.commit()
    db.refresh(u)
    raw_key = make_api_key(db, u.id)
    return {**_user_public(u), "api_key": raw_key}


@app.get("/api/auth/utilisateurs")
def lister_utilisateurs(
    _admin: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Liste tous les comptes. Ne retourne jamais api_key_hash."""
    return [_user_public(u) for u in db.query(Utilisateur).order_by(Utilisateur.id).all()]


@app.post("/api/auth/utilisateurs/{uid}/revoquer")
def revoquer_utilisateur(
    uid: int,
    _admin: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Désactive un compte et révoque sa clé API."""
    u = db.get(Utilisateur, uid)
    if not u:
        raise HTTPException(404, "Utilisateur introuvable")
    if u.role == "admin" and _count_active_admins(db) <= 1:
        raise HTTPException(409, "Impossible : dernier administrateur actif")
    u.actif = False
    revoke_api_key(db, u.id)
    db.commit()
    return _user_public(u)


@app.post("/api/auth/utilisateurs/{uid}/reactiver")
def reactiver_utilisateur(
    uid: int,
    _admin: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Réactive un compte désactivé."""
    u = db.get(Utilisateur, uid)
    if not u:
        raise HTTPException(404, "Utilisateur introuvable")
    u.actif = True
    db.commit()
    return _user_public(u)


@app.post("/api/auth/utilisateurs/{uid}/regenerer")
def regenerer_cle(
    uid: int,
    _admin: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Génère une nouvelle clé API pour l'employé. L'ancienne est invalidée immédiatement."""
    u = db.get(Utilisateur, uid)
    if not u:
        raise HTTPException(404, "Utilisateur introuvable")
    raw_key = make_api_key(db, u.id)
    return {**_user_public(u), "api_key": raw_key}


@app.delete("/api/auth/utilisateurs/{uid}")
def supprimer_utilisateur(
    uid: int,
    _admin: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Supprime définitivement un compte. Privilégier la désactivation."""
    u = db.get(Utilisateur, uid)
    if not u:
        raise HTTPException(404, "Utilisateur introuvable")
    if u.role == "admin" and _count_active_admins(db) <= 1:
        raise HTTPException(409, "Impossible : dernier administrateur actif")
    db.delete(u)
    db.commit()
    return {"ok": True, "id": uid}


# ══════════════════════════════════════════════════════════════════════════════
# OAuth 2.0 — Authorization Code Flow (Google + architecture multi-fournisseurs)
# ══════════════════════════════════════════════════════════════════════════════
#
# Politique de création de compte :
#   Un email inconnu de la base est REFUSÉ. L'admin doit créer le compte
#   (via POST /api/auth/utilisateurs) avec l'email du collaborateur AVANT
#   que celui-ci tente de se connecter via OAuth. Cette politique évite
#   qu'un email Google aléatoire accède à l'application.
#
# État anti-CSRF :
#   Dictionnaire en mémoire {state → {provider, expires_at}}.
#   TTL = 10 minutes. Nettoyé à chaque vérification.
#
# Session après OAuth :
#   Le même système de cookie session (create_session) est utilisé — aucune
#   modification du middleware AuthMiddleware nécessaire.
#
# Dépendances : google-auth (déjà installé) + httpx (déjà installé)
# ══════════════════════════════════════════════════════════════════════════════

import secrets as _secrets
import time    as _time
from urllib.parse import urlencode as _urlencode

_OAUTH_STATES: dict = {}   # {state_str: {"provider": str, "exp": float}}
_STATE_TTL_S  = 600        # 10 minutes

_OAUTH_CFG = {
    "google": {
        "auth_url":    "https://accounts.google.com/o/oauth2/v2/auth",
        "token_url":   "https://oauth2.googleapis.com/token",
        "userinfo_url":"https://www.googleapis.com/oauth2/v3/userinfo",
        "scope":       "openid email profile",
        "client_id":     lambda: os.getenv("GOOGLE_CLIENT_ID",     ""),
        "client_secret": lambda: os.getenv("GOOGLE_CLIENT_SECRET", ""),
    },
    "microsoft": {
        "auth_url":    "https://login.microsoftonline.com/common/oauth2/v2.0/authorize",
        "token_url":   "https://login.microsoftonline.com/common/oauth2/v2.0/token",
        "userinfo_url":"https://graph.microsoft.com/oidc/userinfo",
        "scope":       "openid email profile",
        "client_id":     lambda: os.getenv("MICROSOFT_CLIENT_ID",     ""),
        "client_secret": lambda: os.getenv("MICROSOFT_CLIENT_SECRET", ""),
    },
}


def _oauth_callback_url(request: Request, provider: str) -> str:
    base = str(request.base_url).rstrip("/")
    return f"{base}/api/auth/oauth/{provider}/callback"


def _clean_states():
    now = _time.time()
    expired = [k for k, v in _OAUTH_STATES.items() if v["exp"] < now]
    for k in expired:
        _OAUTH_STATES.pop(k, None)


@app.get("/api/auth/oauth/providers", include_in_schema=True)
def oauth_providers():
    """
    Retourne les fournisseurs OAuth actifs (client_id défini).
    Endpoint public — ne divulgue aucun secret.
    """
    return {p: bool(cfg["client_id"]()) for p, cfg in _OAUTH_CFG.items()}


@app.get("/api/auth/oauth/{provider}/login", include_in_schema=True)
def oauth_login(provider: str, request: Request):
    """
    Démarre le flux OAuth : redirige vers la page de consentement du fournisseur.
    Un `state` anti-CSRF est généré et stocké côté serveur.
    """
    if provider not in _OAUTH_CFG:
        return RedirectResponse(url="/login?oauth_error=unknown_provider")
    cfg = _OAUTH_CFG[provider]
    client_id = cfg["client_id"]()
    if not client_id:
        # Redirige vers login avec message d'erreur plutôt que 503 brut
        return RedirectResponse(url="/login?oauth_error=oauth_not_configured")

    _clean_states()
    state = _secrets.token_urlsafe(32)
    _OAUTH_STATES[state] = {"provider": provider, "exp": _time.time() + _STATE_TTL_S}

    params = {
        "client_id":     client_id,
        "redirect_uri":  _oauth_callback_url(request, provider),
        "response_type": "code",
        "scope":         cfg["scope"],
        "state":         state,
        "prompt":        "select_account",
    }
    return RedirectResponse(url=f"{cfg['auth_url']}?{_urlencode(params)}")


@app.get("/api/auth/oauth/{provider}/callback", include_in_schema=True)
def oauth_callback(
    provider: str,
    request: Request,
    response: Response,
    db: Session = Depends(get_db),
    code:  Optional[str] = None,
    state: Optional[str] = None,
    error: Optional[str] = None,
):
    """
    Callback OAuth :
    1. Vérifie le state anti-CSRF
    2. Échange le code contre un token
    3. Récupère le profil (email, sub, nom)
    4. Rapproche avec un compte existant (email connu requis)
    5. Crée une session et redirige vers l'interface
    """
    import httpx

    # Erreur explicite du fournisseur
    if error:
        return RedirectResponse(url=f"/login?oauth_error={error}")

    # Validation state anti-CSRF
    _clean_states()
    stored = _OAUTH_STATES.pop(state or "", None)
    if not stored or stored.get("provider") != provider:
        return RedirectResponse(url="/login?oauth_error=invalid_state")

    if provider not in _OAUTH_CFG:
        return RedirectResponse(url="/login?oauth_error=unknown_provider")
    cfg = _OAUTH_CFG[provider]

    client_id     = cfg["client_id"]()
    client_secret = cfg["client_secret"]()
    redirect_uri  = _oauth_callback_url(request, provider)

    # Échange du code → access_token + id_token
    try:
        tok_resp = httpx.post(
            cfg["token_url"],
            data={
                "grant_type":    "authorization_code",
                "code":          code,
                "redirect_uri":  redirect_uri,
                "client_id":     client_id,
                "client_secret": client_secret,
            },
            timeout=10,
        )
        tok_resp.raise_for_status()
        tokens = tok_resp.json()
    except Exception:
        return RedirectResponse(url="/login?oauth_error=token_exchange_failed")

    access_token = tokens.get("access_token")
    if not access_token:
        return RedirectResponse(url="/login?oauth_error=no_access_token")

    # Récupération du profil utilisateur
    try:
        info_resp = httpx.get(
            cfg["userinfo_url"],
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=10,
        )
        info_resp.raise_for_status()
        profile = info_resp.json()
    except Exception:
        return RedirectResponse(url="/login?oauth_error=userinfo_failed")

    email    = (profile.get("email") or "").lower().strip()
    sub      = profile.get("sub") or profile.get("id", "")
    name     = profile.get("name") or profile.get("displayName") or ""
    verified = profile.get("email_verified", True)  # Microsoft ne renvoie pas ce champ

    if not email or not verified:
        return RedirectResponse(url="/login?oauth_error=email_not_verified")

    # Rapprochement : cherche d'abord par oauth_sub, puis par email
    user = db.query(Utilisateur).filter_by(oauth_sub=sub, oauth_provider=provider).first()
    if not user:
        user = db.query(Utilisateur).filter_by(email=email).first()
        if not user:
            # Compte inconnu → refus (politique de sécurité)
            return RedirectResponse(url="/login?oauth_error=account_not_found")
        # Première connexion OAuth : lie l'identité externe au compte existant
        user.oauth_provider = provider
        user.oauth_sub      = sub
        if not user.nom_complet and name:
            user.nom_complet = name
        db.commit()

    if not user.actif:
        return RedirectResponse(url="/login?oauth_error=account_disabled")

    # ── Vérification en deux étapes (OTP) après OAuth ────────────────
    if OTP_ENABLED and user.email:
        try:
            code, pending_token = create_otp(db, user.id)
            send_otp_email(user.nom_complet or user.email, user.email, code)
        except (ValueError, RuntimeError) as exc:
            import traceback; traceback.print_exc()
            log.warning("OTP impossible pour %s via OAuth : %s", _mask_email(user.email), exc)
            return RedirectResponse(url=f"/login?oauth_error=otp_failed", status_code=302)
        except Exception as exc:
            import traceback; traceback.print_exc()
            log.error("Exception inattendue OTP OAuth : %s", exc)
            return RedirectResponse(url=f"/login?oauth_error=otp_failed", status_code=302)
        log_event(db, OTP_SENT, user_id=user.id,
                  ip_address=request.client.host if request.client else None)
        hint = url_quote(_mask_email(user.email), safe="")
        redir = RedirectResponse(url=f"/login?otp=1&hint={hint}", status_code=302)
        redir.set_cookie(
            OTP_PENDING_COOKIE, pending_token,
            httponly=True, samesite="lax", max_age=OTP_PENDING_MAX_AGE, path="/",
        )
        return redir

    # ── Connexion directe (OTP désactivé) ────────────────────────────
    ip = request.client.host if request.client else None
    ua = request.headers.get("user-agent", "")
    session_token = create_session(db, user.id, ip_address=ip, user_agent=ua)
    log_event(db, LOGIN_SUCCESS, user_id=user.id, ip_address=ip)
    redir = RedirectResponse(url="/", status_code=302)
    redir.set_cookie(
        SESSION_COOKIE, session_token,
        httponly=True, samesite="lax", max_age=7 * 24 * 3600, path="/",
    )
    return redir


@app.get("/login", include_in_schema=False)
def serve_login():
    html_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "frontend", "login.html"
    )
    return FileResponse(html_path, media_type="text/html")


@app.get("/api/audit/pdf", include_in_schema=False)
def serve_audit_pdf():
    path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "frontend",
        "audit_station_suivi_meters_2026-06-17.pdf",
    )
    return FileResponse(path, media_type="application/pdf",
                        filename="audit_station_suivi_meters_2026-06-17.pdf")


@app.get("/api/audit/xlsx", include_in_schema=False)
def serve_audit_xlsx():
    path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "frontend",
        "audit_station_suivi_meters_2026-06-17.xlsx",
    )
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="audit_station_suivi_meters_2026-06-17.xlsx",
    )


# ---------- Schemas ----------
class PompeIn(BaseModel):
    nom: str


class ProduitIn(BaseModel):
    nom: str
    prix_gallon: float = 0


class ReleveIn(BaseModel):
    date: date_type
    periode: str
    pompe_id: int
    prix_gallon: float
    metter_avant: float
    metter_apres: float


# ---------- Produits & Pompes ----------
@app.get("/api/produits")
def list_produits(db: Session = Depends(get_db)):
    # Bug 10 fix : ne retourner que les produits et pompes actifs
    out = []
    for p in db.query(Produit).filter_by(actif=True).all():
        out.append({
            "id": p.id, "nom": p.nom, "prix_gallon": p.prix_gallon,
            "pompes": [{"id": q.id, "nom": q.nom} for q in p.pompes if q.actif],
        })
    return out


@app.post("/api/produits")
def create_produit(data: ProduitIn, db: Session = Depends(get_db)):
    if db.query(Produit).filter_by(nom=data.nom).first():
        raise HTTPException(400, "Produit existe deja")
    p = Produit(nom=data.nom, prix_gallon=data.prix_gallon)
    db.add(p); db.commit(); db.refresh(p)
    return {"id": p.id, "nom": p.nom, "prix_gallon": p.prix_gallon}


@app.post("/api/produits/{produit_id}/pompes")
def add_pompe(produit_id: int, data: PompeIn, db: Session = Depends(get_db)):
    if not db.query(Produit).get(produit_id):
        raise HTTPException(404, "Produit introuvable")
    pompe = Pompe(produit_id=produit_id, nom=data.nom)
    db.add(pompe); db.commit(); db.refresh(pompe)
    return {"id": pompe.id, "nom": pompe.nom}


@app.delete("/api/pompes/{pompe_id}")
def delete_pompe(pompe_id: int, db: Session = Depends(get_db)):
    pompe = db.query(Pompe).get(pompe_id)
    if not pompe:
        raise HTTPException(404, "Pompe introuvable")
    db.delete(pompe); db.commit()
    return {"ok": True}


# ---------- Releves ----------
@app.post("/api/releves")
def upsert_releve(data: ReleveIn, db: Session = Depends(get_db)):
    # Bug 7 fix : validations métier à la frontière API (message clair avant la DB)
    if data.periode not in PERIODES:
        raise HTTPException(400, f"Periode invalide — valeurs acceptées : {PERIODES}")
    if data.prix_gallon < 0:
        raise HTTPException(400, "prix_gallon doit être ≥ 0")
    if data.metter_avant < 0:
        raise HTTPException(400, "metter_avant doit être ≥ 0")
    if data.metter_apres < data.metter_avant:
        raise HTTPException(
            400,
            f"metter_apres ({data.metter_apres:.3f}) doit être ≥ metter_avant "
            f"({data.metter_avant:.3f}) — un compteur ne peut pas reculer."
        )

    r = (db.query(Releve)
         .filter_by(date=data.date, periode=data.periode, pompe_id=data.pompe_id)
         .first())
    if not r:
        r = Releve(date=data.date, periode=data.periode, pompe_id=data.pompe_id,
                   nb_modifications=0)
        db.add(r)
    else:
        # Bug 9 fix : utilisation de la constante module
        if r.nb_modifications >= MAX_MODIFICATIONS_PAR_RELEVE:
            raise HTTPException(
                403,
                f"Limite atteinte : ce relevé a déjà été modifié "
                f"{MAX_MODIFICATIONS_PAR_RELEVE} fois."
            )
        r.nb_modifications += 1
    r.prix_gallon  = data.prix_gallon
    r.metter_avant = data.metter_avant
    r.metter_apres = data.metter_apres
    db.commit(); db.refresh(r)
    return _releve_dict(r)


@app.get("/api/releves")
def get_releves(date: date_type, periode: Optional[str] = None,
                produit_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(Releve).filter(Releve.date == date)
    if periode:
        q = q.filter(Releve.periode == periode)
    rows = q.all()
    if produit_id:
        rows = [r for r in rows if r.pompe.produit_id == produit_id]
    return [_releve_dict(r) for r in rows]


def _releve_dict(r: Releve):
    return {
        "id": r.id, "date": str(r.date), "periode": r.periode,
        "pompe_id": r.pompe_id, "pompe_nom": r.pompe.nom,
        "produit_id": r.pompe.produit_id, "produit_nom": r.pompe.produit.nom,
        "prix_gallon": r.prix_gallon,
        "metter_avant": r.metter_avant, "metter_apres": r.metter_apres,
        "quantite": r.quantite, "montant_vente": r.montant_vente,
        "nb_modifications": r.nb_modifications,
    }


# ---------- Rapport / Dashboard ----------
@app.get("/api/rapport")
def rapport(date: date_type, db: Session = Depends(get_db)):
    """Synthese d'une journee : par produit, par periode."""
    produits = db.query(Produit).all()
    result = {"date": str(date), "produits": [], "total_cash": 0}
    for p in produits:
        pompe_ids = [q.id for q in p.pompes]
        bloc = {"produit_id": p.id, "produit_nom": p.nom, "periodes": []}
        prod_cash = 0
        for per in PERIODES:
            releves = (db.query(Releve)
                       .filter(Releve.date == date, Releve.periode == per,
                               Releve.pompe_id.in_(pompe_ids)).all())
            cash = sum(r.montant_vente for r in releves)
            prod_cash += cash
            bloc["periodes"].append({
                "periode": per,
                "releves": [_releve_dict(r) for r in releves],
                "total_cash": round(cash, 2),
            })
            result["total_cash"] += cash
        bloc["total_cash_produit"] = round(prod_cash, 2)
        result["produits"].append(bloc)
    result["total_cash"] = round(result["total_cash"], 2)
    return result


# ---------- Stats (source de verite pour le chatbot) ----------
@app.get("/api/stats")
def stats_endpoint(date_debut: date_type, date_fin: date_type,
                   produit_id: Optional[int] = None,
                   pompe_id: Optional[int] = None,
                   periode: Optional[str] = None,
                   db: Session = Depends(get_db)):
    from stats import compute_stats
    return compute_stats(db, date_debut, date_fin, produit_id, periode, pompe_id)


# ---------- Chatbot de rapports ----------
class ChatIn(BaseModel):
    message: str
    historique: list = []


@app.post("/api/chat")
def chat_endpoint(data: ChatIn, db: Session = Depends(get_db)):
    from chatbot import chat
    return chat(db, data.message, data.historique)


# ---------- Detection d'anomalies dans la suite des meters ----------
# PERIODE_ORDRE supprimé — utiliser PERIODE_RANG défini plus bas (identique).
# Seuil du saut anormal : une quantite > SEUIL_SAUT x la moyenne de la pompe
# est signalee comme avertissement (pas comme erreur bloquante).
SEUIL_SAUT = 5
# Bug 12 fix : n'activer la détection de saut qu'après un minimum de données.
# Avec 1-2 relevés, la moyenne est peu fiable et génère des faux positifs.
SEUIL_MIN_RELEVES_POUR_SAUT = 5


@app.get("/api/anomalies")
def anomalies(date: date_type, db: Session = Depends(get_db)):
    """
    Moteur d'anomalies unifié : analyse les compteurs ET la cohérence stock.

    Anomalies compteurs (inchangées) :
      QUANTITE_NEGATIVE, REGRESSION_METER, SAUT_ANORMAL

    Anomalies stock (nouvelles) :
      VENTE_SANS_STOCK, STOCK_NEGATIF, PRIX_MANQUANT, DECALAGE_STOCK

    Corrélation : SAUT_ANORMAL + DECALAGE_STOCK le même jour/produit
    → champ incident_lie ajouté aux deux anomalies.
    """
    # ── 1. Relevés compteurs ──────────────────────────────────────────
    releves = db.query(Releve).filter(Releve.date <= date).all()

    # Précharger les pompes pour éviter le lazy loading
    pompes_cache: dict[int, Pompe] = {}
    for r in releves:
        if r.pompe_id not in pompes_cache:
            pompes_cache[r.pompe_id] = r.pompe

    par_pompe: dict[int, list] = {}
    for r in releves:
        par_pompe.setdefault(r.pompe_id, []).append(r)

    anom_compteurs = []

    for pompe_id, liste in par_pompe.items():
        pompe_obj = pompes_cache[pompe_id]
        liste.sort(key=lambda r: (r.date, PERIODE_RANG.get(r.periode, 9)))

        quantites  = [r.quantite for r in liste if r.quantite > 0]
        moyenne    = sum(quantites) / len(quantites) if quantites else 0
        saut_actif = len(quantites) >= SEUIL_MIN_RELEVES_POUR_SAUT

        meter_apres_precedent = None
        releve_courant_valide = True

        for r in liste:
            nom = pompe_obj.nom

            if r.metter_apres < r.metter_avant:
                releve_courant_valide = False
                anom_compteurs.append({
                    "type":               "QUANTITE_NEGATIVE",
                    "gravite":            "erreur",
                    "pompe_nom":          nom,
                    "produit_id":         pompe_obj.produit_id,
                    "date":               str(r.date),
                    "periode":            r.periode,
                    "valeur_attendue_min": float(r.metter_avant),
                    "valeur_saisie":      float(r.metter_apres),
                    "message": (
                        f"Le meter apres ({r.metter_apres}) est inferieur au "
                        f"meter avant ({r.metter_avant}). La quantite vendue "
                        f"serait negative, ce qui est impossible."
                    ),
                })
            else:
                releve_courant_valide = True

            if (meter_apres_precedent is not None
                    and r.metter_avant < meter_apres_precedent):
                anom_compteurs.append({
                    "type":               "REGRESSION_METER",
                    "gravite":            "erreur",
                    "pompe_nom":          nom,
                    "produit_id":         pompe_obj.produit_id,
                    "date":               str(r.date),
                    "periode":            r.periode,
                    "valeur_attendue_min": round(meter_apres_precedent, 3),
                    "valeur_saisie":      float(r.metter_avant),
                    "message": (
                        f"Le meter avant ({r.metter_avant}) est inferieur au "
                        f"meter apres precedent ({round(meter_apres_precedent, 3)}). "
                        f"Le compteur ne peut pas reculer."
                    ),
                })

            if saut_actif and moyenne > 0 and r.quantite > SEUIL_SAUT * moyenne:
                anom_compteurs.append({
                    "type":          "SAUT_ANORMAL",
                    "gravite":       "avertissement",
                    "pompe_nom":     nom,
                    "produit_id":    pompe_obj.produit_id,
                    "date":          str(r.date),
                    "periode":       r.periode,
                    "seuil_utilise": round(SEUIL_SAUT * moyenne, 3),
                    "valeur_saisie": round(r.quantite, 3),
                    "message": (
                        f"Quantite vendue ({round(r.quantite, 3)} gal) anormalement "
                        f"elevee par rapport a la moyenne de cette pompe "
                        f"({round(moyenne, 3)} gal). A verifier."
                    ),
                })

            if releve_courant_valide:
                meter_apres_precedent = r.metter_apres

    # ── 2. Anomalies stock ────────────────────────────────────────────
    anom_stk = anomalies_stock(db, date)

    # ── 3. Corrélation SAUT_ANORMAL ↔ DECALAGE_STOCK ─────────────────
    anom_compteurs, anom_stk = corr_saut_decalage(anom_compteurs, anom_stk)

    # ── 4. Anomalies bar/restaurant ───────────────────────────────────
    try:
        from pos_service import detecter_anomalies_bar
        anom_bar = detecter_anomalies_bar(date, db)
    except Exception:
        anom_bar = []   # module bar non initialisé — ne pas bloquer les anomalies carburant

    # ── 5. Fusion et tri ──────────────────────────────────────────────
    # Erreurs avant avertissements, puis par date.
    gravite_ordre = {"erreur": 0, "avertissement": 1}
    toutes = anom_compteurs + anom_stk + anom_bar
    toutes.sort(key=lambda a: (gravite_ordre.get(a["gravite"], 2), a.get("date", "")))

    return {
        "date":           str(date),
        "nb_anomalies":   len(toutes),
        "nb_compteurs":   len(anom_compteurs),
        "nb_stock":       len(anom_stk),
        "nb_bar":         len(anom_bar),
        "anomalies":      toutes,
    }


# ---------- Série temporelle (7 jours) ----------
@app.get("/api/serie")
def serie_endpoint(
    date_fin: Optional[str] = None,
    jours: int = 7,
    produit_id: Optional[int] = None,
    periode: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    Bug 8 fix : remplace N appels compute_stats() par 2 requêtes SQL uniques.
    Pour 90 jours : 180 requêtes → 2 requêtes.
    """
    from datetime import timedelta, date as dt, datetime
    from collections import defaultdict

    if jours < 1:
        raise HTTPException(400, "jours doit être ≥ 1")

    d_fin = (
        datetime.strptime(date_fin, "%Y-%m-%d").date()
        if date_fin
        else dt.today()
    )
    d_debut = d_fin - timedelta(days=jours - 1)

    # ── Période courante — une seule requête ─────────────────────────
    q = db.query(Releve).filter(Releve.date >= d_debut, Releve.date <= d_fin)
    if periode:
        q = q.filter(Releve.periode == periode)
    releves = q.all()
    if produit_id:
        releves = [r for r in releves if r.pompe.produit_id == produit_id]

    par_date: dict = defaultdict(lambda: {"total_montant": 0.0, "total_quantite": 0.0})
    for r in releves:
        ds = str(r.date)
        par_date[ds]["total_montant"]  += r.montant_vente
        par_date[ds]["total_quantite"] += r.quantite

    result_jours = []
    for i in range(jours):
        d  = d_debut + timedelta(days=i)
        ds = d.strftime("%Y-%m-%d")
        day = par_date.get(ds, {"total_montant": 0.0, "total_quantite": 0.0})
        result_jours.append({
            "date":           ds,
            "total_montant":  round(day["total_montant"],  2),
            "total_quantite": round(day["total_quantite"], 3),
        })

    total_periode = round(sum(j["total_montant"] for j in result_jours), 2)

    # ── Période précédente — une seule requête ───────────────────────
    prev_fin   = d_debut - timedelta(days=1)
    prev_debut = prev_fin - timedelta(days=jours - 1)
    pq = db.query(Releve).filter(Releve.date >= prev_debut, Releve.date <= prev_fin)
    if periode:
        pq = pq.filter(Releve.periode == periode)
    prev_releves = pq.all()
    if produit_id:
        prev_releves = [r for r in prev_releves if r.pompe.produit_id == produit_id]
    prev_total = round(sum(r.montant_vente for r in prev_releves), 2)

    variation_pct = None
    if prev_total > 0:
        variation_pct = round((total_periode - prev_total) / prev_total * 100, 1)

    return {
        "jours":         result_jours,
        "total_periode": total_periode,
        "variation_pct": variation_pct,
    }


# ---------- Journal des meters + analyse des décalages ----------
# Consolidation : PERIODE_ORDRE (anomalies) et PERIODE_RANG (journal) étaient
# identiques — on garde un seul nom, PERIODE_RANG, utilisé partout.
PERIODE_RANG = {"Matin": 0, "Apres-midi": 1}


def _build_journal_entries(
    db: Session,
    d_debut,
    d_fin,
    produit_id: Optional[int] = None,
    pompe_id_filter: Optional[List[int]] = None,
    periode_filter: Optional[str] = None,
) -> list:
    """
    Source unique de vérité pour la construction des entrées du journal.
    Utilisée par /api/journal ET /api/journal/pdf (Bug 4 fix : plus de duplication).

    Corrections appliquées :
    - Bug 1 : tri de période par PERIODE_RANG (Matin=0 < Apres-midi=1) au lieu
      du tri alphabétique .desc() qui mettait Matin avant Apres-midi.
      Méthode : cherche d'abord la dernière DATE, puis trie les sessions
      de ce jour par PERIODE_RANG pour choisir la vraie dernière session.
    - Bug 2 : un relevé invalide (apres < avant) ne propage pas son ap
      comme référence — la chaîne de comparaison reste propre.
    """
    q = db.query(Releve).filter(Releve.date >= d_debut, Releve.date <= d_fin)
    if pompe_id_filter:
        q = q.filter(Releve.pompe_id.in_(pompe_id_filter))
    # periode_filter N'EST PAS appliqué ici : on a besoin de TOUS les relevés
    # dans l'ordre chronologique pour que l'analyse de continuité (pompe_last)
    # soit correcte. Le filtre est appliqué en post-traitement ci-dessous.
    releves = q.all()
    if produit_id:
        releves = [r for r in releves if r.pompe.produit_id == produit_id]

    # Tri chronologique strict : pompe → date → rang période
    releves.sort(key=lambda r: (r.pompe_id, r.date, PERIODE_RANG.get(r.periode, 99)))

    # ── Dernier metter_apres connu avant la plage (Bug 1 fix) ────────
    # Algorithme en deux temps pour éviter le tri alphabétique incorrect :
    #   1. Trouver la date maximale avant d_debut pour cette pompe (SQL efficace)
    #   2. Charger toutes les sessions de cette date et trier par PERIODE_RANG
    #      → garantit que c'est bien l'Après-midi qui est pris si présent.
    pompe_last: dict = {}
    for pid in {r.pompe_id for r in releves}:
        last_date = (
            db.query(Releve.date)
            .filter(Releve.pompe_id == pid, Releve.date < d_debut)
            .order_by(Releve.date.desc())
            .limit(1)
            .scalar()
        )
        if last_date is None:
            pompe_last[pid] = None
        else:
            last_day = (
                db.query(Releve)
                .filter(Releve.pompe_id == pid, Releve.date == last_date)
                .all()
            )
            last_day.sort(key=lambda r: PERIODE_RANG.get(r.periode, 99))
            pompe_last[pid] = float(last_day[-1].metter_apres)

    entries = []
    for r in releves:
        pid  = r.pompe_id
        av   = float(r.metter_avant)
        ap   = float(r.metter_apres)
        qte  = float(r.quantite)
        mnt  = float(r.montant_vente)
        prec = pompe_last.get(pid)
        decalage = round(av - prec, 3) if prec is not None else None

        releve_valide = av <= ap  # Bug 2 : flag de validité interne

        if not releve_valide:
            statut        = "erreur"
            type_anomalie = "saisie_invalide"
            commentaire   = (
                f"Le meter après ({ap:.3f}) est inférieur au meter avant ({av:.3f}). "
                "Un compteur ne peut physiquement pas reculer — saisie à corriger immédiatement."
            )
            recommandation = (
                f"Ouvrez Saisie, activez la modification de cette pompe et corrigez la valeur. "
                f"La valeur après doit être ≥ {av:.3f}."
            )
            # Bug 2 fix : NE PAS mettre à jour pompe_last — ap est corrompu.
            # On conserve la dernière référence valide connue.

        elif decalage is not None and decalage < -0.001:
            statut        = "erreur"
            type_anomalie = "recul_compteur"
            commentaire   = (
                f"Le compteur a reculé entre les sessions. "
                f"La session précédente s'est terminée à {prec:.3f} mais cette session "
                f"démarre à {av:.3f} (recul de {abs(decalage):.3f})."
            )
            recommandation = (
                f"Vérifiez si le compteur a été remplacé ou manipulé. "
                f"En cas d'erreur de saisie, corrigez le meter avant à {prec:.3f}. "
                "En cas de remplacement confirmé, documentez la nouvelle base."
            )
            pompe_last[pid] = ap

        elif decalage is not None and decalage > 0.001:
            statut        = "alerte"
            type_anomalie = "saut_compteur"
            commentaire   = (
                f"Saut du compteur : attendu {prec:.3f} (fin de la session précédente) "
                f"mais le meter avant saisi est {av:.3f}. "
                f"Écart positif de {decalage:.3f} — correspond à {decalage:.3f} gallons non comptabilisés."
            )
            recommandation = (
                f"Vérifiez si une session (matin ou après-midi) a été omise. "
                f"Si le compteur a été remplacé, enregistrez la nouvelle base {av:.3f} "
                "et annulez l'alerte en ajoutant une note explicative."
            )
            pompe_last[pid] = ap

        else:
            statut = "ok"
            if prec is None:
                type_anomalie  = "premiere_lecture"
                commentaire    = f"Première lecture enregistrée pour cette pompe. Base de référence : {av:.3f}."
                recommandation = "Aucune action requise. Ce relevé servira de base de comparaison."
            else:
                type_anomalie  = "ok"
                commentaire    = (
                    f"Continuité parfaite : le compteur reprend exactement là où "
                    f"la session précédente s'est arrêtée ({prec:.3f} → {av:.3f})."
                )
                recommandation = "Aucune action requise."
            pompe_last[pid] = ap

        entries.append({
            "id":               r.id,
            "date":             str(r.date),
            "periode":          r.periode,
            "pompe_id":         pid,
            "pompe_nom":        r.pompe.nom,
            "produit_nom":      r.pompe.produit.nom,
            "prix_gallon":      float(r.prix_gallon),
            "metter_avant":     av,
            "metter_apres":     ap,
            "metter_attendu":   prec,
            "decalage":         decalage,
            "ecart_gallons":    round(abs(decalage), 3) if decalage is not None else None,
            "quantite":         round(qte, 3),
            "montant_vente":    round(mnt, 2),
            "nb_modifications": r.nb_modifications,
            "statut":           statut,
            "type_anomalie":    type_anomalie,
            "commentaire":      commentaire,
            "recommandation":   recommandation,
        })

    # Post-filtre période : appliqué APRÈS l'analyse de continuité pour ne pas
    # casser la chaîne Matin → Après-midi → Matin suivant.
    if periode_filter:
        entries = [e for e in entries if e["periode"] == periode_filter]

    return entries


@app.get("/api/journal")
def journal_endpoint(
    date_debut: Optional[date_type] = None,
    date_fin:   Optional[date_type] = None,
    produit_id: Optional[int] = None,
    pompe_id:   List[int] = Query(default=[]),
    periode:    Optional[str] = None,
    db: Session = Depends(get_db),
):
    from datetime import date as dt

    d_fin   = date_fin   or dt.today()
    d_debut = date_debut or dt(d_fin.year, d_fin.month, 1)

    entries = _build_journal_entries(db, d_debut, d_fin, produit_id, pompe_id or None, periode)

    nb_alertes = sum(1 for e in entries if e["statut"] == "alerte")
    nb_erreurs = sum(1 for e in entries if e["statut"] == "erreur")
    nb_ok      = len(entries) - nb_alertes - nb_erreurs
    taux_conformite = round(nb_ok / len(entries) * 100, 1) if entries else 100.0

    return {
        "date_debut": str(d_debut),
        "date_fin":   str(d_fin),
        "entries":    entries,
        "resume": {
            "total_entrees":    len(entries),
            "nb_ok":            nb_ok,
            "nb_alertes":       nb_alertes,
            "nb_erreurs":       nb_erreurs,
            "taux_conformite":  taux_conformite,
        },
    }


@app.get("/api/journal/pdf")
def journal_pdf(
    date_debut: Optional[date_type] = None,
    date_fin:   Optional[date_type] = None,
    produit_id: Optional[int] = None,
    pompe_id:   List[int] = Query(default=[]),
    periode:    Optional[str] = None,
    db: Session = Depends(get_db),
):
    import io
    from datetime import date as dt
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    # ── Récupérer les données via la fonction partagée (Bug 4 fix) ───
    d_fin   = date_fin   or dt.today()
    d_debut = date_debut or dt(d_fin.year, d_fin.month, 1)

    entries    = _build_journal_entries(db, d_debut, d_fin, produit_id, pompe_id or None, periode)
    nb_alertes = sum(1 for e in entries if e["statut"] == "alerte")
    nb_erreurs = sum(1 for e in entries if e["statut"] == "erreur")
    nb_ok      = len(entries) - nb_alertes - nb_erreurs

    # ── Couleurs ─────────────────────────────────────────────────────
    C_DARK    = colors.HexColor("#0f172a")
    C_BLUE    = colors.HexColor("#3b82f6")
    C_GREEN   = colors.HexColor("#10b981")
    C_AMBER   = colors.HexColor("#f7a93b")
    C_RED     = colors.HexColor("#e0536a")
    C_GRAY    = colors.HexColor("#64748b")
    C_LIGHT   = colors.HexColor("#f1f5f9")
    C_WHITE   = colors.white

    # ── Styles ───────────────────────────────────────────────────────
    styles = getSampleStyleSheet()
    sTitle = ParagraphStyle("title", fontSize=18, textColor=C_DARK, spaceAfter=2,
                            fontName="Helvetica-Bold", alignment=TA_LEFT)
    sSub   = ParagraphStyle("sub",   fontSize=10, textColor=C_GRAY, spaceAfter=12,
                            fontName="Helvetica")
    sH2    = ParagraphStyle("h2",    fontSize=12, textColor=C_DARK, spaceBefore=14,
                            spaceAfter=6, fontName="Helvetica-Bold")
    sSmall = ParagraphStyle("small", fontSize=8,  textColor=C_GRAY, fontName="Helvetica")
    sBody  = ParagraphStyle("body",  fontSize=9,  textColor=C_DARK, fontName="Helvetica",
                            leading=13)
    sAlHead = ParagraphStyle("alh", fontSize=10, textColor=C_WHITE, fontName="Helvetica-Bold")
    sAlBody = ParagraphStyle("alb", fontSize=8,  textColor=C_DARK,  fontName="Helvetica", leading=12)

    # ── Assemblage du document ───────────────────────────────────────
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    W = landscape(A4)[0] - 3*cm   # largeur utile

    story = []

    # En-tête
    story.append(Paragraph("Journal des Meters", sTitle))
    periode_txt = f"{d_debut.strftime('%d/%m/%Y')} → {d_fin.strftime('%d/%m/%Y')}"
    story.append(Paragraph(f"Période : {periode_txt}  ·  Généré le {dt.today().strftime('%d/%m/%Y')}", sSub))
    story.append(HRFlowable(width="100%", thickness=1, color=C_BLUE, spaceAfter=12))

    # Cartes résumé
    def stat_cell(num, label, col):
        return [
            Paragraph(f'<font color="{col.hexval()}" size="18"><b>{num}</b></font>', styles["Normal"]),
            Paragraph(f'<font color="#64748b" size="8">{label}</font>', styles["Normal"]),
        ]

    sum_data = [[
        stat_cell(len(entries), "Relevés", C_BLUE),
        stat_cell(nb_ok,        "Conformes", C_GREEN),
        stat_cell(nb_alertes,   "Alertes",   C_AMBER),
        stat_cell(nb_erreurs,   "Erreurs",   C_RED),
    ]]
    sum_table = Table(sum_data, colWidths=[W/4]*4)
    sum_table.setStyle(TableStyle([
        ("BOX",       (0,0), (0,0), 0.5, C_BLUE),
        ("BOX",       (1,0), (1,0), 0.5, C_GREEN),
        ("BOX",       (2,0), (2,0), 0.5, C_AMBER),
        ("BOX",       (3,0), (3,0), 0.5, C_RED),
        ("BACKGROUND",(0,0), (0,0), colors.HexColor("#eff6ff")),
        ("BACKGROUND",(1,0), (1,0), colors.HexColor("#ecfdf5")),
        ("BACKGROUND",(2,0), (2,0), colors.HexColor("#fffbeb")),
        ("BACKGROUND",(3,0), (3,0), colors.HexColor("#fff1f3")),
        ("ALIGN",     (0,0), (-1,-1), "CENTER"),
        ("VALIGN",    (0,0), (-1,-1), "MIDDLE"),
        ("ROWPADDING",(0,0), (-1,-1), 10),
    ]))
    story.append(sum_table)
    story.append(Spacer(1, 16))

    # Section alertes/erreurs
    problemes = [e for e in entries if e["statut"] != "ok"]
    if problemes:
        story.append(Paragraph(f"Analyse des décalages ({len(problemes)} problème(s))", sH2))
        for e in problemes:
            is_err = e["statut"] == "erreur"
            bg_col = colors.HexColor("#fff1f3") if is_err else colors.HexColor("#fffbeb")
            bd_col = C_RED if is_err else C_AMBER
            label  = "ERREUR" if is_err else "ALERTE"
            dec_str = f"{e['decalage']:+.3f}" if e["decalage"] is not None else "—"
            att_str = f"{e['metter_attendu']:.3f}" if e["metter_attendu"] is not None else "—"

            al_data = [
                [Paragraph(f"{label} — {e['pompe_nom']} · {e['produit_nom']} | {e['date']} {e['periode']} | Décalage : {dec_str}", sAlHead)],
                [Paragraph(
                    f"<b>Meter avant :</b> {e['metter_avant']:.3f}  |  "
                    f"<b>Attendu :</b> {att_str}  |  "
                    f"<b>Meter après :</b> {e['metter_apres']:.3f}  |  "
                    f"<b>Qté :</b> {e['quantite']:.3f} gal  |  "
                    f"<b>Montant :</b> {e['montant_vente']:,.0f} G",
                    sAlBody
                )],
                [Paragraph(f"<b>Commentaire :</b> {e['commentaire']}", sAlBody)],
                [Paragraph(f"<b>Recommandation :</b> {e['recommandation']}", sAlBody)],
            ]
            al_table = Table(al_data, colWidths=[W])
            al_table.setStyle(TableStyle([
                ("BACKGROUND",  (0,0), (-1,0), bd_col),
                ("BACKGROUND",  (0,1), (-1,-1), bg_col),
                ("BOX",         (0,0), (-1,-1), 0.5, bd_col),
                ("TOPPADDING",  (0,0), (-1,-1), 6),
                ("BOTTOMPADDING",(0,0),(-1,-1), 6),
                ("LEFTPADDING", (0,0), (-1,-1), 10),
                ("RIGHTPADDING",(0,0), (-1,-1), 10),
            ]))
            story.append(al_table)
            story.append(Spacer(1, 6))
        story.append(Spacer(1, 8))
    else:
        story.append(Paragraph("Aucun décalage détecté — tous les relevés présentent une continuité parfaite.", sBody))
        story.append(Spacer(1, 12))

    # Table journal complet
    story.append(Paragraph("Journal complet", sH2))
    if not entries:
        story.append(Paragraph("Aucun relevé pour cette période.", sBody))
    else:
        hdr = ["Date", "Période", "Pompe", "Produit",
               "Meter avant", "Meter après", "Qté (gal)", "Montant (G)",
               "Décalage", "Attendu", "Statut"]
        t_data = [hdr]
        for e in entries:
            dec = f"{e['decalage']:+.3f}" if e["decalage"] is not None else "—"
            att = f"{e['metter_attendu']:.3f}" if e["metter_attendu"] is not None else "—"
            per = "Matin" if e["periode"] == "Matin" else "Après-midi"
            t_data.append([
                e["date"][5:].replace("-", "/") + "/" + e["date"][:4],
                per,
                e["pompe_nom"],
                e["produit_nom"],
                f"{e['metter_avant']:.3f}",
                f"{e['metter_apres']:.3f}",
                f"{e['quantite']:.3f}",
                f"{e['montant_vente']:,.0f}",
                dec,
                att,
                "OK" if e["statut"] == "ok" else e["statut"].upper(),
            ])

        col_w = [2.4*cm, 2.2*cm, 2.4*cm, 2.2*cm,
                 2.2*cm, 2.2*cm, 2.2*cm, 2.4*cm,
                 2.0*cm, 2.0*cm, 1.8*cm]
        tbl = Table(t_data, colWidths=col_w, repeatRows=1)

        ts = [
            ("BACKGROUND",   (0,0), (-1,0), C_DARK),
            ("TEXTCOLOR",    (0,0), (-1,0), C_WHITE),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,0), 7.5),
            ("ALIGN",        (0,0), (-1,-1), "CENTER"),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
            ("FONTSIZE",     (0,1), (-1,-1), 7.5),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [C_WHITE, C_LIGHT]),
            ("GRID",         (0,0), (-1,-1), 0.3, colors.HexColor("#e2e8f0")),
            ("TOPPADDING",   (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
        ]
        # Coloriser les lignes erreur/alerte
        for i, e in enumerate(entries, start=1):
            if e["statut"] == "erreur":
                ts.append(("BACKGROUND", (0,i), (-1,i), colors.HexColor("#fde8ec")))
                ts.append(("TEXTCOLOR",  (-1,i), (-1,i), C_RED))
            elif e["statut"] == "alerte":
                ts.append(("BACKGROUND", (0,i), (-1,i), colors.HexColor("#fff8e6")))
                ts.append(("TEXTCOLOR",  (-1,i), (-1,i), C_AMBER))
            else:
                ts.append(("TEXTCOLOR",  (-1,i), (-1,i), C_GREEN))

        tbl.setStyle(TableStyle(ts))
        story.append(tbl)

    # Pied de page info
    story.append(Spacer(1, 16))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_GRAY))
    story.append(Paragraph(
        f"Station Carburant · Journal des Meters · {periode_txt} · {len(entries)} relevé(s)",
        sSmall
    ))

    doc.build(story)
    buf.seek(0)

    fname = f"journal_{d_debut}_{d_fin}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------- Export Excel ----------
@app.get("/api/releves/export")
def export_releves_xlsx(
    date_debut: Optional[date_type] = None,
    date_fin:   Optional[date_type] = None,
    produit_id: Optional[int] = None,
    pompe_id:   List[int] = Query(default=[]),
    periode:    Optional[str] = None,
    db: Session = Depends(get_db),
):
    import io
    from collections import defaultdict
    from datetime import date as date_cls
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    from openpyxl.utils import get_column_letter

    # ── Requête ────────────────────────────────────────────────────────
    q = (db.query(Releve)
           .join(Pompe, Releve.pompe_id == Pompe.id)
           .join(Produit, Pompe.produit_id == Produit.id))
    if date_debut:  q = q.filter(Releve.date >= date_debut)
    if date_fin:    q = q.filter(Releve.date <= date_fin)
    if produit_id:  q = q.filter(Pompe.produit_id == produit_id)
    if pompe_id:    q = q.filter(Releve.pompe_id.in_(pompe_id))
    if periode:     q = q.filter(Releve.periode == periode)
    releves = q.order_by(Releve.date.desc(), Releve.periode, Releve.pompe_id).all()

    # ── Styles helpers ─────────────────────────────────────────────────
    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color="2D3748", size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def _align(h="left"):
        return Alignment(horizontal=h, vertical="center")

    def _border():
        s = Side(style="thin", color="D5E0E8")
        return Border(left=s, right=s, top=s, bottom=s)

    HDR_FILL  = _fill("F7A93B");  HDR_FONT  = _font(bold=True, color="1A1208", size=11)
    TOT_FILL  = _fill("FFF0CC");  TOT_FONT  = _font(bold=True, color="1A1208", size=11)
    ODD_FILL  = _fill("FFFDF5");  EVN_FILL  = _fill("FFFFFF")
    GRN_HDR   = _fill("3DB88A");  GRN_FILL  = _fill("F0FBF7")
    BASE_FONT = _font()
    NUM_FMT   = "#,##0.000"
    MNT_FMT   = "#,##0.00"
    BD        = _border()

    # ── Agrégations ────────────────────────────────────────────────────
    rows_data = []
    by_prod   = defaultdict(lambda: {"nb": 0, "qte": 0.0, "mnt": 0.0})
    by_pump   = defaultdict(lambda: {"produit": "", "nb": 0, "qte": 0.0, "mnt": 0.0})
    for r in releves:
        qte = round(float(r.metter_apres) - float(r.metter_avant), 3)
        mnt = round(qte * float(r.prix_gallon), 2)
        rows_data.append((r, qte, mnt))
        pn = r.pompe.produit.nom
        po = r.pompe.nom
        by_prod[pn]["nb"]  += 1;  by_prod[pn]["qte"] += qte;  by_prod[pn]["mnt"] += mnt
        by_pump[po]["produit"] = pn
        by_pump[po]["nb"]  += 1;  by_pump[po]["qte"] += qte;  by_pump[po]["mnt"] += mnt

    # ══════════════════════════════════════════════════════════════════
    # Feuille 1 — Relevés
    # ══════════════════════════════════════════════════════════════════
    wb = Workbook()
    ws1 = wb.active;  ws1.title = "Relevés"

    ws1.cell(1, 1, "PétroSync — Relevés de compteurs").font = Font(name="Calibri", bold=True, size=13, color="F7A93B")
    ws1.merge_cells("A1:J1");  ws1.row_dimensions[1].height = 22

    info = f"Exporté le {date_cls.today().isoformat()}  ·  {len(releves)} relevé(s)"
    if date_debut or date_fin:
        info += f"  ·  Période : {date_debut or '…'} → {date_fin or '…'}"
    ws1.cell(2, 1, info).font = Font(name="Calibri", size=9, color="7A8CA0")
    ws1.merge_cells("A2:J2");  ws1.row_dimensions[2].height = 14

    HDR1  = ["#", "Date", "Période", "Produit", "Pompe",
             "Prix/Gallon (G)", "Meter Avant", "Meter Après", "Quantité (gal)", "Montant (G)"]
    WDT1  = [6, 13, 13, 16, 16, 18, 14, 14, 15, 15]
    ws1.row_dimensions[3].height = 26
    for col, (h, w) in enumerate(zip(HDR1, WDT1), 1):
        c = ws1.cell(3, col, h)
        c.font = HDR_FONT;  c.fill = HDR_FILL;  c.alignment = _align("center");  c.border = BD
        ws1.column_dimensions[get_column_letter(col)].width = w

    total_qte = total_mnt = 0.0
    for ri, (r, qte, mnt) in enumerate(rows_data, 4):
        rf = ODD_FILL if ri % 2 else EVN_FILL
        total_qte += qte;  total_mnt += mnt
        vals = [r.id, str(r.date), r.periode, r.pompe.produit.nom, r.pompe.nom,
                float(r.prix_gallon), float(r.metter_avant), float(r.metter_apres), qte, mnt]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(ri, col, val)
            c.font = BASE_FONT;  c.fill = rf;  c.border = BD
            if col in (6, 7, 8, 9): c.number_format = NUM_FMT;  c.alignment = _align("right")
            elif col == 10:          c.number_format = MNT_FMT;  c.alignment = _align("right")
            elif col == 1:           c.alignment = _align("center")
            else:                    c.alignment = _align("left")
        ws1.row_dimensions[ri].height = 17

    tr1 = len(rows_data) + 4
    for col in range(1, 11):
        c = ws1.cell(tr1, col);  c.fill = TOT_FILL;  c.font = TOT_FONT;  c.border = BD
    ws1.cell(tr1, 1, "TOTAL").alignment = _align("left")
    ws1.cell(tr1, 2, f"{len(rows_data)} relevé(s)").alignment = _align("left")
    ws1.cell(tr1, 9, round(total_qte, 3)).number_format = NUM_FMT;  ws1.cell(tr1, 9).alignment = _align("right")
    ws1.cell(tr1, 10, round(total_mnt, 2)).number_format = MNT_FMT; ws1.cell(tr1, 10).alignment = _align("right")
    ws1.row_dimensions[tr1].height = 22
    ws1.freeze_panes = "A4"

    # ══════════════════════════════════════════════════════════════════
    # Feuille 2 — Par Produit
    # ══════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Par Produit")
    ws2.cell(1, 1, "Résumé par produit").font = Font(name="Calibri", bold=True, size=12, color="3DB88A")
    ws2.merge_cells("A1:E1");  ws2.row_dimensions[1].height = 22
    HDR2 = ["Produit", "Nb Relevés", "Volume total (gal)", "Montant total (G)", "Prix moy. (G/gal)"]
    WDT2 = [18, 13, 20, 20, 18]
    ws2.row_dimensions[2].height = 24
    for col, (h, w) in enumerate(zip(HDR2, WDT2), 1):
        c = ws2.cell(2, col, h)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c.fill = GRN_HDR;  c.alignment = _align("center");  c.border = BD
        ws2.column_dimensions[get_column_letter(col)].width = w

    t2_nb = t2_qte = t2_mnt = 0
    for ri, (pnom, v) in enumerate(sorted(by_prod.items()), 3):
        rf = GRN_FILL if ri % 2 else EVN_FILL
        pm = round(v["mnt"] / v["qte"], 3) if v["qte"] else 0
        t2_nb += v["nb"];  t2_qte += v["qte"];  t2_mnt += v["mnt"]
        for col, val in enumerate([pnom, v["nb"], round(v["qte"], 3), round(v["mnt"], 2), pm], 1):
            c = ws2.cell(ri, col, val);  c.font = BASE_FONT;  c.fill = rf;  c.border = BD
            if col == 1:  c.alignment = _align("left")
            else:
                c.alignment = _align("right")
                c.number_format = "#,##0" if col == 2 else (MNT_FMT if col == 4 else NUM_FMT)
        ws2.row_dimensions[ri].height = 17

    tr2 = len(by_prod) + 3
    for col in range(1, 6):
        c = ws2.cell(tr2, col);  c.fill = TOT_FILL;  c.font = TOT_FONT;  c.border = BD
    ws2.cell(tr2, 1, "TOTAL").alignment = _align("left")
    ws2.cell(tr2, 2, t2_nb).alignment = _align("right");  ws2.cell(tr2, 2).number_format = "#,##0"
    ws2.cell(tr2, 3, round(t2_qte, 3)).number_format = NUM_FMT; ws2.cell(tr2, 3).alignment = _align("right")
    ws2.cell(tr2, 4, round(t2_mnt, 2)).number_format = MNT_FMT; ws2.cell(tr2, 4).alignment = _align("right")
    ws2.row_dimensions[tr2].height = 22
    ws2.freeze_panes = "A3"

    # ══════════════════════════════════════════════════════════════════
    # Feuille 3 — Par Pompe
    # ══════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Par Pompe")
    ws3.cell(1, 1, "Résumé par pompe").font = Font(name="Calibri", bold=True, size=12, color="F7A93B")
    ws3.merge_cells("A1:E1");  ws3.row_dimensions[1].height = 22
    HDR3 = ["Pompe", "Produit", "Nb Relevés", "Volume total (gal)", "Montant total (G)"]
    WDT3 = [18, 16, 13, 20, 20]
    ws3.row_dimensions[2].height = 24
    for col, (h, w) in enumerate(zip(HDR3, WDT3), 1):
        c = ws3.cell(2, col, h)
        c.font = HDR_FONT;  c.fill = HDR_FILL;  c.alignment = _align("center");  c.border = BD
        ws3.column_dimensions[get_column_letter(col)].width = w

    t3_nb = t3_qte = t3_mnt = 0
    for ri, (ponom, v) in enumerate(sorted(by_pump.items()), 3):
        rf = ODD_FILL if ri % 2 else EVN_FILL
        t3_nb += v["nb"];  t3_qte += v["qte"];  t3_mnt += v["mnt"]
        for col, val in enumerate([ponom, v["produit"], v["nb"], round(v["qte"], 3), round(v["mnt"], 2)], 1):
            c = ws3.cell(ri, col, val);  c.font = BASE_FONT;  c.fill = rf;  c.border = BD
            if col <= 2: c.alignment = _align("left")
            else:
                c.alignment = _align("right")
                c.number_format = "#,##0" if col == 3 else (MNT_FMT if col == 5 else NUM_FMT)
        ws3.row_dimensions[ri].height = 17

    tr3 = len(by_pump) + 3
    for col in range(1, 6):
        c = ws3.cell(tr3, col);  c.fill = TOT_FILL;  c.font = TOT_FONT;  c.border = BD
    ws3.cell(tr3, 1, "TOTAL").alignment = _align("left")
    ws3.cell(tr3, 3, t3_nb).alignment = _align("right");  ws3.cell(tr3, 3).number_format = "#,##0"
    ws3.cell(tr3, 4, round(t3_qte, 3)).number_format = NUM_FMT; ws3.cell(tr3, 4).alignment = _align("right")
    ws3.cell(tr3, 5, round(t3_mnt, 2)).number_format = MNT_FMT; ws3.cell(tr3, 5).alignment = _align("right")
    ws3.row_dimensions[tr3].height = 22
    ws3.freeze_panes = "A3"

    # ── Stream ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf);  buf.seek(0)

    parts = ["petrosync_releves"]
    if date_debut: parts.append(str(date_debut))
    if date_fin:   parts.append(str(date_fin))
    fname = "_".join(parts) + ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------- Rapport analytique journalier ----------

def _build_rapport_data(db: Session, date_rapport: date_type) -> dict:
    """Agrège toutes les données nécessaires au rapport PDF + Excel."""
    from datetime import timedelta
    from collections import defaultdict

    d = date_rapport
    hier      = d - timedelta(days=1)
    d7_debut  = d - timedelta(days=6)
    d30_debut = d - timedelta(days=29)
    mois_debut = date_type(d.year, d.month, 1)

    def stats(debut, fin):
        q = db.query(Releve).filter(Releve.date >= debut, Releve.date <= fin).all()
        return {
            "nb_releves":    len(q),
            "total_montant": round(sum(r.montant_vente for r in q), 2),
            "total_quantite": round(sum(r.quantite for r in q), 3),
            "jours_actifs":  len({str(r.date) for r in q}),
        }

    auj   = stats(d,          d)
    v     = stats(hier,       hier)
    s7    = stats(d7_debut,   d)
    s30   = stats(d30_debut,  d)
    mois  = stats(mois_debut, d)

    # Variation vs veille
    if v["total_montant"] > 0:
        var_hier = round((auj["total_montant"] - v["total_montant"]) / v["total_montant"] * 100, 1)
    else:
        var_hier = None

    # Détail par pompe (mois)
    mois_releves = db.query(Releve).filter(Releve.date >= mois_debut, Releve.date <= d).all()
    par_pompe = defaultdict(lambda: {"quantite": 0.0, "montant": 0.0, "produit": ""})
    par_periode = {"Matin": {"quantite": 0.0, "montant": 0.0},
                   "Apres-midi": {"quantite": 0.0, "montant": 0.0}}
    par_produit = defaultdict(lambda: {"quantite": 0.0, "montant": 0.0})
    for r in mois_releves:
        par_pompe[r.pompe.nom]["quantite"]  += r.quantite
        par_pompe[r.pompe.nom]["montant"]   += r.montant_vente
        par_pompe[r.pompe.nom]["produit"]    = r.pompe.produit.nom
        par_periode[r.periode]["quantite"]  += r.quantite
        par_periode[r.periode]["montant"]   += r.montant_vente
        par_produit[r.pompe.produit.nom]["quantite"] += r.quantite
        par_produit[r.pompe.produit.nom]["montant"]  += r.montant_vente

    # Série 30 jours (pour tableau)
    q30 = db.query(Releve).filter(Releve.date >= d30_debut, Releve.date <= d).all()
    serie: dict = defaultdict(lambda: {"montant": 0.0, "quantite": 0.0})
    for r in q30:
        serie[str(r.date)]["montant"]   += r.montant_vente
        serie[str(r.date)]["quantite"]  += r.quantite
    jours_actifs = {k: v for k, v in serie.items() if v["montant"] > 0}
    moy_jours_actifs = (
        round(sum(v["montant"] for v in jours_actifs.values()) / len(jours_actifs), 2)
        if jours_actifs else 0
    )

    # Anomalies du jour
    anom_q = db.query(Releve).filter(Releve.date <= d).all()
    nb_anom = 0
    anom_list = []
    for pid in {r.pompe_id for r in anom_q}:
        prev_ap = None
        for r in sorted(
            [x for x in anom_q if x.pompe_id == pid],
            key=lambda x: (x.date, PERIODE_RANG.get(x.periode, 9))
        ):
            av, ap = float(r.metter_avant), float(r.metter_apres)
            if av > ap:
                nb_anom += 1
                anom_list.append({
                    "type": "Saisie invalide",
                    "pompe": r.pompe.nom,
                    "date": str(r.date),
                    "periode": r.periode,
                    "detail": f"meter avant {av:.0f} > meter après {ap:.0f}",
                })
            elif prev_ap is not None and av < prev_ap - 0.001:
                nb_anom += 1
                anom_list.append({
                    "type": "Régression compteur",
                    "pompe": r.pompe.nom,
                    "date": str(r.date),
                    "periode": r.periode,
                    "detail": f"attendu {prev_ap:.0f}, saisi {av:.0f}",
                })
            if av <= ap:
                prev_ap = ap

    prix_gal = round(mois_releves[0].prix_gallon, 2) if mois_releves else 0

    return {
        "date_rapport": str(d),
        "prix_gallon":  float(prix_gal),
        "aujourd_hui":  auj,
        "veille":       v,
        "variation_hier": var_hier,
        "semaine_7j":   s7,
        "mois_30j":     s30,
        "mois_courant": mois,
        "mois_debut":   str(mois_debut),
        "par_pompe":    {k: {kk: round(vv, 2) if isinstance(vv, float) else vv
                             for kk, vv in vals.items()}
                         for k, vals in par_pompe.items()},
        "par_periode":  par_periode,
        "par_produit":  dict(par_produit),
        "jours_actifs_30j": len(jours_actifs),
        "moy_jours_actifs": moy_jours_actifs,
        "serie_30j":    dict(serie),
        "nb_anomalies": nb_anom,
        "anomalies":    anom_list,
    }


@app.get("/api/rapport/pdf")
def rapport_pdf(
    date: Optional[date_type] = None,
    db: Session = Depends(get_db),
    request: Request = None,
):
    """Rapport analytique journalier complet au format PDF."""
    from datetime import date as dt
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    d = date or dt.today()
    data = _build_rapport_data(db, d)

    C_BG    = colors.HexColor("#0f172a")
    C_BLUE  = colors.HexColor("#3b82f6")
    C_RED   = colors.HexColor("#ef4444")
    C_GREEN = colors.HexColor("#22c55e")
    C_AMBER = colors.HexColor("#f97316")
    C_GRAY  = colors.HexColor("#64748b")
    C_LIGHT = colors.HexColor("#f1f5f9")
    C_WHITE = colors.white

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=18,
                                 textColor=C_WHITE, spaceAfter=4, alignment=TA_CENTER)
    sub_style   = ParagraphStyle("sub",   fontName="Helvetica",      fontSize=10,
                                 textColor=C_GRAY,  spaceAfter=2,  alignment=TA_CENTER)
    h2_style    = ParagraphStyle("h2",    fontName="Helvetica-Bold", fontSize=12,
                                 textColor=C_BLUE,  spaceBefore=12, spaceAfter=6)
    body_style  = ParagraphStyle("body",  fontName="Helvetica",      fontSize=9,
                                 textColor=colors.HexColor("#334155"), spaceAfter=4)
    warn_style  = ParagraphStyle("warn",  fontName="Helvetica-Bold", fontSize=9,
                                 textColor=C_RED,   spaceAfter=3)

    def tbl(data_rows, col_widths, header=True):
        t = Table(data_rows, colWidths=col_widths)
        cmds = [
            ("BACKGROUND",  (0, 0), (-1, 0 if header else -1), C_BG),
            ("TEXTCOLOR",   (0, 0), (-1, 0 if header else -1), C_WHITE),
            ("FONTNAME",    (0, 0), (-1, 0 if header else -1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 8),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1 if header else 0), (-1, -1), [C_LIGHT, C_WHITE]),
            ("GRID",        (0, 0), (-1, -1), 0.4, C_GRAY),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        t.setStyle(TableStyle(cmds))
        return t

    def g(n):
        return f"{n:,.0f} G".replace(",", " ")
    def gal(n):
        return f"{n:.3f} gal"
    def pct(n):
        if n is None: return "N/A"
        sign = "+" if n >= 0 else ""
        return f"{sign}{n:.1f}%"

    story = []
    buf   = io.BytesIO()
    doc   = SimpleDocTemplate(buf, pagesize=A4,
                               leftMargin=1.8*cm, rightMargin=1.8*cm,
                               topMargin=1.5*cm, bottomMargin=1.5*cm)

    # ── En-tête ──────────────────────────────────────────────────────
    header_tbl = Table([[
        Paragraph("PétroSync", title_style),
        Paragraph(f"Rapport Analytique · {d.strftime('%d %B %Y')}", sub_style),
        Paragraph("Complexe Commercial Pillatre", sub_style),
    ]], colWidths=[5*cm, 9*cm, 5*cm])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), C_BG),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 0.4*cm))

    # ── Résumé exécutif ───────────────────────────────────────────────
    story.append(Paragraph("📌 Résumé Exécutif", h2_style))
    auj = data["aujourd_hui"]
    v   = data["veille"]
    var = data["variation_hier"]
    if auj["nb_releves"] == 0:
        exec_txt = (
            f"<b>Aucune saisie enregistrée pour le {d.strftime('%d/%m/%Y')}.</b> "
            f"Dernière journée active : {data['veille']['total_montant']:,.0f} G (veille). "
            f"CA mois courant : <b>{g(data['mois_courant']['total_montant'])}</b> "
            f"sur {data['mois_courant']['jours_actifs']} jours actifs. "
            f"Anomalies détectées : <b>{data['nb_anomalies']}</b>."
        )
    else:
        exec_txt = (
            f"CA du jour : <b>{g(auj['total_montant'])}</b> · "
            f"{gal(auj['total_quantite'])} vendus · "
            f"Variation vs veille : <b>{pct(var)}</b>. "
            f"CA mois courant : <b>{g(data['mois_courant']['total_montant'])}</b>. "
            f"Anomalies : <b>{data['nb_anomalies']}</b>."
        )
    story.append(Paragraph(exec_txt, body_style))
    story.append(Spacer(1, 0.3*cm))

    # ── Statistiques clés ─────────────────────────────────────────────
    story.append(Paragraph("📊 Statistiques Clés", h2_style))
    kpi_rows = [
        ["Indicateur",          "Valeur",                            "Détail"],
        ["CA aujourd'hui",      g(auj["total_montant"]),             f"{auj['nb_releves']} relevé(s)"],
        ["CA veille",           g(v["total_montant"]),               f"{v['nb_releves']} relevé(s)"],
        ["Variation vs veille", pct(var),                            ""],
        ["CA 7 derniers jours", g(data["semaine_7j"]["total_montant"]), f"{data['semaine_7j']['jours_actifs']} j actifs"],
        ["CA mois courant",     g(data["mois_courant"]["total_montant"]), f"depuis {data['mois_debut']}"],
        ["Moy. jours actifs",   g(data["moy_jours_actifs"]),        f"{data['jours_actifs_30j']} j actifs/30j"],
        ["Prix au gallon",      f"{data['prix_gallon']:.2f} G/gal", ""],
        ["Anomalies détectées", str(data["nb_anomalies"]),           "Voir section Alertes"],
    ]
    story.append(tbl(kpi_rows, [6*cm, 5*cm, 6*cm]))
    story.append(Spacer(1, 0.4*cm))

    # ── Répartition par pompe ─────────────────────────────────────────
    story.append(Paragraph("📈 Répartition par Pompe (mois courant)", h2_style))
    pompe_rows = [["Pompe", "Produit", "Volume (gal)", "CA (G)", "Part CA %"]]
    total_m = data["mois_courant"]["total_montant"] or 1
    for nom, vals in data["par_pompe"].items():
        part = round(vals["montant"] / total_m * 100, 1)
        pompe_rows.append([nom, vals["produit"], gal(vals["quantite"]),
                           g(vals["montant"]), f"{part}%"])
    story.append(tbl(pompe_rows, [4*cm, 3.5*cm, 3.5*cm, 4*cm, 2*cm]))
    story.append(Spacer(1, 0.4*cm))

    # ── Répartition par période ───────────────────────────────────────
    story.append(Paragraph("📈 Répartition par Période (mois courant)", h2_style))
    per_rows = [["Période", "Volume (gal)", "CA (G)", "Part CA %"]]
    for p, vals in data["par_periode"].items():
        part = round(vals["montant"] / total_m * 100, 1)
        per_rows.append([p, gal(vals["quantite"]), g(vals["montant"]), f"{part}%"])
    story.append(tbl(per_rows, [4*cm, 4*cm, 5*cm, 4*cm]))
    story.append(Spacer(1, 0.4*cm))

    # ── Alertes ──────────────────────────────────────────────────────
    story.append(Paragraph("⚠️ Alertes et Anomalies", h2_style))
    alerts = []
    if auj["nb_releves"] == 0:
        alerts.append("Aucune saisie enregistrée pour aujourd'hui.")
    if data["par_periode"]["Apres-midi"]["montant"] == 0:
        alerts.append("Session Après-midi : 0 G enregistré sur tout le mois — saisie manquante ou station fermée l'après-midi.")
    if data["mois_courant"]["jours_actifs"] < 5:
        alerts.append(f"Seulement {data['mois_courant']['jours_actifs']} jour(s) actif(s) ce mois — historique incomplet.")
    if data["nb_anomalies"] > 0:
        alerts.append(f"{data['nb_anomalies']} anomalie(s) compteur détectée(s) — corriger dans Saisie.")
    for a in data["anomalies"]:
        alerts.append(f"• {a['type']} | {a['pompe']} | {a['date']} {a['periode']} : {a['detail']}")
    # Concentration pompe
    for nom, vals in data["par_pompe"].items():
        part = round(vals["montant"] / total_m * 100, 1)
        if part > 85:
            alerts.append(f"Concentration élevée : {nom} représente {part}% du CA — risque en cas de panne.")

    if not alerts:
        story.append(Paragraph("Aucune alerte détectée.", body_style))
    else:
        for a in alerts:
            story.append(Paragraph(f"⚠ {a}", warn_style))
    story.append(Spacer(1, 0.4*cm))

    # ── Pistes de décision ────────────────────────────────────────────
    story.append(Paragraph("✅ Pistes de Décision", h2_style))
    dec_rows = [["Priorité", "Action", "Objectif", "Impact"]]
    if data["nb_anomalies"] > 0:
        dec_rows.append(["🔴 Immédiat", "Corriger anomalies compteur", "Intégrité des données", "Élevé"])
    if auj["nb_releves"] == 0:
        dec_rows.append(["🔴 Immédiat", "Saisir relevés du jour", "Couverture 100%", "Élevé"])
    if data["par_periode"]["Apres-midi"]["montant"] == 0:
        dec_rows.append(["🟡 Cette semaine", "Clarifier politique Après-midi", "Doubler CA potentiel", "Très élevé"])
    if data["mois_courant"]["jours_actifs"] < 10:
        dec_rows.append(["🟡 Cette semaine", "Reconstituer l'historique", "Analyses fiables", "Élevé"])
    dec_rows.append(["🟢 Ce mois", "Former opérateurs à la saisie", "Données complètes", "Moyen"])
    story.append(tbl(dec_rows, [2.5*cm, 5*cm, 5*cm, 4.5*cm]))

    # ── Pied de page ─────────────────────────────────────────────────
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_GRAY))
    story.append(Paragraph(
        f"Généré automatiquement par PétroSync · {d.strftime('%d/%m/%Y')} · Complexe Commercial Pillatre",
        ParagraphStyle("footer", fontName="Helvetica", fontSize=7, textColor=C_GRAY, alignment=TA_CENTER)
    ))

    doc.build(story)
    buf.seek(0)
    fname = f"Rapport_Ventes_{d}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@app.get("/api/rapport/xlsx")
def rapport_xlsx(
    date: Optional[date_type] = None,
    db: Session = Depends(get_db),
    request: Request = None,
):
    """Rapport analytique journalier complet au format Excel."""
    from datetime import date as dt
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    d    = date or dt.today()
    data = _build_rapport_data(db, d)

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────
    NAVY  = "0F172A"
    BLUE  = "3B82F6"
    RED   = "EF4444"
    GREEN = "22C55E"
    AMBER = "F97316"
    GRAY  = "64748B"
    LGRAY = "F1F5F9"
    WHITE = "FFFFFF"

    def hdr_font(color=WHITE):      return Font(bold=True, color=color, name="Calibri", size=10)
    def norm_font(bold=False):      return Font(bold=bold, name="Calibri", size=10)
    def fill(hex_col):              return PatternFill("solid", fgColor=hex_col)
    def border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    def center():                   return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left():                     return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def write_hdr(ws, row, col, val, bg=NAVY, fg=WHITE):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = hdr_font(fg)
        c.fill      = fill(bg)
        c.alignment = center()
        c.border    = border()
        return c

    def write_cell(ws, row, col, val, bold=False, bg=None, num_fmt=None, align_fn=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = norm_font(bold)
        c.border    = border()
        c.alignment = (align_fn or left)()
        if bg:      c.fill = fill(bg)
        if num_fmt: c.number_format = num_fmt
        return c

    # ═══════════════════════════════════════════════════════════════
    # Feuille 1 : Résumé
    # ═══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.sheet_view.showGridLines = False

    # Titre
    ws1.merge_cells("A1:E1")
    c = ws1["A1"]
    c.value     = f"RAPPORT ANALYTIQUE · {d.strftime('%d/%m/%Y')} · Complexe Commercial Pillatre"
    c.font      = Font(bold=True, color=WHITE, name="Calibri", size=13)
    c.fill      = fill(NAVY)
    c.alignment = center()
    ws1.row_dimensions[1].height = 28

    # KPI
    kpis = [
        ("CA aujourd'hui",          data["aujourd_hui"]["total_montant"],    "# ### ##0 G"),
        ("CA veille",               data["veille"]["total_montant"],         "# ### ##0 G"),
        ("Variation vs veille",
         (data["variation_hier"] or 0) / 100 if data["variation_hier"] is not None else "",
         "0.0%;-0.0%"),
        ("CA 7 derniers jours",     data["semaine_7j"]["total_montant"],     "# ### ##0 G"),
        ("CA mois courant",         data["mois_courant"]["total_montant"],   "# ### ##0 G"),
        ("Volume mois (gal)",       data["mois_courant"]["total_quantite"],  "0.000"),
        ("Prix au gallon (G)",      data["prix_gallon"],                     "0.00"),
        ("Jours actifs / 30j",      data["jours_actifs_30j"],               "0"),
        ("Moy. jours actifs",       data["moy_jours_actifs"],               "# ### ##0 G"),
        ("Anomalies",               data["nb_anomalies"],                    "0"),
    ]
    write_hdr(ws1, 3, 1, "Indicateur", BLUE)
    write_hdr(ws1, 3, 2, "Valeur",     BLUE)
    write_hdr(ws1, 3, 3, "Contexte",   BLUE)
    for i, (lbl, val, fmt) in enumerate(kpis, start=4):
        bg = LGRAY if i % 2 == 0 else WHITE
        write_cell(ws1, i, 1, lbl, bold=True, bg=bg)
        c = write_cell(ws1, i, 2, val, bg=bg, num_fmt=fmt, align_fn=center)
        write_cell(ws1, i, 3, "", bg=bg)

    ws1.column_dimensions["A"].width = 2
    ws1.column_dimensions["B"].width = 32
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 22
    ws1.column_dimensions["E"].width = 20

    # ═══════════════════════════════════════════════════════════════
    # Feuille 2 : Données détaillées (série 30j)
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Données")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:E1")
    c = ws2["A1"]
    c.value     = "SÉRIE 30 JOURS · Détail par journée"
    c.font      = Font(bold=True, color=WHITE, name="Calibri", size=12)
    c.fill      = fill(NAVY)
    c.alignment = center()
    ws2.row_dimensions[1].height = 24

    hdrs = ["Date", "CA (G)", "Volume (gal)", "Statut"]
    for ci, h in enumerate(hdrs, 1):
        write_hdr(ws2, 3, ci, h, BLUE)

    from datetime import timedelta
    cursor = d - timedelta(days=29)
    row = 4
    while cursor <= d:
        ds    = str(cursor)
        vals  = data["serie_30j"].get(ds, {"montant": 0.0, "quantite": 0.0})
        m     = vals["montant"]
        q     = vals["quantite"]
        statut = "Actif" if m > 0 else "Sans saisie"
        bg     = WHITE if m > 0 else LGRAY
        write_cell(ws2, row, 1, cursor, bg=bg, num_fmt="DD/MM/YYYY")
        write_cell(ws2, row, 2, m,      bg=bg, num_fmt="# ### ##0 G", align_fn=center)
        write_cell(ws2, row, 3, q,      bg=bg, num_fmt="0.000",        align_fn=center)
        write_cell(ws2, row, 4, statut, bg=bg)
        row += 1
        cursor += timedelta(days=1)

    for ci, w in enumerate([14, 18, 16, 14], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ═══════════════════════════════════════════════════════════════
    # Feuille 3 : Répartition
    # ═══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Répartition")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:E1")
    c = ws3["A1"]
    c.value     = "RÉPARTITION · Pompes et Périodes (mois courant)"
    c.font      = Font(bold=True, color=WHITE, name="Calibri", size=12)
    c.fill      = fill(NAVY)
    c.alignment = center()
    ws3.row_dimensions[1].height = 24

    # Par pompe
    write_hdr(ws3, 3, 1, "Pompe",       BLUE)
    write_hdr(ws3, 3, 2, "Produit",     BLUE)
    write_hdr(ws3, 3, 3, "Volume (gal)", BLUE)
    write_hdr(ws3, 3, 4, "CA (G)",       BLUE)
    write_hdr(ws3, 3, 5, "Part CA",      BLUE)
    total_m = data["mois_courant"]["total_montant"] or 1
    row = 4
    for nom, vals in data["par_pompe"].items():
        part = vals["montant"] / total_m
        bg   = LGRAY if row % 2 == 0 else WHITE
        write_cell(ws3, row, 1, nom,              bg=bg, bold=True)
        write_cell(ws3, row, 2, vals["produit"],  bg=bg)
        write_cell(ws3, row, 3, vals["quantite"], bg=bg, num_fmt="0.000",       align_fn=center)
        write_cell(ws3, row, 4, vals["montant"],  bg=bg, num_fmt="# ### ##0 G", align_fn=center)
        write_cell(ws3, row, 5, part,             bg=bg, num_fmt="0.0%",         align_fn=center)
        row += 1

    row += 1
    write_hdr(ws3, row, 1, "Période", BLUE)
    write_hdr(ws3, row, 3, "Volume (gal)", BLUE)
    write_hdr(ws3, row, 4, "CA (G)", BLUE)
    write_hdr(ws3, row, 5, "Part CA", BLUE)
    row += 1
    for p, vals in data["par_periode"].items():
        part = vals["montant"] / total_m
        bg   = LGRAY if row % 2 == 0 else WHITE
        write_cell(ws3, row, 1, p,              bg=bg, bold=True)
        write_cell(ws3, row, 3, vals["quantite"], bg=bg, num_fmt="0.000",        align_fn=center)
        write_cell(ws3, row, 4, vals["montant"],  bg=bg, num_fmt="# ### ##0 G",  align_fn=center)
        write_cell(ws3, row, 5, part,             bg=bg, num_fmt="0.0%",          align_fn=center)
        row += 1

    for ci, w in enumerate([20, 14, 16, 18, 12], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # ═══════════════════════════════════════════════════════════════
    # Feuille 4 : Alertes
    # ═══════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Alertes")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:D1")
    c = ws4["A1"]
    c.value     = "ALERTES ET ANOMALIES"
    c.font      = Font(bold=True, color=WHITE, name="Calibri", size=12)
    c.fill      = fill(RED)
    c.alignment = center()
    ws4.row_dimensions[1].height = 24

    write_hdr(ws4, 3, 1, "Type",    RED)
    write_hdr(ws4, 3, 2, "Pompe",   RED)
    write_hdr(ws4, 3, 3, "Date",    RED)
    write_hdr(ws4, 3, 4, "Détail",  RED)
    row = 4
    if not data["anomalies"]:
        ws4.merge_cells(f"A{row}:D{row}")
        c = ws4.cell(row=row, column=1, value="Aucune anomalie détectée")
        c.font = Font(color=GREEN, bold=True, name="Calibri", size=10)
        c.fill = fill(LGRAY)
        c.alignment = center()
    for a in data["anomalies"]:
        write_cell(ws4, row, 1, a["type"],   bold=True, bg="FEE2E2")
        write_cell(ws4, row, 2, a["pompe"],  bg="FEE2E2")
        write_cell(ws4, row, 3, f"{a['date']} {a['periode']}", bg="FEE2E2")
        write_cell(ws4, row, 4, a["detail"], bg="FEE2E2")
        row += 1
    for ci, w in enumerate([24, 16, 20, 36], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    # ═══════════════════════════════════════════════════════════════
    # Feuille 5 : Décisions
    # ═══════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Décisions")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:E1")
    c = ws5["A1"]
    c.value     = "PISTES DE DÉCISION CLASSÉES PAR PRIORITÉ"
    c.font      = Font(bold=True, color=WHITE, name="Calibri", size=12)
    c.fill      = fill("22C55E")
    c.alignment = center()
    ws5.row_dimensions[1].height = 24

    write_hdr(ws5, 3, 1, "Priorité",  "22C55E")
    write_hdr(ws5, 3, 2, "Action",    "22C55E")
    write_hdr(ws5, 3, 3, "Objectif",  "22C55E")
    write_hdr(ws5, 3, 4, "Impact",    "22C55E")
    write_hdr(ws5, 3, 5, "Statut",    "22C55E")

    decisions = []
    if data["nb_anomalies"] > 0:
        decisions.append(("🔴 Immédiat", "Corriger anomalies compteur", "Intégrité des données", "Élevé", "À faire"))
    if data["aujourd_hui"]["nb_releves"] == 0:
        decisions.append(("🔴 Immédiat", "Saisir relevés du jour", "Couverture 100%", "Élevé", "À faire"))
    if data["par_periode"]["Apres-midi"]["montant"] == 0:
        decisions.append(("🟡 Cette semaine", "Clarifier saisie Après-midi", "Récupérer CA manquant", "Très élevé", "À investiguer"))
    if data["mois_courant"]["jours_actifs"] < 10:
        decisions.append(("🟡 Cette semaine", "Reconstituer historique", "Analyses fiables", "Élevé", "En cours"))
    decisions.append(("🟢 Ce mois", "Former opérateurs", "Saisie quotidienne 100%", "Moyen", "À planifier"))
    decisions.append(("🟢 Ce mois", "Vérifier pompes sous-utilisées", "Optimiser capacité", "Moyen", "À analyser"))

    row = 4
    for i, (pri, act, obj, imp, stat) in enumerate(decisions):
        bg = LGRAY if i % 2 == 0 else WHITE
        write_cell(ws5, row, 1, pri,  bold=True, bg=bg)
        write_cell(ws5, row, 2, act,  bg=bg)
        write_cell(ws5, row, 3, obj,  bg=bg)
        write_cell(ws5, row, 4, imp,  bold=True, bg=bg)
        write_cell(ws5, row, 5, stat, bg=bg)
        row += 1

    for ci, w in enumerate([18, 30, 28, 14, 16], 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w

    # Retour
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Rapport_Ventes_{d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ══════════════════════════════════════════════════════════════════
# PRÉVISION DES VENTES
# ══════════════════════════════════════════════════════════════════

@app.get("/api/forecast")
def forecast_endpoint(
    horizon:    int            = 14,
    metric:     str            = "montant",
    produit_id: Optional[int]  = None,
    pompe_id:   Optional[int]  = None,
    objectif:   Optional[float]= None,
    db: Session = Depends(get_db),
):
    """
    Prévision statistique des ventes.

    Paramètres :
      horizon    : jours à prévoir (1–90, défaut 14)
      metric     : "montant" (CA en G) | "quantite" (gallons)
      produit_id : filtrer par produit
      pompe_id   : filtrer par pompe
      objectif   : seuil pour calcul P(CA > objectif)
    """
    from forecasting import run_forecast
    horizon = max(1, min(90, horizon))
    return run_forecast(db, horizon, metric, produit_id, pompe_id, objectif)


# ══════════════════════════════════════════════════════════════════════════
# MODULE STOCK & RENTABILITÉ
# ══════════════════════════════════════════════════════════════════════════
from stock_service import (
    gallons_vendus,
    gallons_livres,
    stock_restant,
    cout_moyen_pondere,
    rentabilite_globale,
    anomalies_stock,
    corr_saut_decalage,
    SEUIL_ALERTE_JOURS_PAR_DEFAUT,
)


# ---------- Schemas stock ----------
class LivraisonIn(BaseModel):
    produit_id:        int
    date_livraison:    str           # YYYY-MM-DD
    gallons_recus:     float
    prix_achat_gallon: float
    fournisseur:       Optional[str] = None
    reference_camion:  Optional[str] = None
    notes:             Optional[str] = None


class PrixVenteIn(BaseModel):
    produit_id:        int
    prix_vente_gallon: float
    date_effet:        str           # YYYY-MM-DD


# ---------- Livraisons ----------
@app.post("/api/livraisons", status_code=201)
def create_livraison(payload: LivraisonIn, db: Session = Depends(get_db)):
    try:
        d = date_type.fromisoformat(payload.date_livraison)
    except ValueError:
        raise HTTPException(400, "date_livraison invalide (format YYYY-MM-DD)")
    if payload.gallons_recus <= 0:
        raise HTTPException(400, "gallons_recus doit être > 0")
    if payload.prix_achat_gallon < 0:
        raise HTTPException(400, "prix_achat_gallon doit être >= 0")
    produit = db.query(Produit).filter(Produit.id == payload.produit_id).first()
    if not produit:
        raise HTTPException(404, "Produit introuvable")

    lv = Livraison(
        produit_id        = payload.produit_id,
        date_livraison    = d,
        gallons_recus     = payload.gallons_recus,
        prix_achat_gallon = payload.prix_achat_gallon,
        fournisseur       = payload.fournisseur,
        reference_camion  = payload.reference_camion,
        notes             = payload.notes,
    )
    try:
        db.add(lv)
        db.commit()
        db.refresh(lv)
    except Exception:
        db.rollback()
        raise
    return {
        "id":                lv.id,
        "produit_id":        lv.produit_id,
        "produit_nom":       produit.nom,
        "date_livraison":    str(lv.date_livraison),
        "gallons_recus":     float(lv.gallons_recus),
        "prix_achat_gallon": float(lv.prix_achat_gallon),
        "fournisseur":       lv.fournisseur,
        "reference_camion":  lv.reference_camion,
        "notes":             lv.notes,
    }


@app.get("/api/livraisons")
def list_livraisons(
    produit_id: Optional[int]  = None,
    date_debut: Optional[str]  = None,
    date_fin:   Optional[str]  = None,
    db: Session = Depends(get_db),
):
    q = db.query(Livraison)
    if produit_id:
        q = q.filter(Livraison.produit_id == produit_id)
    if date_debut:
        try:
            q = q.filter(Livraison.date_livraison >= date_type.fromisoformat(date_debut))
        except ValueError:
            raise HTTPException(400, "date_debut invalide")
    if date_fin:
        try:
            q = q.filter(Livraison.date_livraison <= date_type.fromisoformat(date_fin))
        except ValueError:
            raise HTTPException(400, "date_fin invalide")
    livraisons = q.order_by(Livraison.date_livraison.desc(), Livraison.id.desc()).all()
    return {
        "nb": len(livraisons),
        "livraisons": [
            {
                "id":                l.id,
                "produit_id":        l.produit_id,
                "produit_nom":       l.produit.nom,
                "date_livraison":    str(l.date_livraison),
                "gallons_recus":     float(l.gallons_recus),
                "prix_achat_gallon": float(l.prix_achat_gallon),
                "fournisseur":       l.fournisseur,
                "reference_camion":  l.reference_camion,
                "notes":             l.notes,
                "created_at":        str(l.created_at),
            }
            for l in livraisons
        ],
    }


@app.delete("/api/livraisons/{livraison_id}", status_code=200)
def delete_livraison(livraison_id: int, db: Session = Depends(get_db)):
    lv = db.query(Livraison).filter(Livraison.id == livraison_id).first()
    if not lv:
        raise HTTPException(404, "Livraison introuvable")
    try:
        db.delete(lv)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"detail": "Livraison supprimée"}


# ---------- Prix de vente ----------
@app.post("/api/prix-vente", status_code=201)
def create_prix_vente(payload: PrixVenteIn, db: Session = Depends(get_db)):
    try:
        d = date_type.fromisoformat(payload.date_effet)
    except ValueError:
        raise HTTPException(400, "date_effet invalide (format YYYY-MM-DD)")
    if payload.prix_vente_gallon <= 0:
        raise HTTPException(400, "prix_vente_gallon doit être > 0")
    produit = db.query(Produit).filter(Produit.id == payload.produit_id).first()
    if not produit:
        raise HTTPException(404, "Produit introuvable")

    pv = PrixVente(
        produit_id        = payload.produit_id,
        prix_vente_gallon = payload.prix_vente_gallon,
        date_effet        = d,
    )
    try:
        db.add(pv)
        db.commit()
        db.refresh(pv)
    except Exception:
        db.rollback()
        raise
    return {
        "id":                pv.id,
        "produit_id":        pv.produit_id,
        "produit_nom":       produit.nom,
        "prix_vente_gallon": float(pv.prix_vente_gallon),
        "date_effet":        str(pv.date_effet),
    }


@app.get("/api/prix-vente")
def list_prix_vente(
    produit_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = db.query(PrixVente)
    if produit_id:
        q = q.filter(PrixVente.produit_id == produit_id)
    prix = q.order_by(PrixVente.produit_id, PrixVente.date_effet.desc()).all()
    return {
        "nb": len(prix),
        "prix_vente": [
            {
                "id":                p.id,
                "produit_id":        p.produit_id,
                "produit_nom":       p.produit.nom,
                "prix_vente_gallon": float(p.prix_vente_gallon),
                "date_effet":        str(p.date_effet),
                "created_at":        str(p.created_at),
            }
            for p in prix
        ],
    }


# ---------- Stock ----------
@app.get("/api/stock")
def stock_endpoint(
    seuil_jours: int           = SEUIL_ALERTE_JOURS_PAR_DEFAUT,
    produit_id:  Optional[int] = None,
    db: Session = Depends(get_db),
):
    """
    Stock restant par produit = Σ(gallons livrés) − Σ(gallons vendus relevés).
    Enrichi des anomalies de cohérence (STOCK_NEGATIF, VENTE_SANS_STOCK, etc.).
    """
    if produit_id:
        produits = db.query(Produit).filter(Produit.id == produit_id, Produit.actif == True).all()
    else:
        produits = db.query(Produit).filter(Produit.actif == True).all()

    aujourd_hui = date_type.today()
    resultats = []
    for p in produits:
        s = stock_restant(db, p.id, seuil_jours=seuil_jours)
        s["produit_nom"] = p.nom
        cmp = cout_moyen_pondere(db, p.id)
        s["cout_moyen_pondere"] = cmp
        s["valeur_stock_gourdes"] = round(float(s["gallons_restants"]) * cmp, 2) if cmp and s["gallons_restants"] > 0 else None
        resultats.append(s)

    nb_alertes = sum(1 for r in resultats if r["alerte_bas"])
    return {
        "date":        str(aujourd_hui),
        "seuil_jours": seuil_jours,
        "nb_alertes":  nb_alertes,
        "stocks":      resultats,
    }


# ---------- Rentabilité ----------
@app.get("/api/rentabilite")
def rentabilite_endpoint(
    date_debut: Optional[str] = None,
    date_fin:   Optional[str] = None,
    produit_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """
    Bénéfice = Revenu (relevés.prix_gallon × quantité) − COGS (WAC × gallons vendus).
    Retourne None si aucune livraison enregistrée (pas de WAC disponible).
    """
    aujourd_hui = date_type.today()
    try:
        d_debut = date_type.fromisoformat(date_debut) if date_debut else date_type(aujourd_hui.year, aujourd_hui.month, 1)
        d_fin   = date_type.fromisoformat(date_fin)   if date_fin   else aujourd_hui
    except ValueError:
        raise HTTPException(400, "Format de date invalide (YYYY-MM-DD)")
    if d_debut > d_fin:
        raise HTTPException(400, "date_debut doit être <= date_fin")

    return rentabilite_globale(db, d_debut, d_fin, produit_id)


# ---------- Rapport complet multi-période ----------
@app.get("/api/rapport/export")
def rapport_export(
    date_debut: str,
    date_fin: str,
    format: str = "pdf",
    produit_id: Optional[int] = None,
    request: Request = None,
    db: Session = Depends(get_db),
):
    """
    Génère un rapport professionnel téléchargeable sur une période.
    Formats acceptés : pdf, docx, xlsx
    """
    from rapport_service import build_report_payload, build_narrative, build_charts
    from rapport_renderers import render_pdf, render_docx, render_xlsx

    try:
        d_debut = date_type.fromisoformat(date_debut)
        d_fin   = date_type.fromisoformat(date_fin)
    except ValueError:
        raise HTTPException(400, "Format de date invalide (YYYY-MM-DD)")

    if d_debut > d_fin:
        raise HTTPException(400, "date_debut doit être <= date_fin")

    fmt = format.lower().strip(".")
    if fmt not in ("pdf", "docx", "xlsx"):
        raise HTTPException(400, "Format invalide — choisir parmi : pdf, docx, xlsx")

    payload   = build_report_payload(db, d_debut, d_fin, produit_id)
    narrative = build_narrative(payload)
    charts    = build_charts(payload)

    filename = f"rapport_station_{date_debut}_{date_fin}.{fmt}"

    if fmt == "pdf":
        data = render_pdf(payload, narrative, charts)
        media = "application/pdf"
    elif fmt == "docx":
        data = render_docx(payload, narrative, charts)
        media = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    else:  # xlsx
        data = render_xlsx(payload, narrative, charts)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        iter([data]),
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------- Audit IA complet ----------
@app.get("/api/audit/export")
def audit_export(
    date_debut: str,
    date_fin: str,
    format: str = "pdf",
    produit_id: Optional[int] = None,
    request: Request = None,
    db: Session = Depends(get_db),
):
    """
    Génère un document d'audit narratif complet rédigé par l'IA.
    Formats : pdf, docx
    L'IA analyse les données réelles et produit un texte d'expert structuré.
    """
    from audit_service   import generate_audit
    from audit_renderer  import render_audit_pdf, render_audit_docx

    try:
        d_debut = date_type.fromisoformat(date_debut)
        d_fin   = date_type.fromisoformat(date_fin)
    except ValueError:
        raise HTTPException(400, "Format de date invalide (YYYY-MM-DD)")

    if d_debut > d_fin:
        raise HTTPException(400, "date_debut doit être <= date_fin")

    fmt = format.lower().strip(".")
    if fmt not in ("pdf", "docx"):
        raise HTTPException(400, "Format invalide — choisir parmi : pdf, docx")

    try:
        result = generate_audit(db, d_debut, d_fin, produit_id)
    except RuntimeError as e:
        raise HTTPException(503, str(e))

    text     = result["text"]
    kpis     = result.get("kpis", {})
    filename = f"audit_station_{date_debut}_{date_fin}.{fmt}"

    if fmt == "pdf":
        data  = render_audit_pdf(text, d_debut, d_fin, kpi_data=kpis)
        media = "application/pdf"
    else:
        data  = render_audit_docx(text, d_debut, d_fin, kpi_data=kpis)
        media = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"

    return StreamingResponse(
        iter([data]),
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ══════════════════════════════════════════════════════════════════
# MODULE EMPLOYÉS
# ══════════════════════════════════════════════════════════════════

class EmployeIn(BaseModel):
    nom:           str
    prenom:        str
    poste:         str
    date_embauche: str
    salaire_base:  float
    type_contrat:  str = "CDI"
    telephone:     Optional[str] = None
    email:         Optional[str] = None
    notes:         Optional[str] = None

class EmployePatch(BaseModel):
    nom:           Optional[str]   = None
    prenom:        Optional[str]   = None
    poste:         Optional[str]   = None
    date_embauche: Optional[str]   = None
    salaire_base:  Optional[float] = None
    type_contrat:  Optional[str]   = None
    telephone:     Optional[str]   = None
    email:         Optional[str]   = None
    actif:         Optional[bool]  = None
    notes:         Optional[str]   = None

_CONTRATS_VALIDES = {"CDI", "CDD", "Temps partiel", "Journalier", "Stage"}

@app.get("/api/employes")
def lister_employes(actif: Optional[bool] = None, db: Session = Depends(get_db)):
    q = db.query(Employe)
    if actif is not None:
        q = q.filter(Employe.actif == actif)
    employes = q.order_by(Employe.nom, Employe.prenom).all()
    return [
        {
            "id": e.id, "nom": e.nom, "prenom": e.prenom,
            "nom_complet": f"{e.prenom} {e.nom}",
            "poste": e.poste, "date_embauche": str(e.date_embauche),
            "salaire_base": float(e.salaire_base),
            "type_contrat": e.type_contrat,
            "telephone": e.telephone, "email": e.email,
            "actif": e.actif, "notes": e.notes,
            "created_at": e.created_at.isoformat() if e.created_at else None,
        }
        for e in employes
    ]

@app.post("/api/employes", status_code=201)
def creer_employe(data: EmployeIn, db: Session = Depends(get_db)):
    if data.type_contrat not in _CONTRATS_VALIDES:
        raise HTTPException(400, f"Type de contrat invalide. Valeurs : {sorted(_CONTRATS_VALIDES)}")
    if data.salaire_base < 0:
        raise HTTPException(400, "Le salaire de base doit être ≥ 0.")
    try:
        from datetime import date as _date
        date_emb = _date.fromisoformat(data.date_embauche)
    except ValueError:
        raise HTTPException(400, "Format de date invalide (attendu YYYY-MM-DD).")
    e = Employe(
        nom=data.nom.strip(), prenom=data.prenom.strip(),
        poste=data.poste.strip(), date_embauche=date_emb,
        salaire_base=data.salaire_base, type_contrat=data.type_contrat,
        telephone=data.telephone, email=data.email, notes=data.notes,
    )
    db.add(e)
    db.commit()
    db.refresh(e)
    return {"id": e.id, "message": "Employé créé."}

@app.put("/api/employes/{employe_id}")
def modifier_employe(employe_id: int, data: EmployePatch, db: Session = Depends(get_db)):
    e = db.query(Employe).filter(Employe.id == employe_id).first()
    if not e:
        raise HTTPException(404, "Employé introuvable.")
    if data.nom          is not None: e.nom          = data.nom.strip()
    if data.prenom       is not None: e.prenom       = data.prenom.strip()
    if data.poste        is not None: e.poste        = data.poste.strip()
    if data.salaire_base is not None:
        if data.salaire_base < 0:
            raise HTTPException(400, "Le salaire doit être ≥ 0.")
        e.salaire_base = data.salaire_base
    if data.type_contrat is not None:
        if data.type_contrat not in _CONTRATS_VALIDES:
            raise HTTPException(400, "Type de contrat invalide.")
        e.type_contrat = data.type_contrat
    if data.date_embauche is not None:
        try:
            from datetime import date as _date
            e.date_embauche = _date.fromisoformat(data.date_embauche)
        except ValueError:
            raise HTTPException(400, "Format de date invalide.")
    if data.telephone is not None: e.telephone = data.telephone
    if data.email     is not None: e.email     = data.email
    if data.actif     is not None: e.actif     = data.actif
    if data.notes     is not None: e.notes     = data.notes
    db.commit()
    return {"message": "Employé mis à jour."}

@app.delete("/api/employes/{employe_id}")
def desactiver_employe(employe_id: int, db: Session = Depends(get_db)):
    e = db.query(Employe).filter(Employe.id == employe_id).first()
    if not e:
        raise HTTPException(404, "Employé introuvable.")
    e.actif = False
    db.commit()
    return {"message": "Employé désactivé."}


# ══════════════════════════════════════════════════════════════════
# MODULE PAYROLL (FICHES DE PAIE)
# ══════════════════════════════════════════════════════════════════

class FichePaieIn(BaseModel):
    employe_id:    int
    periode_debut: str
    periode_fin:   str
    salaire_base:  float
    heures_sup:    float = 0.0
    taux_hs:       float = 0.0
    primes:        float = 0.0
    deductions:    float = 0.0
    notes:         Optional[str] = None

class PayerFicheIn(BaseModel):
    date_paiement: str

def _calc_net(salaire_base: float, heures_sup: float, taux_hs: float,
              primes: float, deductions: float) -> float:
    return round(salaire_base + (heures_sup * taux_hs) + primes - deductions, 2)

@app.get("/api/payroll")
def lister_fiches(
    employe_id: Optional[int] = None,
    statut: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(FichePaie)
    if employe_id: q = q.filter(FichePaie.employe_id == employe_id)
    if statut:     q = q.filter(FichePaie.statut == statut)
    fiches = q.order_by(FichePaie.periode_debut.desc()).all()
    return [
        {
            "id": f.id,
            "employe_id":    f.employe_id,
            "employe_nom":   f"{f.employe.prenom} {f.employe.nom}",
            "employe_poste": f.employe.poste,
            "periode_debut": str(f.periode_debut),
            "periode_fin":   str(f.periode_fin),
            "salaire_base":  float(f.salaire_base),
            "heures_sup":    float(f.heures_sup),
            "taux_hs":       float(f.taux_hs),
            "primes":        float(f.primes),
            "deductions":    float(f.deductions),
            "net_a_payer":   float(f.net_a_payer),
            "statut":        f.statut,
            "date_paiement": str(f.date_paiement) if f.date_paiement else None,
            "notes":         f.notes,
            "created_at":    f.created_at.isoformat() if f.created_at else None,
        }
        for f in fiches
    ]

@app.post("/api/payroll", status_code=201)
def creer_fiche(data: FichePaieIn, db: Session = Depends(get_db)):
    e = db.query(Employe).filter(Employe.id == data.employe_id, Employe.actif == True).first()
    if not e:
        raise HTTPException(404, "Employé introuvable ou inactif.")
    for val, label in [(data.salaire_base, "salaire_base"), (data.heures_sup, "heures_sup"),
                       (data.taux_hs, "taux_hs"), (data.primes, "primes"), (data.deductions, "deductions")]:
        if val < 0:
            raise HTTPException(400, f"Le champ {label} doit être ≥ 0.")
    try:
        from datetime import date as _date
        pd = _date.fromisoformat(data.periode_debut)
        pf = _date.fromisoformat(data.periode_fin)
    except ValueError:
        raise HTTPException(400, "Format de date invalide (attendu YYYY-MM-DD).")
    if pf < pd:
        raise HTTPException(400, "La date de fin doit être ≥ à la date de début.")
    net = _calc_net(data.salaire_base, data.heures_sup, data.taux_hs, data.primes, data.deductions)
    if net < 0:
        raise HTTPException(400, "Le net à payer est négatif — vérifiez les déductions.")
    f = FichePaie(
        employe_id=data.employe_id, periode_debut=pd, periode_fin=pf,
        salaire_base=data.salaire_base, heures_sup=data.heures_sup,
        taux_hs=data.taux_hs, primes=data.primes, deductions=data.deductions,
        net_a_payer=net, notes=data.notes,
    )
    db.add(f)
    db.commit()
    db.refresh(f)
    return {"id": f.id, "net_a_payer": net, "message": "Fiche de paie créée."}

@app.put("/api/payroll/{fiche_id}/payer")
def marquer_payee(fiche_id: int, data: PayerFicheIn, db: Session = Depends(get_db)):
    f = db.query(FichePaie).filter(FichePaie.id == fiche_id).first()
    if not f:
        raise HTTPException(404, "Fiche introuvable.")
    if f.statut == "paye":
        raise HTTPException(409, "Fiche déjà marquée comme payée.")
    try:
        from datetime import date as _date
        f.date_paiement = _date.fromisoformat(data.date_paiement)
    except ValueError:
        raise HTTPException(400, "Format de date invalide.")
    f.statut = "paye"
    db.commit()
    return {"message": "Fiche marquée comme payée."}

@app.delete("/api/payroll/{fiche_id}")
def supprimer_fiche(fiche_id: int, db: Session = Depends(get_db)):
    f = db.query(FichePaie).filter(FichePaie.id == fiche_id).first()
    if not f:
        raise HTTPException(404, "Fiche introuvable.")
    if f.statut == "paye":
        raise HTTPException(409, "Impossible de supprimer une fiche déjà payée.")
    db.delete(f)
    db.commit()
    return {"message": "Fiche supprimée."}

@app.get("/api/payroll/stats")
def stats_payroll(db: Session = Depends(get_db)):
    fiches = db.query(FichePaie).all()
    total_paye    = sum(float(f.net_a_payer) for f in fiches if f.statut == "paye")
    total_pending = sum(float(f.net_a_payer) for f in fiches if f.statut == "brouillon")
    nb_employes   = db.query(Employe).filter(Employe.actif == True).count()
    return {
        "nb_employes_actifs": nb_employes,
        "total_paye_htg":     round(total_paye, 2),
        "total_en_attente_htg": round(total_pending, 2),
        "nb_fiches_brouillon": sum(1 for f in fiches if f.statut == "brouillon"),
        "nb_fiches_payees":    sum(1 for f in fiches if f.statut == "paye"),
    }


# ══════════════════════════════════════════════════════════════════
# MODULE DÉPENSES
# ══════════════════════════════════════════════════════════════════

_CATEGORIES_DEPENSE = {
    "Salaires", "Maintenance", "Fournitures", "Electricite",
    "Eau", "Loyer", "Transport", "Taxes", "Assurance", "Divers",
}

# ── Helpers limite ────────────────────────────────────────────────

def _get_param_depense(db: Session) -> ParametreDepense:
    """Retourne le singleton ParametreDepense (crée si absent)."""
    p = db.query(ParametreDepense).filter(ParametreDepense.id == 1).first()
    if not p:
        p = ParametreDepense(id=1, limite=None, active=True)
        db.add(p)
        db.commit()
        db.refresh(p)
    return p

def _total_depenses_mois_courant(db: Session, exclure_id: Optional[int] = None) -> float:
    """Somme des dépenses du mois calendaire en cours."""
    from datetime import date as _date
    today = _date.today()
    debut_mois = today.replace(day=1)
    q = db.query(Depense).filter(
        Depense.date_depense >= debut_mois,
        Depense.date_depense <= today,
    )
    if exclure_id:
        q = q.filter(Depense.id != exclure_id)
    return round(sum(float(d.montant) for d in q.all()), 2)

def _verifier_limite(db: Session, montant_nouveau: float, exclure_id: Optional[int] = None):
    """Lève 403 si l'ajout dépasserait la limite mensuelle active."""
    p = _get_param_depense(db)
    if not p.active or p.limite is None:
        return
    total = _total_depenses_mois_courant(db, exclure_id)
    if total + montant_nouveau > float(p.limite):
        disponible = max(0.0, float(p.limite) - total)
        raise HTTPException(
            403,
            f"Limite mensuelle atteinte ({float(p.limite):,.2f} G). "
            f"Dépenses ce mois : {total:,.2f} G. "
            f"Montant disponible : {disponible:,.2f} G."
        )

# ── Endpoints limite (PDG only) ───────────────────────────────────

@app.get("/api/depenses/limite")
def get_limite_depenses(db: Session = Depends(get_db)):
    """Retourne la configuration de la limite mensuelle."""
    from datetime import date as _date
    p = _get_param_depense(db)
    today = _date.today()
    total_mois = _total_depenses_mois_courant(db)
    return {
        "limite":      float(p.limite) if p.limite is not None else None,
        "active":      p.active,
        "updated_at":  p.updated_at.isoformat() if p.updated_at else None,
        "updated_by":  p.updated_by,
        "total_mois":  total_mois,
        "mois":        today.strftime("%B %Y"),
        "pct":         round(total_mois / float(p.limite) * 100, 1) if p.limite else None,
    }

class LimiteDepenseIn(BaseModel):
    limite: Optional[float] = None   # None = supprimer la limite
    active: bool = True

@app.put("/api/depenses/limite")
def set_limite_depenses(
    data: LimiteDepenseIn,
    request: Request,
    _pdg: Utilisateur = Depends(require_pdg),
    db: Session = Depends(get_db),
):
    """Définit ou supprime la limite mensuelle. PDG uniquement."""
    from datetime import datetime as _dt
    if data.limite is not None and data.limite <= 0:
        raise HTTPException(400, "La limite doit être > 0.")
    p = _get_param_depense(db)
    p.limite     = data.limite
    p.active     = data.active
    p.updated_at = _dt.utcnow()
    p.updated_by = _pdg.username
    db.commit()
    return {"ok": True, "limite": float(p.limite) if p.limite else None, "active": p.active}

# ──────────────────────────────────────────────────────────────────

class DepenseIn(BaseModel):
    categorie:    str
    description:  str
    montant:      float = Field(gt=0)
    date_depense: str
    beneficiaire: Optional[str] = None
    reference:    Optional[str] = None
    notes:        Optional[str] = None

class DepensePatch(BaseModel):
    categorie:    Optional[str]   = None
    description:  Optional[str]   = None
    montant:      Optional[float] = Field(None, gt=0)
    date_depense: Optional[str]   = None
    beneficiaire: Optional[str]   = None
    reference:    Optional[str]   = None
    notes:        Optional[str]   = None

@app.get("/api/depenses")
def lister_depenses(
    date_debut: Optional[str] = None,
    date_fin:   Optional[str] = None,
    categorie:  Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(Depense)
    if date_debut:
        try:
            from datetime import date as _date
            q = q.filter(Depense.date_depense >= _date.fromisoformat(date_debut))
        except ValueError:
            raise HTTPException(400, "date_debut invalide.")
    if date_fin:
        try:
            from datetime import date as _date
            q = q.filter(Depense.date_depense <= _date.fromisoformat(date_fin))
        except ValueError:
            raise HTTPException(400, "date_fin invalide.")
    if categorie:
        q = q.filter(Depense.categorie == categorie)
    depenses = q.order_by(Depense.date_depense.desc()).all()
    total = round(sum(float(d.montant) for d in depenses), 2)
    return {
        "total": total,
        "nb": len(depenses),
        "depenses": [
            {
                "id": d.id, "categorie": d.categorie,
                "description": d.description, "montant": float(d.montant),
                "date_depense": str(d.date_depense),
                "beneficiaire": d.beneficiaire, "reference": d.reference,
                "notes": d.notes,
                "created_at": d.created_at.isoformat() if d.created_at else None,
            }
            for d in depenses
        ],
    }

@app.post("/api/depenses", status_code=201)
def creer_depense(data: DepenseIn, db: Session = Depends(get_db)):
    if data.categorie not in _CATEGORIES_DEPENSE:
        raise HTTPException(400, f"Catégorie invalide. Valeurs : {sorted(_CATEGORIES_DEPENSE)}")
    if data.montant <= 0:
        raise HTTPException(400, "Le montant doit être > 0.")
    try:
        from datetime import date as _date
        date_d = _date.fromisoformat(data.date_depense)
    except ValueError:
        raise HTTPException(400, "Format de date invalide.")
    _verifier_limite(db, data.montant)
    d = Depense(
        categorie=data.categorie, description=data.description.strip(),
        montant=data.montant, date_depense=date_d,
        beneficiaire=data.beneficiaire, reference=data.reference, notes=data.notes,
    )
    db.add(d)
    db.commit()
    db.refresh(d)
    return {"id": d.id, "message": "Dépense enregistrée."}

@app.put("/api/depenses/{depense_id}")
def modifier_depense(depense_id: int, data: DepensePatch, db: Session = Depends(get_db)):
    d = db.query(Depense).filter(Depense.id == depense_id).first()
    if not d:
        raise HTTPException(404, "Dépense introuvable.")
    if data.categorie is not None:
        if data.categorie not in _CATEGORIES_DEPENSE:
            raise HTTPException(400, "Catégorie invalide.")
        d.categorie = data.categorie
    if data.description  is not None: d.description  = data.description.strip()
    if data.montant is not None:
        if data.montant <= 0:
            raise HTTPException(400, "Le montant doit être > 0.")
        _verifier_limite(db, data.montant, exclure_id=depense_id)
        d.montant = data.montant
    if data.date_depense is not None:
        try:
            from datetime import date as _date
            d.date_depense = _date.fromisoformat(data.date_depense)
        except ValueError:
            raise HTTPException(400, "Format de date invalide.")
    if data.beneficiaire is not None: d.beneficiaire = data.beneficiaire
    if data.reference    is not None: d.reference    = data.reference
    if data.notes        is not None: d.notes        = data.notes
    db.commit()
    return {"message": "Dépense mise à jour."}

@app.delete("/api/depenses/{depense_id}")
def supprimer_depense(depense_id: int, db: Session = Depends(get_db)):
    d = db.query(Depense).filter(Depense.id == depense_id).first()
    if not d:
        raise HTTPException(404, "Dépense introuvable.")
    db.delete(d)
    db.commit()
    return {"message": "Dépense supprimée."}

@app.get("/api/depenses/stats")
def stats_depenses(
    date_debut: Optional[str] = None,
    date_fin:   Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(Depense)
    if date_debut:
        try:
            from datetime import date as _date
            q = q.filter(Depense.date_depense >= _date.fromisoformat(date_debut))
        except ValueError:
            raise HTTPException(400, "date_debut invalide.")
    if date_fin:
        try:
            from datetime import date as _date
            q = q.filter(Depense.date_depense <= _date.fromisoformat(date_fin))
        except ValueError:
            raise HTTPException(400, "date_fin invalide.")
    depenses = q.all()
    par_cat: dict[str, float] = {}
    for d in depenses:
        par_cat[d.categorie] = round(par_cat.get(d.categorie, 0) + float(d.montant), 2)
    return {
        "total": round(sum(par_cat.values()), 2),
        "nb": len(depenses),
        "par_categorie": [
            {"categorie": k, "total": v, "pct": round(v / sum(par_cat.values()) * 100, 1) if par_cat else 0}
            for k, v in sorted(par_cat.items(), key=lambda x: -x[1])
        ],
    }


# ══════════════════════════════════════════════════════════════════
# EXPORTS DÉPENSES — PDF + XLSX
# ══════════════════════════════════════════════════════════════════

def _query_depenses_filtrees(db, date_debut, date_fin, categorie):
    """Requête dépenses avec filtres optionnels, triées par date desc."""
    from datetime import date as _date
    q = db.query(Depense)
    if date_debut:
        try:    q = q.filter(Depense.date_depense >= _date.fromisoformat(date_debut))
        except ValueError: pass
    if date_fin:
        try:    q = q.filter(Depense.date_depense <= _date.fromisoformat(date_fin))
        except ValueError: pass
    if categorie:
        q = q.filter(Depense.categorie == categorie)
    return q.order_by(Depense.date_depense.desc()).all()


@app.get("/api/depenses/export/xlsx")
def export_depenses_xlsx(
    date_debut: Optional[str] = None,
    date_fin:   Optional[str] = None,
    categorie:  Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Export Excel des dépenses avec feuille détail + feuille récapitulatif par catégorie."""
    import io
    from datetime import date as date_cls
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    depenses = _query_depenses_filtrees(db, date_debut, date_fin, categorie)

    # ── Styles ──────────────────────────────────────────────────
    def _fill(hex_c):
        return PatternFill("solid", fgColor=hex_c)
    def _font(bold=False, color="1E293B", size=10, italic=False):
        return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
    def _al(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def _bd():
        s = Side(style="thin", color="CBD5E1")
        return Border(left=s, right=s, top=s, bottom=s)

    HDR_BG   = _fill("F87171")   # rouge dépenses
    HDR_FG   = _font(bold=True, color="FFFFFF", size=11)
    TOT_BG   = _fill("FEE2E2")
    TOT_FG   = _font(bold=True, color="7F1D1D", size=11)
    ODD_BG   = _fill("FFF8F8")
    EVN_BG   = _fill("FFFFFF")
    CAT_BG   = _fill("FEF2F2")

    MONEY_FMT = "#,##0.00"
    DATE_FMT  = "DD/MM/YYYY"

    # ══════════════════════════════════════════════════════════
    # Feuille 1 — Détail des dépenses
    # ══════════════════════════════════════════════════════════
    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Dépenses"

    # Titre
    ws1.merge_cells("A1:H1")
    c = ws1.cell(1, 1, "PétroSync — Rapport des Dépenses Opérationnelles")
    c.font      = Font(name="Calibri", bold=True, size=14, color="DC2626")
    c.alignment = _al("center")
    ws1.row_dimensions[1].height = 24

    # Sous-titre
    periode_txt = ""
    if date_debut or date_fin:
        periode_txt = f"  ·  Période : {date_debut or '…'} → {date_fin or '…'}"
    if categorie:
        periode_txt += f"  ·  Catégorie : {categorie}"
    info = f"Exporté le {date_cls.today().strftime('%d/%m/%Y')}  ·  {len(depenses)} dépense(s){periode_txt}"
    ws1.merge_cells("A2:H2")
    c2 = ws1.cell(2, 1, info)
    c2.font      = Font(name="Calibri", size=9, color="94A3B8", italic=True)
    c2.alignment = _al("center")
    ws1.row_dimensions[2].height = 14

    ws1.row_dimensions[3].height = 6  # spacer

    # En-têtes colonnes
    headers = ["#", "Date", "Catégorie", "Description", "Bénéficiaire", "Référence", "Montant (HTG)", "Notes"]
    col_widths = [5, 14, 18, 40, 24, 18, 16, 30]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws1.cell(4, ci, h)
        c.font      = HDR_FG
        c.fill      = HDR_BG
        c.alignment = _al("center" if ci in (1, 7) else "left")
        c.border    = _bd()
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[4].height = 18

    # Lignes de données
    total = 0.0
    par_cat = defaultdict(float)
    for ri, d in enumerate(depenses, 5):
        bg = ODD_BG if ri % 2 == 1 else EVN_BG
        vals = [ri - 4, d.date_depense, d.categorie, d.description,
                d.beneficiaire or "", d.reference or "", float(d.montant), d.notes or ""]
        for ci, val in enumerate(vals, 1):
            c = ws1.cell(ri, ci, val)
            c.fill      = bg
            c.border    = _bd()
            c.alignment = _al("center" if ci == 1 else ("right" if ci == 7 else "left"),
                               wrap=ci in (4, 8))
            c.font      = _font()
            if ci == 7:
                c.number_format = MONEY_FMT
            if ci == 2 and isinstance(val, __import__('datetime').date):
                c.number_format = DATE_FMT
        ws1.row_dimensions[ri].height = 16
        total += float(d.montant)
        par_cat[d.categorie] += float(d.montant)

    # Ligne total
    row_tot = len(depenses) + 5
    ws1.merge_cells(f"A{row_tot}:F{row_tot}")
    c = ws1.cell(row_tot, 1, "TOTAL")
    c.font = TOT_FG;  c.fill = TOT_BG;  c.alignment = _al("right");  c.border = _bd()
    ct = ws1.cell(row_tot, 7, round(total, 2))
    ct.font = TOT_FG;  ct.fill = TOT_BG;  ct.alignment = _al("right")
    ct.border = _bd();  ct.number_format = MONEY_FMT
    ws1.cell(row_tot, 8).fill = TOT_BG
    ws1.row_dimensions[row_tot].height = 18

    # Figer la ligne d'en-têtes
    ws1.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════════
    # Feuille 2 — Récapitulatif par catégorie
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Par Catégorie")

    ws2.merge_cells("A1:D1")
    c = ws2.cell(1, 1, "Récapitulatif par catégorie")
    c.font = Font(name="Calibri", bold=True, size=13, color="DC2626")
    c.alignment = _al("center")
    ws2.row_dimensions[1].height = 22

    ws2.row_dimensions[2].height = 6
    hdr2 = ["Catégorie", "Nb dépenses", "Montant total (HTG)", "Part (%)"]
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 12
    for ci, h in enumerate(hdr2, 1):
        c = ws2.cell(3, ci, h)
        c.font = HDR_FG;  c.fill = HDR_BG
        c.alignment = _al("center" if ci > 1 else "left")
        c.border = _bd()
    ws2.row_dimensions[3].height = 18

    cats_sorted = sorted(par_cat.items(), key=lambda x: -x[1])
    grand = sum(par_cat.values()) or 1
    nb_by_cat = defaultdict(int)
    for d in depenses:
        nb_by_cat[d.categorie] += 1

    for ri, (cat, mnt) in enumerate(cats_sorted, 4):
        pct = round(mnt / grand * 100, 1)
        bg = ODD_BG if ri % 2 == 1 else EVN_BG
        row_data = [cat, nb_by_cat[cat], round(mnt, 2), pct]
        for ci, val in enumerate(row_data, 1):
            c = ws2.cell(ri, ci, val)
            c.fill = bg;  c.border = _bd()
            c.alignment = _al("center" if ci > 1 else "left")
            c.font = _font()
            if ci == 3: c.number_format = MONEY_FMT
            if ci == 4: c.number_format = "0.0\"%\""
        ws2.row_dimensions[ri].height = 15

    # Ligne total récap
    row_t2 = len(cats_sorted) + 4
    ws2.cell(row_t2, 1, "TOTAL").font = TOT_FG
    ws2.cell(row_t2, 1).fill  = TOT_BG;  ws2.cell(row_t2, 1).border = _bd()
    ws2.cell(row_t2, 2, sum(nb_by_cat.values())).font = TOT_FG
    ws2.cell(row_t2, 2).fill  = TOT_BG;  ws2.cell(row_t2, 2).border = _bd()
    ws2.cell(row_t2, 2).alignment = _al("center")
    c_t = ws2.cell(row_t2, 3, round(grand, 2))
    c_t.font = TOT_FG;  c_t.fill = TOT_BG;  c_t.border = _bd()
    c_t.number_format = MONEY_FMT;  c_t.alignment = _al("center")
    ws2.cell(row_t2, 4, 100).font = TOT_FG
    ws2.cell(row_t2, 4).fill = TOT_BG;  ws2.cell(row_t2, 4).border = _bd()
    ws2.cell(row_t2, 4).alignment = _al("center")
    ws2.row_dimensions[row_t2].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"depenses_{date_cls.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/api/depenses/export/pdf")
def export_depenses_pdf(
    date_debut: Optional[str] = None,
    date_fin:   Optional[str] = None,
    categorie:  Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Export PDF des dépenses : entête, tableau détaillé, récapitulatif par catégorie."""
    import io
    from datetime import date as date_cls
    from collections import defaultdict
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    depenses = _query_depenses_filtrees(db, date_debut, date_fin, categorie)

    # ── Palette ────────────────────────────────────────────────
    C_RED    = colors.HexColor("#DC2626")
    C_RED2   = colors.HexColor("#F87171")
    C_DARK   = colors.HexColor("#0F172A")
    C_GRAY   = colors.HexColor("#64748B")
    C_LIGHT  = colors.HexColor("#FEF2F2")
    C_LIGHT2 = colors.HexColor("#FFF8F8")
    C_WHITE  = colors.white
    C_TOTAL  = colors.HexColor("#FEE2E2")
    C_GREEN  = colors.HexColor("#16A34A")
    C_BORDER = colors.HexColor("#FECACA")

    # ── Styles texte ───────────────────────────────────────────
    st = ParagraphStyle
    S_TITLE  = st("title",  fontName="Helvetica-Bold", fontSize=18, textColor=C_RED,  alignment=TA_CENTER, spaceAfter=2)
    S_SUB    = st("sub",    fontName="Helvetica",      fontSize=9,  textColor=C_GRAY, alignment=TA_CENTER, spaceAfter=0)
    S_H2     = st("h2",     fontName="Helvetica-Bold", fontSize=11, textColor=C_RED,  spaceBefore=14, spaceAfter=6)
    S_CELL   = st("cell",   fontName="Helvetica",      fontSize=8,  textColor=C_DARK, leading=10)
    S_TOTAL  = st("tot",    fontName="Helvetica-Bold", fontSize=9,  textColor=C_RED)
    S_NOTE   = st("note",   fontName="Helvetica-Oblique", fontSize=7.5, textColor=C_GRAY, alignment=TA_CENTER)

    def _tbl_style(has_total=False):
        cmds = [
            ("BACKGROUND",   (0, 0), (-1, 0),  C_DARK),
            ("TEXTCOLOR",    (0, 0), (-1, 0),  C_WHITE),
            ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 8),
            ("ALIGN",        (0, 0), (-1, -1), "LEFT"),
            ("ALIGN",        (0, 0), (0, -1),  "CENTER"),
            ("ALIGN",        (5, 0), (6, -1),  "RIGHT"),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2 if has_total else -1), [C_LIGHT2, C_WHITE]),
            ("GRID",         (0, 0), (-1, -1), 0.35, C_BORDER),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ]
        if has_total:
            cmds += [
                ("BACKGROUND",  (0, -1), (-1, -1), C_TOTAL),
                ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
                ("TEXTCOLOR",   (0, -1), (-1, -1), C_RED),
            ]
        t = TableStyle(cmds)
        return t

    def _htg_fmt(v):
        return f"{v:,.2f} HTG"

    # ── Construction du document ───────────────────────────────
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    W = landscape(A4)[0] - 3*cm   # largeur utile

    story = []

    # ── En-tête document ──────────────────────────────────────
    story.append(Paragraph("PétroSync", ParagraphStyle("brand", fontName="Helvetica-Bold", fontSize=10, textColor=C_GRAY, alignment=TA_CENTER)))
    story.append(Paragraph("Rapport des Dépenses Opérationnelles", S_TITLE))

    periode_txt = ""
    if date_debut or date_fin:
        periode_txt = f"Période : {date_debut or '…'} → {date_fin or '…'}  ·  "
    if categorie:
        periode_txt += f"Catégorie : {categorie}  ·  "
    periode_txt += f"{len(depenses)} dépense(s)  ·  Généré le {date_cls.today().strftime('%d/%m/%Y')}"
    story.append(Paragraph(periode_txt, S_SUB))
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width=W, thickness=1.5, color=C_RED2, spaceAfter=8))

    # ── KPI résumé (1 ligne de 3 blocs) ──────────────────────
    total = round(sum(float(d.montant) for d in depenses), 2)
    par_cat = defaultdict(float)
    for d in depenses:
        par_cat[d.categorie] += float(d.montant)
    cat_max = max(par_cat, key=par_cat.get) if par_cat else "—"

    def _kpi(label, val):
        return [Paragraph(f"<b><font size=7 color='#64748B'>{label}</font></b><br/>"
                          f"<font size=13 color='#DC2626'><b>{val}</b></font>",
                          ParagraphStyle("kpi", fontName="Helvetica", fontSize=9,
                                         alignment=TA_CENTER, leading=16))]

    kpi_data = [[_kpi("TOTAL DÉPENSES", _htg_fmt(total)),
                 _kpi("NOMBRE DE DÉPENSES", str(len(depenses))),
                 _kpi("CATÉGORIE PRINCIPALE", cat_max)]]
    kpi_tbl = Table(kpi_data, colWidths=[W/3]*3)
    kpi_tbl.setStyle(TableStyle([
        ("BOX",        (0, 0), (-1, -1), 0.5, C_BORDER),
        ("LINEAFTER",  (0, 0), (1, 0),   0.5, C_BORDER),
        ("BACKGROUND", (0, 0), (-1, -1), C_LIGHT),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("ROUNDEDCORNERS", [4]),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 0.4*cm))

    # ── Tableau détail ────────────────────────────────────────
    story.append(Paragraph("Détail des dépenses", S_H2))

    col_w = [1*cm, 2.2*cm, 2.8*cm, 6.5*cm, 3.5*cm, 2.5*cm, 3.2*cm]
    hdr = [["#", "Date", "Catégorie", "Description", "Bénéficiaire", "Référence", "Montant (HTG)"]]
    rows = hdr[:]
    for i, d in enumerate(depenses, 1):
        rows.append([
            str(i),
            d.date_depense.strftime("%d/%m/%Y") if d.date_depense else "",
            d.categorie,
            Paragraph(d.description, S_CELL),
            d.beneficiaire or "—",
            d.reference    or "—",
            _htg_fmt(float(d.montant)),
        ])
    rows.append(["", "", "", "", "", "TOTAL", _htg_fmt(total)])

    t_detail = Table(rows, colWidths=col_w, repeatRows=1)
    t_detail.setStyle(_tbl_style(has_total=True))
    story.append(t_detail)
    story.append(Spacer(1, 0.5*cm))

    # ── Récapitulatif par catégorie ───────────────────────────
    if len(par_cat) > 0:
        story.append(HRFlowable(width=W, thickness=0.5, color=C_BORDER, spaceAfter=6))
        story.append(Paragraph("Récapitulatif par catégorie", S_H2))

        cats_sorted = sorted(par_cat.items(), key=lambda x: -x[1])
        nb_by_cat = defaultdict(int)
        for d in depenses:
            nb_by_cat[d.categorie] += 1

        cat_col_w = [(W - 6*cm) / 2, 2*cm, 3*cm, 3*cm]
        cat_rows = [["Catégorie", "Nb", "Montant (HTG)", "Part (%)"]]
        for cat, mnt in cats_sorted:
            pct = round(mnt / total * 100, 1) if total else 0
            cat_rows.append([cat, str(nb_by_cat[cat]), _htg_fmt(mnt), f"{pct} %"])
        cat_rows.append(["TOTAL", str(len(depenses)), _htg_fmt(total), "100.0 %"])

        t_cat = Table(cat_rows, colWidths=cat_col_w, repeatRows=1)
        cat_style = TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0),  C_DARK),
            ("TEXTCOLOR",    (0, 0), (-1, 0),  C_WHITE),
            ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 8),
            ("ALIGN",        (1, 0), (-1, -1), "CENTER"),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [C_LIGHT2, C_WHITE]),
            ("BACKGROUND",   (0, -1), (-1, -1), C_TOTAL),
            ("FONTNAME",     (0, -1), (-1, -1), "Helvetica-Bold"),
            ("TEXTCOLOR",    (0, -1), (-1, -1), C_RED),
            ("GRID",         (0, 0), (-1, -1), 0.35, C_BORDER),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ])
        t_cat.setStyle(cat_style)
        story.append(t_cat)

    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BORDER, spaceAfter=4))
    story.append(Paragraph(
        f"Document généré automatiquement par PétroSync  ·  {date_cls.today().strftime('%d/%m/%Y')}",
        S_NOTE
    ))

    doc.build(story)
    buf.seek(0)
    fname = f"depenses_{date_cls.today().isoformat()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ══════════════════════════════════════════════════════════════════
# MODULE ACHATS
# ══════════════════════════════════════════════════════════════════

_CATEGORIES_ACHAT = {
    "Equipement", "Pieces detachees", "Fournitures bureau",
    "Informatique", "Securite", "Nettoyage", "Autre",
}

class AchatIn(BaseModel):
    fournisseur: str
    description: str
    categorie:   str
    montant:     float
    date_achat:  str
    reference:   Optional[str] = None
    notes:       Optional[str] = None

class AchatPatch(BaseModel):
    fournisseur: Optional[str]   = None
    description: Optional[str]   = None
    categorie:   Optional[str]   = None
    montant:     Optional[float] = None
    date_achat:  Optional[str]   = None
    reference:   Optional[str]   = None
    notes:       Optional[str]   = None

@app.get("/api/achats")
def lister_achats(
    date_debut:  Optional[str] = None,
    date_fin:    Optional[str] = None,
    categorie:   Optional[str] = None,
    fournisseur: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(Achat)
    if date_debut:
        try:
            from datetime import date as _date
            q = q.filter(Achat.date_achat >= _date.fromisoformat(date_debut))
        except ValueError:
            raise HTTPException(400, "date_debut invalide.")
    if date_fin:
        try:
            from datetime import date as _date
            q = q.filter(Achat.date_achat <= _date.fromisoformat(date_fin))
        except ValueError:
            raise HTTPException(400, "date_fin invalide.")
    if categorie:
        q = q.filter(Achat.categorie == categorie)
    if fournisseur:
        q = q.filter(Achat.fournisseur.ilike(f"%{fournisseur}%"))
    achats = q.order_by(Achat.date_achat.desc()).all()
    total = round(sum(float(a.montant) for a in achats), 2)
    return {
        "total": total,
        "nb": len(achats),
        "achats": [
            {
                "id": a.id, "fournisseur": a.fournisseur,
                "description": a.description, "categorie": a.categorie,
                "montant": float(a.montant), "date_achat": str(a.date_achat),
                "reference": a.reference, "notes": a.notes,
                "created_at": a.created_at.isoformat() if a.created_at else None,
            }
            for a in achats
        ],
    }

@app.post("/api/achats", status_code=201)
def creer_achat(data: AchatIn, db: Session = Depends(get_db)):
    if data.categorie not in _CATEGORIES_ACHAT:
        raise HTTPException(400, f"Catégorie invalide. Valeurs : {sorted(_CATEGORIES_ACHAT)}")
    if data.montant <= 0:
        raise HTTPException(400, "Le montant doit être > 0.")
    try:
        from datetime import date as _date
        date_a = _date.fromisoformat(data.date_achat)
    except ValueError:
        raise HTTPException(400, "Format de date invalide.")
    a = Achat(
        fournisseur=data.fournisseur.strip(), description=data.description.strip(),
        categorie=data.categorie, montant=data.montant,
        date_achat=date_a, reference=data.reference, notes=data.notes,
    )
    db.add(a)
    db.commit()
    db.refresh(a)
    return {"id": a.id, "message": "Achat enregistré."}

@app.put("/api/achats/{achat_id}")
def modifier_achat(achat_id: int, data: AchatPatch, db: Session = Depends(get_db)):
    a = db.query(Achat).filter(Achat.id == achat_id).first()
    if not a:
        raise HTTPException(404, "Achat introuvable.")
    if data.categorie is not None:
        if data.categorie not in _CATEGORIES_ACHAT:
            raise HTTPException(400, "Catégorie invalide.")
        a.categorie = data.categorie
    if data.fournisseur is not None: a.fournisseur = data.fournisseur.strip()
    if data.description is not None: a.description = data.description.strip()
    if data.montant is not None:
        if data.montant <= 0:
            raise HTTPException(400, "Le montant doit être > 0.")
        a.montant = data.montant
    if data.date_achat is not None:
        try:
            from datetime import date as _date
            a.date_achat = _date.fromisoformat(data.date_achat)
        except ValueError:
            raise HTTPException(400, "Format de date invalide.")
    if data.reference is not None: a.reference = data.reference
    if data.notes     is not None: a.notes     = data.notes
    db.commit()
    return {"message": "Achat mis à jour."}

@app.delete("/api/achats/{achat_id}")
def supprimer_achat(achat_id: int, db: Session = Depends(get_db)):
    a = db.query(Achat).filter(Achat.id == achat_id).first()
    if not a:
        raise HTTPException(404, "Achat introuvable.")
    db.delete(a)
    db.commit()
    return {"message": "Achat supprimé."}


# ══════════════════════════════════════════════════════════
# RAPPORT DE CAISSE — Synthèse financière consolidée
# ══════════════════════════════════════════════════════════
@app.get("/api/caisse/rapport")
def rapport_caisse(
    date_debut: Optional[date_type] = None,
    date_fin:   Optional[date_type] = None,
    db: Session = Depends(get_db),
):
    """
    Synthèse financière : Ventes – Achats – Payroll – Dépenses = Cash disponible.
    Sans filtre de dates : toutes les données en base.
    """
    from datetime import date as _date

    # ── Ventes (releves) ──────────────────────────────────
    q_rel = db.query(Releve)
    if date_debut: q_rel = q_rel.filter(Releve.date >= date_debut)
    if date_fin:   q_rel = q_rel.filter(Releve.date <= date_fin)
    releves = q_rel.all()
    total_ventes = round(sum(float(r.montant_vente) for r in releves), 2)
    total_gallons = round(sum(float(r.quantite) for r in releves), 3)

    # Ventilation par produit
    prod_map = {}
    for r in releves:
        nom = r.pompe.produit.nom
        e = prod_map.setdefault(nom, {"gallons": 0.0, "montant": 0.0})
        e["gallons"]  += float(r.quantite)
        e["montant"]  += float(r.montant_vente)
    ventes_par_produit = [
        {"produit": k, "gallons": round(v["gallons"], 3), "montant": round(v["montant"], 2)}
        for k, v in prod_map.items()
    ]

    # ── Achats ───────────────────────────────────────────
    q_ach = db.query(Achat)
    if date_debut: q_ach = q_ach.filter(Achat.date_achat >= date_debut)
    if date_fin:   q_ach = q_ach.filter(Achat.date_achat <= date_fin)
    achats = q_ach.all()
    total_achats = round(sum(float(a.montant) for a in achats), 2)

    # ── Payroll payé ─────────────────────────────────────
    q_pay = db.query(FichePaie).filter(FichePaie.statut == "paye")
    if date_debut: q_pay = q_pay.filter(FichePaie.date_paiement >= date_debut)
    if date_fin:   q_pay = q_pay.filter(FichePaie.date_paiement <= date_fin)
    fiches_payees = q_pay.all()
    total_payroll_paye = round(sum(float(f.net_a_payer) for f in fiches_payees), 2)

    # Payroll brouillon (engagé mais non encore versé)
    q_brou = db.query(FichePaie).filter(FichePaie.statut == "brouillon")
    if date_debut: q_brou = q_brou.filter(FichePaie.periode_debut >= date_debut)
    if date_fin:   q_brou = q_brou.filter(FichePaie.periode_fin   <= date_fin)
    fiches_brou = q_brou.all()
    total_payroll_engage = round(sum(float(f.net_a_payer) for f in fiches_brou), 2)

    # ── Dépenses ─────────────────────────────────────────
    q_dep = db.query(Depense)
    if date_debut: q_dep = q_dep.filter(Depense.date_depense >= date_debut)
    if date_fin:   q_dep = q_dep.filter(Depense.date_depense <= date_fin)
    depenses = q_dep.all()
    total_depenses = round(sum(float(d.montant) for d in depenses), 2)

    # Ventilation dépenses par catégorie
    cat_map = {}
    for d in depenses:
        cat_map[d.categorie] = round(cat_map.get(d.categorie, 0.0) + float(d.montant), 2)
    depenses_par_cat = [{"categorie": k, "montant": v} for k, v in sorted(cat_map.items(), key=lambda x: -x[1])]

    # ── Synthèse ─────────────────────────────────────────
    total_sorties = round(total_achats + total_payroll_paye + total_depenses, 2)
    cash_disponible = round(total_ventes - total_sorties, 2)
    cash_apres_engage = round(cash_disponible - total_payroll_engage, 2)

    return {
        "periode": {
            "debut": str(date_debut) if date_debut else None,
            "fin":   str(date_fin)   if date_fin   else None,
        },
        "ventes": {
            "total":   total_ventes,
            "gallons": total_gallons,
            "nb_releves": len(releves),
            "par_produit": ventes_par_produit,
        },
        "achats": {
            "total": total_achats,
            "nb":    len(achats),
        },
        "payroll": {
            "paye":    total_payroll_paye,
            "engage":  total_payroll_engage,
            "nb_fiches_payees":  len(fiches_payees),
            "nb_fiches_engage":  len(fiches_brou),
        },
        "depenses": {
            "total": total_depenses,
            "nb":    len(depenses),
            "par_categorie": depenses_par_cat,
        },
        "synthese": {
            "total_entrees":    total_ventes,
            "total_sorties":    total_sorties,
            "cash_disponible":  cash_disponible,
            "cash_apres_engage": cash_apres_engage,
        },
    }


# ══════════════════════════════════════════════════════════════════
# TABLEAU DE BORD GI — Données agrégées pour le dashboard interactif
# ══════════════════════════════════════════════════════════════════
@app.get("/api/gi/dashboard")
def gi_dashboard(
    date_debut:  Optional[date_type] = None,
    date_fin:    Optional[date_type] = None,
    produit_id:  Optional[int]       = None,
    db: Session = Depends(get_db),
):
    """
    Agrège en un seul appel toutes les données du dashboard GI :
    ventes, trend quotidien, employés, payroll, dépenses, achats, cash.
    """
    from datetime import date as _date, timedelta
    from collections import defaultdict

    today = _date.today()
    if not date_debut:
        date_debut = _date(today.year, today.month, 1)   # début du mois courant
    if not date_fin:
        date_fin = today

    nb_jours = (date_fin - date_debut).days + 1

    # ── Ventes (releves) ───────────────────────────────────────
    q_releves = (db.query(Releve)
                 .filter(Releve.date >= date_debut, Releve.date <= date_fin))
    if produit_id:
        from sqlalchemy.orm import joinedload
        q_releves = q_releves.join(Releve.pompe).filter(Pompe.produit_id == produit_id)
    releves = q_releves.order_by(Releve.date).all()

    total_ventes  = round(sum(float(r.montant_vente) for r in releves), 2)
    total_gallons = round(sum(float(r.quantite) for r in releves), 3)
    nb_releves    = len(releves)

    # Trend quotidien (regroupe par jour)
    trend_ventes: dict[str, float] = {}
    trend_gallons: dict[str, float] = {}
    for r in releves:
        k = str(r.date)
        trend_ventes[k]  = round(trend_ventes.get(k, 0)  + float(r.montant_vente), 2)
        trend_gallons[k] = round(trend_gallons.get(k, 0) + float(r.quantite), 3)

    # Ventes par produit
    prod_map: dict = {}
    for r in releves:
        nom = r.pompe.produit.nom
        e = prod_map.setdefault(nom, {"gallons": 0.0, "montant": 0.0})
        e["gallons"]  += float(r.quantite)
        e["montant"]  += float(r.montant_vente)
    ventes_par_produit = [
        {"produit": k, "gallons": round(v["gallons"], 3), "montant": round(v["montant"], 2)}
        for k, v in prod_map.items()
    ]

    # ── Employés ───────────────────────────────────────────────
    tous_employes  = db.query(Employe).all()
    employes_actifs = [e for e in tous_employes if e.actif]
    masse_salariale = round(sum(float(e.salaire_base) for e in employes_actifs), 2)
    repartition_postes: dict = {}
    for e in employes_actifs:
        repartition_postes[e.poste] = repartition_postes.get(e.poste, 0) + 1

    # ── Payroll ────────────────────────────────────────────────
    fiches_payees = (db.query(FichePaie)
                     .filter(FichePaie.statut == "paye",
                             FichePaie.date_paiement >= date_debut,
                             FichePaie.date_paiement <= date_fin)
                     .all())
    fiches_brou   = db.query(FichePaie).filter(FichePaie.statut == "brouillon").all()
    total_payroll_paye   = round(sum(float(f.net_a_payer) for f in fiches_payees), 2)
    total_payroll_engage = round(sum(float(f.net_a_payer) for f in fiches_brou), 2)

    # ── Dépenses ───────────────────────────────────────────────
    depenses = (db.query(Depense)
                .filter(Depense.date_depense >= date_debut,
                        Depense.date_depense <= date_fin)
                .all())
    total_depenses = round(sum(float(d.montant) for d in depenses), 2)
    dep_par_cat: dict[str, float] = {}
    for d in depenses:
        dep_par_cat[d.categorie] = round(dep_par_cat.get(d.categorie, 0) + float(d.montant), 2)
    dep_trend: dict[str, float] = {}
    for d in depenses:
        k = str(d.date_depense)
        dep_trend[k] = round(dep_trend.get(k, 0) + float(d.montant), 2)

    # ── Achats ─────────────────────────────────────────────────
    achats = (db.query(Achat)
              .filter(Achat.date_achat >= date_debut,
                      Achat.date_achat <= date_fin)
              .all())
    total_achats = round(sum(float(a.montant) for a in achats), 2)
    ach_par_cat: dict[str, float] = {}
    for a in achats:
        ach_par_cat[a.categorie] = round(ach_par_cat.get(a.categorie, 0) + float(a.montant), 2)

    # ── Synthèse cash ──────────────────────────────────────────
    total_sorties   = round(total_achats + total_payroll_paye + total_depenses, 2)
    cash_disponible = round(total_ventes - total_sorties, 2)
    marge_pct       = round(cash_disponible / total_ventes * 100, 1) if total_ventes else 0.0

    # Série de jours pour le graphique (tous les jours de la période)
    jours_serie = []
    cur = date_debut
    while cur <= date_fin:
        k = str(cur)
        jours_serie.append({
            "date":     k,
            "ventes":   trend_ventes.get(k, 0),
            "gallons":  trend_gallons.get(k, 0),
            "depenses": dep_trend.get(k, 0),
        })
        cur += timedelta(days=1)

    return {
        "periode": {"debut": str(date_debut), "fin": str(date_fin), "nb_jours": nb_jours},
        "ventes": {
            "total":       total_ventes,
            "gallons":     total_gallons,
            "nb_releves":  nb_releves,
            "par_produit": ventes_par_produit,
        },
        "employes": {
            "total":           len(tous_employes),
            "actifs":          len(employes_actifs),
            "masse_salariale": masse_salariale,
            "par_poste":       [{"poste": k, "nb": v} for k, v in
                                sorted(repartition_postes.items(), key=lambda x: -x[1])],
        },
        "payroll": {
            "paye":                total_payroll_paye,
            "engage":              total_payroll_engage,
            "nb_fiches_payees":    len(fiches_payees),
            "nb_fiches_engage":    len(fiches_brou),
        },
        "depenses": {
            "total":       total_depenses,
            "nb":          len(depenses),
            "par_categorie": sorted(
                [{"categorie": k, "montant": v} for k, v in dep_par_cat.items()],
                key=lambda x: -x["montant"],
            ),
        },
        "achats": {
            "total":       total_achats,
            "nb":          len(achats),
            "par_categorie": sorted(
                [{"categorie": k, "montant": v} for k, v in ach_par_cat.items()],
                key=lambda x: -x["montant"],
            ),
        },
        "synthese": {
            "total_sorties":    total_sorties,
            "cash_disponible":  cash_disponible,
            "marge_pct":        marge_pct,
            "cash_apres_engage": round(cash_disponible - total_payroll_engage, 2),
        },
        "trend": jours_serie,
    }


# ══════════════════════════════════════════════════════════════════
# STATISTIQUES AVANCÉES
# ══════════════════════════════════════════════════════════════════
@app.get("/api/statistiques")
async def get_statistiques(
    date_debut:  Optional[str] = None,
    date_fin:    Optional[str] = None,
    produit_id:  Optional[int] = None,
    db:          Session = Depends(get_db),
):
    import statistics as _stat
    from collections import defaultdict
    from datetime import timedelta

    today = date_type.today()
    fin   = date_type.fromisoformat(date_fin)   if date_fin   else today
    debut = date_type.fromisoformat(date_debut) if date_debut else fin - timedelta(days=29)
    nb_jours = (fin - debut).days + 1

    prev_fin   = debut - timedelta(days=1)
    prev_debut = prev_fin - timedelta(days=nb_jours - 1)

    # ── Requête relevés ──────────────────────────────────────────
    def _q_releves(d0, d1):
        q = db.query(Releve).filter(Releve.date >= d0, Releve.date <= d1)
        if produit_id:
            ids = [p.id for p in db.query(Pompe).filter(Pompe.produit_id == produit_id).all()]
            q = q.filter(Releve.pompe_id.in_(ids)) if ids else q.filter(False)
        return q.all()

    releves  = _q_releves(debut, fin)
    prev_rel = _q_releves(prev_debut, prev_fin)

    ca_total  = sum(r.montant_vente for r in releves)
    vol_total = sum(r.quantite      for r in releves)
    ca_prev   = sum(r.montant_vente for r in prev_rel)
    vol_prev  = sum(r.quantite      for r in prev_rel)

    # ── Série journalière ────────────────────────────────────────
    daily = defaultdict(lambda: {"montant": 0.0, "quantite": 0.0})
    for r in releves:
        ds = r.date.isoformat() if hasattr(r.date, "isoformat") else str(r.date)
        daily[ds]["montant"]  += r.montant_vente
        daily[ds]["quantite"] += r.quantite

    daily_montants  = [v["montant"] for v in daily.values()]
    nb_jours_actifs = len(daily)

    std_ca = _stat.stdev(daily_montants) if len(daily_montants) > 1 else 0.0
    moy_ca = _stat.mean(daily_montants)  if daily_montants else 0.0
    cv_ca  = (std_ca / moy_ca * 100) if moy_ca > 0 else 0.0

    max_j = max(daily.items(), key=lambda x: x[1]["montant"]) if daily else (None, {"montant": 0})
    min_j = min(daily.items(), key=lambda x: x[1]["montant"]) if daily else (None, {"montant": 0})

    matin_ca = sum(r.montant_vente for r in releves if r.periode == "Matin")
    soir_ca  = sum(r.montant_vente for r in releves if r.periode == "Apres-midi")

    ca_var  = ((ca_total  - ca_prev)  / ca_prev  * 100) if ca_prev  > 0 else None
    vol_var = ((vol_total - vol_prev) / vol_prev * 100) if vol_prev > 0 else None

    # ── Performance pompes ───────────────────────────────────────
    pompe_map = defaultdict(lambda: {"ca": 0.0, "vol": 0.0, "nom": "", "produit": ""})
    for r in releves:
        pompe_map[r.pompe_id]["ca"]  += r.montant_vente
        pompe_map[r.pompe_id]["vol"] += r.quantite
        if r.pompe:
            pompe_map[r.pompe_id]["nom"]     = r.pompe.nom
            pompe_map[r.pompe_id]["produit"] = r.pompe.produit.nom if r.pompe.produit else "—"

    pompes_list = [{
        "id": pid, "nom": ps["nom"], "produit": ps["produit"],
        "ca":      round(ps["ca"],  2),
        "vol":     round(ps["vol"], 3),
        "pct_ca":  round(ps["ca"]  / ca_total  * 100, 1) if ca_total  > 0 else 0,
        "pct_vol": round(ps["vol"] / vol_total * 100, 1) if vol_total > 0 else 0,
    } for pid, ps in sorted(pompe_map.items(), key=lambda x: x[1]["ca"], reverse=True)]

    # ── Marges par produit ───────────────────────────────────────
    produits = db.query(Produit).filter(Produit.actif == True).all()
    marges = []
    for prod in produits:
        px_hist = db.query(PrixVente).filter(PrixVente.produit_id == prod.id).order_by(PrixVente.date_effet.desc()).first()
        px_vente = float(px_hist.prix_vente_gallon) if px_hist else float(prod.prix_gallon)
        last_deliv = db.query(Livraison).filter(Livraison.produit_id == prod.id).order_by(Livraison.date_livraison.desc()).first()
        cout_achat = float(last_deliv.prix_achat_gallon) if last_deliv else None
        marge_gal  = (px_vente - cout_achat) if cout_achat is not None else None
        marge_pct  = (marge_gal / px_vente * 100) if marge_gal is not None and px_vente > 0 else None
        marges.append({
            "produit":    prod.nom,
            "px_vente":   round(px_vente,  2),
            "cout_achat": round(cout_achat, 2) if cout_achat is not None else None,
            "marge_gal":  round(marge_gal,  2) if marge_gal  is not None else None,
            "marge_pct":  round(marge_pct,  1) if marge_pct  is not None else None,
        })

    # ── Livraisons & coûts ───────────────────────────────────────
    lq = db.query(Livraison).filter(Livraison.date_livraison >= debut, Livraison.date_livraison <= fin)
    if produit_id:
        lq = lq.filter(Livraison.produit_id == produit_id)
    livraisons = lq.all()
    vol_recu_total = sum(float(l.gallons_recus) for l in livraisons)
    cout_achats    = sum(float(l.gallons_recus) * float(l.prix_achat_gallon) for l in livraisons)

    # ── Dépenses ────────────────────────────────────────────────
    depenses = db.query(Depense).filter(Depense.date_depense >= debut, Depense.date_depense <= fin).all()
    total_dep = sum(float(d.montant) for d in depenses)

    marge_brute      = ca_total - cout_achats
    marge_brute_pct  = (marge_brute / ca_total * 100) if ca_total > 0 else 0.0
    resultat_net     = marge_brute - total_dep
    roi_pct          = (resultat_net / cout_achats * 100) if cout_achats > 0 else None

    cat_map = defaultdict(float)
    for d in depenses:
        cat_map[d.categorie] += float(d.montant)
    dep_cats = [
        {"categorie": cat, "montant": round(m, 2), "pct": round(m / total_dep * 100, 1) if total_dep > 0 else 0}
        for cat, m in sorted(cat_map.items(), key=lambda x: x[1], reverse=True)
    ]

    # ── Point mort ───────────────────────────────────────────────
    cout_var_moy  = (cout_achats / vol_recu_total) if vol_recu_total > 0 else None
    px_vente_moy  = (ca_total / vol_total)          if vol_total > 0      else None
    marge_contrib = (px_vente_moy - cout_var_moy)   if px_vente_moy and cout_var_moy else None
    seuil_gal     = (total_dep / marge_contrib)      if marge_contrib and marge_contrib > 0 else None
    seuil_ca      = (seuil_gal * px_vente_moy)       if seuil_gal and px_vente_moy else None
    couv_pct      = (vol_total / seuil_gal * 100)    if seuil_gal else None

    # ── Rotation des stocks ──────────────────────────────────────
    rotation = []
    for prod in produits:
        prod_ids    = {p.id for p in prod.pompes}
        prod_rel    = [r for r in releves if r.pompe_id in prod_ids]
        vol_vendu   = sum(r.quantite for r in prod_rel)
        prod_livs   = [l for l in livraisons if l.produit_id == prod.id]
        vol_recv    = sum(float(l.gallons_recus) for l in prod_livs)
        stock_moy   = vol_recv / 2 if vol_recv > 0 else None
        taux_rot    = (vol_vendu / stock_moy) if stock_moy else None
        jours_cov   = (stock_moy / (vol_vendu / nb_jours)) if stock_moy and vol_vendu > 0 else None
        rotation.append({
            "produit":          prod.nom,
            "vol_vendu":        round(vol_vendu, 3),
            "vol_recu":         round(vol_recv,  3),
            "taux_rotation":    round(taux_rot,  2) if taux_rot  is not None else None,
            "jours_couverture": round(jours_cov, 1) if jours_cov is not None else None,
        })

    # ── Série temporelle ─────────────────────────────────────────
    serie, cur = [], debut
    while cur <= fin:
        ds = cur.isoformat()
        serie.append({"date": ds, "montant": round(daily[ds]["montant"], 2), "quantite": round(daily[ds]["quantite"], 3)})
        cur += timedelta(days=1)

    # ── Alertes automatiques ─────────────────────────────────────
    alertes = []
    if ca_var is not None and ca_var >= 5:
        alertes.append({"type": "ok",      "msg": f"Croissance du CA de +{ca_var:.1f}% vs période précédente"})
    if cv_ca > 40:
        alertes.append({"type": "warning", "msg": f"Forte variabilité des ventes — CV={cv_ca:.1f}% (seuil normal < 40%)"})
    if ca_var is not None and ca_var < -10:
        alertes.append({"type": "danger",  "msg": f"Baisse du CA de {abs(ca_var):.1f}% vs période précédente"})
    if seuil_gal and vol_total < seuil_gal:
        alertes.append({"type": "danger",  "msg": f"Volume vendu ({vol_total:.0f} gal) inférieur au point mort ({seuil_gal:.0f} gal)"})
    if resultat_net < 0:
        alertes.append({"type": "danger",  "msg": f"Résultat net négatif : {resultat_net:,.0f} G"})
    if 0 < marge_brute_pct < 15:
        alertes.append({"type": "warning", "msg": f"Marge brute faible : {marge_brute_pct:.1f}% (idéal ≥ 15%)"})
    if not alertes:
        alertes.append({"type": "ok", "msg": "Aucune alerte critique — indicateurs dans les normes"})

    return {
        "periode": {
            "debut": debut.isoformat(), "fin": fin.isoformat(),
            "nb_jours": nb_jours, "nb_jours_actifs": nb_jours_actifs,
            "prev_debut": prev_debut.isoformat(), "prev_fin": prev_fin.isoformat(),
        },
        "synthese": {
            "ca_total":       round(ca_total,  2),
            "ca_moyen_jour":  round(ca_total / nb_jours, 2),
            "ca_moyen_actif": round(ca_total / nb_jours_actifs, 2) if nb_jours_actifs else 0,
            "vol_total":      round(vol_total, 3),
            "vol_moyen_jour": round(vol_total / nb_jours, 3),
            "nb_releves":     len(releves),
            "matin_pct":      round(matin_ca / ca_total * 100, 1) if ca_total > 0 else 0,
            "soir_pct":       round(soir_ca  / ca_total * 100, 1) if ca_total > 0 else 0,
        },
        "croissance": {
            "ca_var_pct":  round(ca_var,  1) if ca_var  is not None else None,
            "vol_var_pct": round(vol_var, 1) if vol_var is not None else None,
            "ca_prev":     round(ca_prev,  2),
            "vol_prev":    round(vol_prev, 3),
        },
        "variabilite": {
            "ecart_type_ca": round(std_ca, 2),
            "cv_ca":         round(cv_ca,  1),
            "max_jour": {"date": max_j[0], "montant": round(max_j[1]["montant"], 2)} if max_j[0] else None,
            "min_jour": {"date": min_j[0], "montant": round(min_j[1]["montant"], 2)} if min_j[0] else None,
            "pic_periode": "Matin" if matin_ca >= soir_ca else "Après-midi",
        },
        "rentabilite": {
            "ca_total":        round(ca_total,       2),
            "cout_achats":     round(cout_achats,    2),
            "marge_brute":     round(marge_brute,    2),
            "marge_brute_pct": round(marge_brute_pct,1),
            "depenses":        round(total_dep,      2),
            "resultat_net":    round(resultat_net,   2),
            "roi_pct":         round(roi_pct, 1) if roi_pct is not None else None,
        },
        "point_mort": {
            "cout_var_moy":       round(cout_var_moy,  2) if cout_var_moy  is not None else None,
            "px_vente_moy":       round(px_vente_moy,  2) if px_vente_moy  is not None else None,
            "marge_contribution": round(marge_contrib,  2) if marge_contrib is not None else None,
            "gallons_seuil":      round(seuil_gal, 1)      if seuil_gal     is not None else None,
            "ca_seuil":           round(seuil_ca,  2)      if seuil_ca      is not None else None,
            "couverture_pct":     round(couv_pct,  1)      if couv_pct      is not None else None,
        },
        "marges":              marges,
        "pompes":              pompes_list,
        "rotation":            rotation,
        "depenses_categories": dep_cats,
        "serie":               serie,
        "alertes":             alertes,
    }


# ---------- Frontend single-file ----------
@app.get("/", include_in_schema=False)
def serve_index():
    html_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "frontend", "index.html"
    )
    return FileResponse(
        html_path,
        media_type="text/html",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )
