"""Routes d'administration : gestion des rôles, comptes, sessions et journal."""
import io
import json
import re
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import desc
from sqlalchemy.orm import Session

from activity_log import (
    log_event,
    SESSION_REVOKED, SESSIONS_CLEARED, PASSWORD_RESET, PIN_RESET,
    USER_CREATED, USER_UPDATED, USER_DISABLED, USER_ENABLED,
)
from auth import hash_code_acces, hash_password, make_api_key
from otp_service import send_welcome_email
from database import get_db
from models import AuditLog, Employe, LoginSecurityEvent, Role, SessionToken, Utilisateur

router = APIRouter(prefix="/api/admin", tags=["admin"])

_EMAIL_RE    = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_USERNAME_RE = re.compile(r"^[a-zA-Z0-9._\-]{3,60}$")
_CODE_RE     = re.compile(r"^\d{9}$")

DOMAINES = ["finance", "bar", "cuisine", "hotel", "employes", "carburant"]
NIVEAUX  = ["aucun", "lecture", "complet"]

_PERMS_VIDES: dict = {d: "aucun" for d in DOMAINES}
_PERMS_VIDES["admin"] = False


def _require_admin(request: Request, db: Session = Depends(get_db)) -> Utilisateur:
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(403, "Non authentifié")
    if user.role == "admin":
        return user
    if user.role_id:
        u = db.get(Utilisateur, user.id)
        if u and u.role_obj and u.role_obj.permissions.get("admin", False):
            return u
    raise HTTPException(403, "Accès réservé aux administrateurs")


def _validate_permissions(perms: dict) -> dict:
    out = dict(_PERMS_VIDES)
    for d in DOMAINES:
        v = perms.get(d, "aucun")
        if v not in NIVEAUX:
            raise HTTPException(400, f"Niveau invalide '{v}' pour '{d}'. Valeurs: {NIVEAUX}")
        out[d] = v
    out["admin"] = bool(perms.get("admin", False))
    return out


def _role_public(r: Role) -> dict:
    return {
        "id":             r.id,
        "nom":            r.nom,
        "description":    r.description,
        "permissions":    r.permissions,
        "est_admin":      r.est_admin,
        "est_systeme":    r.est_systeme,
        "date_creation":  r.date_creation.isoformat() if r.date_creation else None,
        "nb_utilisateurs": len(r.utilisateurs),
    }


def _user_public(u: Utilisateur, db: Session = None) -> dict:
    employe_id = None
    if db:
        emp = db.query(Employe).filter_by(utilisateur_id=u.id).first()
        employe_id = emp.id if emp else None
    return {
        "id":             u.id,
        "username":       u.username,
        "nom_complet":    u.nom_complet,
        "email":          u.email,
        "role":           u.role,
        "poste":          u.poste,
        "role_id":        u.role_id,
        "role_nom":       u.role_obj.nom if u.role_obj else u.poste,
        "actif":          u.actif,
        "created_at":     u.created_at.isoformat() if u.created_at else None,
        "oauth_provider": u.oauth_provider,
        "employe_id":     employe_id,
    }


# ═══════════════════════════════════════
# RÔLES — CRUD
# ═══════════════════════════════════════

class RoleIn(BaseModel):
    nom:         str
    description: Optional[str] = None
    permissions: dict = {}


@router.get("/roles")
def lister_roles(
    _admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    return [_role_public(r) for r in db.query(Role).order_by(Role.id).all()]


@router.post("/roles", status_code=201)
def creer_role(
    data: RoleIn,
    _admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    nom = data.nom.strip()
    if not nom:
        raise HTTPException(400, "Le nom du rôle est requis")
    if db.query(Role).filter_by(nom=nom).first():
        raise HTTPException(409, f"Un rôle nommé '{nom}' existe déjà")
    perms = _validate_permissions(data.permissions)
    r = Role(
        nom=nom,
        description=data.description.strip() if data.description else None,
        permissions=perms,
        est_admin=bool(perms.get("admin", False)),
        est_systeme=False,
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    return _role_public(r)


@router.put("/roles/{role_id}")
def modifier_role(
    role_id: int,
    data: RoleIn,
    _admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    r = db.get(Role, role_id)
    if not r:
        raise HTTPException(404, "Rôle introuvable")
    nom = data.nom.strip()
    if not nom:
        raise HTTPException(400, "Le nom du rôle est requis")
    if db.query(Role).filter(Role.nom == nom, Role.id != role_id).first():
        raise HTTPException(409, f"Un rôle nommé '{nom}' existe déjà")
    perms = _validate_permissions(data.permissions)
    r.nom         = nom
    r.description = data.description.strip() if data.description else None
    r.permissions = perms
    r.est_admin   = bool(perms.get("admin", False))
    db.commit()
    db.refresh(r)
    return _role_public(r)


@router.delete("/roles/{role_id}")
def supprimer_role(
    role_id: int,
    _admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    r = db.get(Role, role_id)
    if not r:
        raise HTTPException(404, "Rôle introuvable")
    if r.est_systeme:
        raise HTTPException(409, "Les rôles système ne peuvent pas être supprimés")
    if db.query(Utilisateur).filter_by(role_id=role_id).count() > 0:
        raise HTTPException(409, "Ce rôle est attribué à des utilisateurs — réassignez-les d'abord")
    db.delete(r)
    db.commit()
    return {"ok": True, "id": role_id}


# ═══════════════════════════════════════
# UTILISATEURS — CRUD
# ═══════════════════════════════════════

class CreateUserIn(BaseModel):
    nom_complet: str
    email:       str
    username:    str
    password:    str
    code_acces:  str
    role_id:     int


class UpdateUserIn(BaseModel):
    nom_complet: Optional[str] = None
    email:       Optional[str] = None
    role_id:     Optional[int] = None
    actif:       Optional[bool] = None
    employe_id:  Optional[int] = None  # 0 = délier ; N = lier à cet employé


class ResetPasswordIn(BaseModel):
    nouveau_mot_de_passe: str


class ResetPinIn(BaseModel):
    nouveau_code: str


@router.get("/users")
def lister_users(
    _admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    return [_user_public(u, db) for u in db.query(Utilisateur).order_by(Utilisateur.id).all()]


@router.post("/users", status_code=201)
def creer_user(
    data: CreateUserIn,
    _admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    role = db.get(Role, data.role_id)
    if not role:
        raise HTTPException(404, "Rôle introuvable")
    email = data.email.strip().lower()
    if not _EMAIL_RE.match(email):
        raise HTTPException(400, "Adresse email invalide")
    username = data.username.strip()
    if not _USERNAME_RE.match(username):
        raise HTTPException(400, "Identifiant invalide (3-60 chars, lettres/chiffres/._-)")
    if not data.nom_complet.strip():
        raise HTTPException(400, "Le nom complet est requis")
    if len(data.password) < 8:
        raise HTTPException(400, "Le mot de passe doit contenir au moins 8 caractères")
    if not _CODE_RE.match(data.code_acces):
        raise HTTPException(400, "Le code PIN doit contenir exactement 9 chiffres")
    if db.query(Utilisateur).filter_by(username=username).first():
        raise HTTPException(409, f"L'identifiant '{username}' est déjà utilisé")
    if db.query(Utilisateur).filter_by(email=email).first():
        raise HTTPException(409, "Cet email est déjà associé à un compte")

    u = Utilisateur(
        username=username,
        password_hash=hash_password(data.password),
        code_acces_hash=hash_code_acces(data.code_acces),
        nom_complet=data.nom_complet.strip(),
        email=email,
        role="admin" if role.est_admin else "operateur",
        role_id=role.id,
        poste=role.nom,
        actif=True,
    )
    db.add(u)
    db.commit()
    db.refresh(u)
    raw_key = make_api_key(db, u.id)
    # Email de bienvenue — livraison non bloquante
    welcome_sent = False
    try:
        send_welcome_email(u.nom_complet or u.username, u.email, u.username)
        welcome_sent = True
    except Exception:
        pass
    result = _user_public(u)
    result["api_key"] = raw_key
    result["welcome_email_sent"] = welcome_sent
    return result


@router.put("/users/{uid}")
def modifier_user(
    uid: int,
    data: UpdateUserIn,
    _admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    u = db.get(Utilisateur, uid)
    if not u:
        raise HTTPException(404, "Utilisateur introuvable")
    if data.nom_complet is not None:
        if not data.nom_complet.strip():
            raise HTTPException(400, "Le nom complet est requis")
        u.nom_complet = data.nom_complet.strip()
    if data.email is not None:
        email = data.email.strip().lower()
        if not _EMAIL_RE.match(email):
            raise HTTPException(400, "Adresse email invalide")
        if db.query(Utilisateur).filter(Utilisateur.email == email, Utilisateur.id != uid).first():
            raise HTTPException(409, "Cet email est déjà associé à un autre compte")
        u.email = email
    if data.role_id is not None:
        role = db.get(Role, data.role_id)
        if not role:
            raise HTTPException(404, "Rôle introuvable")
        u.role_id = role.id
        u.poste   = role.nom
        u.role    = "admin" if role.est_admin else "operateur"
    if data.actif is not None:
        if not data.actif and u.role == "admin":
            nb_admins = db.query(Utilisateur).filter_by(role="admin", actif=True).count()
            if nb_admins <= 1:
                raise HTTPException(409, "Impossible : dernier administrateur actif")
        u.actif = data.actif
    if data.employe_id is not None:
        # Délier l'ancienne association si elle existait
        ancien = db.query(Employe).filter_by(utilisateur_id=uid).first()
        if ancien:
            ancien.utilisateur_id = None
        if data.employe_id > 0:
            emp = db.get(Employe, data.employe_id)
            if not emp:
                raise HTTPException(404, "Employé introuvable")
            emp.utilisateur_id = uid
    db.commit()
    db.refresh(u)
    return _user_public(u, db)


@router.post("/users/{uid}/reset-password")
def reset_password(
    uid: int,
    request: Request,
    data: ResetPasswordIn,
    admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    u = db.get(Utilisateur, uid)
    if not u:
        raise HTTPException(404, "Utilisateur introuvable")
    if len(data.nouveau_mot_de_passe) < 8:
        raise HTTPException(400, "Le mot de passe doit contenir au moins 8 caractères")
    u.password_hash = hash_password(data.nouveau_mot_de_passe)
    db.commit()
    log_event(db, PASSWORD_RESET, user_id=admin.id, target_user_id=uid,
              ip_address=request.client.host if request.client else None)
    return {"ok": True}


@router.post("/users/{uid}/reset-pin")
def reset_pin(
    uid: int,
    request: Request,
    data: ResetPinIn,
    admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    u = db.get(Utilisateur, uid)
    if not u:
        raise HTTPException(404, "Utilisateur introuvable")
    if not _CODE_RE.match(data.nouveau_code):
        raise HTTPException(400, "Le code PIN doit contenir exactement 9 chiffres")
    u.code_acces_hash = hash_code_acces(data.nouveau_code)
    db.commit()
    log_event(db, PIN_RESET, user_id=admin.id, target_user_id=uid,
              ip_address=request.client.host if request.client else None)
    return {"ok": True}


@router.post("/users/{uid}/test-email")
def test_email(
    uid: int,
    _admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    """Envoie un email de test à l'utilisateur pour vérifier la livraison."""
    u = db.get(Utilisateur, uid)
    if not u:
        raise HTTPException(404, "Utilisateur introuvable")
    if not u.email:
        raise HTTPException(400, "Cet utilisateur n'a pas d'adresse email.")
    bad_tlds = {".local", ".localhost", ".internal", ".test", ".example", ".invalid", ".lan"}
    if any(u.email.lower().endswith(t) for t in bad_tlds):
        raise HTTPException(400, f"Domaine email invalide ({u.email}) — mettez à jour l'email.")
    try:
        send_welcome_email(u.nom_complet or u.username, u.email, u.username)
    except RuntimeError as e:
        raise HTTPException(503, str(e))
    return {"ok": True, "email": u.email}


# ══════════════════════════════════════════════════════════════════
# GESTION DES SESSIONS ACTIVES
# ══════════════════════════════════════════════════════════════════

@router.get("/sessions")
def list_sessions(
    _admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    """Liste toutes les sessions actives avec info utilisateur."""
    now = datetime.now(timezone.utc)
    rows = (
        db.query(SessionToken)
        .filter(SessionToken.expires_at > now)
        .order_by(desc(SessionToken.created_at))
        .all()
    )
    result = []
    for s in rows:
        u = db.get(Utilisateur, s.user_id)
        if not u:
            continue
        exp     = s.expires_at.replace(tzinfo=timezone.utc) if s.expires_at.tzinfo is None else s.expires_at
        created = s.created_at.replace(tzinfo=timezone.utc) if s.created_at.tzinfo is None else s.created_at
        last_act = None
        if s.last_activity_at:
            last_act = s.last_activity_at.replace(tzinfo=timezone.utc) if s.last_activity_at.tzinfo is None else s.last_activity_at
        result.append({
            "session_id":       s.id,
            "user_id":          u.id,
            "nom_complet":      u.nom_complet or u.username,
            "email":            u.email,
            "role":             u.role,
            "ip_address":       s.ip_address,
            "user_agent":       s.user_agent,
            "created_at":       created.isoformat(),
            "expires_at":       exp.isoformat(),
            "last_activity_at": last_act.isoformat() if last_act else None,
        })
    return result


@router.delete("/sessions/purge-inactive")
def purge_inactive_sessions(
    request: Request,
    admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    """Supprime les sessions expirées et celles sans activité depuis plus de 2 heures."""
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(hours=2)

    # Sessions expirées
    expired = db.query(SessionToken).filter(SessionToken.expires_at <= now)
    count_expired = expired.count()
    expired.delete(synchronize_session=False)

    # Sessions abandonnées (last_activity_at > 2h et non nulles)
    abandoned = db.query(SessionToken).filter(
        SessionToken.last_activity_at != None,
        SessionToken.last_activity_at < cutoff,
    )
    count_abandoned = abandoned.count()
    abandoned.delete(synchronize_session=False)

    db.commit()
    total = count_expired + count_abandoned
    log_event(db, SESSION_REVOKED, user_id=admin.id,
              ip_address=request.client.host if request.client else None,
              details={"purge_expirees": count_expired, "purge_abandonnees": count_abandoned})
    return {"ok": True, "sessions_supprimees": total}


@router.delete("/sessions/{session_id}")
def revoke_session(
    session_id: int,
    request: Request,
    admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    """Déconnecte une session spécifique."""
    s = db.get(SessionToken, session_id)
    if not s:
        raise HTTPException(404, "Session introuvable")
    target_uid = s.user_id
    db.delete(s)
    db.commit()
    log_event(db, SESSION_REVOKED, user_id=admin.id, target_user_id=target_uid,
              ip_address=request.client.host if request.client else None,
              details={"session_id": session_id})
    return {"ok": True}


@router.delete("/sessions/user/{uid}")
def revoke_user_sessions(
    uid: int,
    request: Request,
    admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    """Déconnecte toutes les sessions d'un utilisateur."""
    u = db.get(Utilisateur, uid)
    if not u:
        raise HTTPException(404, "Utilisateur introuvable")
    count = db.query(SessionToken).filter(SessionToken.user_id == uid).delete()
    db.commit()
    log_event(db, SESSIONS_CLEARED, user_id=admin.id, target_user_id=uid,
              ip_address=request.client.host if request.client else None,
              details={"sessions_supprimees": count})
    return {"ok": True, "sessions_supprimees": count}


# ══════════════════════════════════════════════════════════════════
# SÉCURITÉ — CONNEXIONS (photo + géolocalisation)
# ══════════════════════════════════════════════════════════════════

@router.get("/login-events")
def list_login_events(
    _admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
    limit: int = Query(100, ge=1, le=500),
):
    """Liste les événements de connexion avec photo et localisation."""
    rows = (
        db.query(LoginSecurityEvent)
        .order_by(desc(LoginSecurityEvent.created_at))
        .limit(limit)
        .all()
    )
    result = []
    for e in rows:
        u = db.get(Utilisateur, e.user_id)
        ts = e.created_at
        if ts and ts.tzinfo is None:
            from datetime import timezone as _tz
            ts = ts.replace(tzinfo=_tz.utc)
        result.append({
            "id":         e.id,
            "user_id":    e.user_id,
            "nom_complet": u.nom_complet if u else "—",
            "email":      u.email if u else "—",
            "role":       u.role if u else "—",
            "has_photo":  bool(e.photo_b64),
            "photo_b64":  e.photo_b64,
            "latitude":   e.latitude,
            "longitude":  e.longitude,
            "ip_address": e.ip_address,
            "user_agent": e.user_agent,
            "created_at": ts.isoformat() if ts else None,
        })
    return result


@router.delete("/login-events/{event_id}")
def delete_login_event(
    event_id: int,
    _admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    """Supprime un événement de connexion (et sa photo)."""
    e = db.get(LoginSecurityEvent, event_id)
    if not e:
        raise HTTPException(404, "Événement introuvable")
    db.delete(e)
    db.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════
# JOURNAL D'ACTIVITÉ
# ══════════════════════════════════════════════════════════════════

_ACTION_LABELS = {
    "login_success":    "Connexion réussie",
    "login_failed":     "Échec de connexion",
    "logout":           "Déconnexion",
    "otp_sent":         "Code OTP envoyé",
    "otp_verified":     "Code OTP vérifié",
    "otp_failed":       "Code OTP incorrect",
    "session_revoked":  "Session révoquée",
    "sessions_cleared": "Toutes les sessions supprimées",
    "password_reset":   "Mot de passe réinitialisé",
    "pin_reset":        "Code PIN réinitialisé",
    "user_created":     "Compte créé",
    "user_updated":     "Compte modifié",
    "user_disabled":    "Compte désactivé",
    "user_enabled":     "Compte activé",
}


@router.get("/audit-log/export")
def export_audit_log_xlsx(
    action:  Optional[str] = Query(None),
    user_id: Optional[int] = Query(None),
    _admin:  Utilisateur   = Depends(_require_admin),
    db:      Session       = Depends(get_db),
):
    """Exporte tout le journal d'activité filtré en fichier Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q = db.query(AuditLog).order_by(desc(AuditLog.created_at))
    if action:
        q = q.filter(AuditLog.action == action)
    if user_id:
        q = q.filter(
            (AuditLog.user_id == user_id) | (AuditLog.target_user_id == user_id)
        )
    rows = q.limit(10_000).all()

    # ── Styles ────────────────────────────────────────────────────────
    NAVY  = "0F172A"
    AMBER = "B45309"
    LGRAY = "F1F5F9"
    WHITE = "FFFFFF"

    def _hf(c=WHITE):   return Font(bold=True, color=c, name="Calibri", size=10)
    def _nf(bold=False): return Font(bold=bold,  name="Calibri", size=10)
    def _fill(h):        return PatternFill("solid", fgColor=h)
    def _brd():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    def _ctr():          return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def _lft():          return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal d'Activité"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # ── En-tête ───────────────────────────────────────────────────────
    headers = ["Date & Heure", "Action", "Code Action", "Utilisateur", "Cible", "Adresse IP", "Détails"]
    widths  = [20,             28,       18,            22,            22,      16,            40]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = _hf()
        c.fill      = _fill(NAVY)
        c.alignment = _ctr()
        c.border    = _brd()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    # ── Données ───────────────────────────────────────────────────────
    ACTION_COLORS = {
        "login_success":    "15803D",
        "login_failed":     "B91C1C",
        "logout":           "334155",
        "otp_sent":         "1D4ED8",
        "otp_verified":     "0369A1",
        "otp_failed":       "C2410C",
        "session_revoked":  "7C3AED",
        "sessions_cleared": "6D28D9",
        "password_reset":   "0E7490",
        "pin_reset":        "0E7490",
        "user_created":     "166534",
        "user_updated":     "1E40AF",
        "user_disabled":    "92400E",
        "user_enabled":     "065F46",
    }

    for ri, r in enumerate(rows, 2):
        actor  = db.get(Utilisateur, r.user_id)        if r.user_id        else None
        target = db.get(Utilisateur, r.target_user_id) if r.target_user_id else None
        created = r.created_at.replace(tzinfo=timezone.utc) if r.created_at.tzinfo is None else r.created_at
        dt_local = created.strftime("%Y-%m-%d %H:%M:%S")

        details_dict = json.loads(r.details) if r.details else {}
        details_str  = " | ".join(f"{k}: {v}" for k, v in details_dict.items()) if details_dict else ""

        bg = "F8FAFC" if ri % 2 == 0 else WHITE
        action_col = ACTION_COLORS.get(r.action, "334155")

        vals = [
            dt_local,
            _ACTION_LABELS.get(r.action, r.action),
            r.action,
            actor.nom_complet  or actor.email  if actor  else "Système",
            target.nom_complet or target.email if target else "",
            r.ip_address or "",
            details_str,
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = _brd()
            cell.alignment = _ctr() if ci in (1, 3, 6) else _lft()
            cell.fill      = _fill(bg)
            if ci == 2:  # Action label colorée
                cell.font = Font(bold=True, color=action_col, name="Calibri", size=10)
            else:
                cell.font = _nf()

    # ── Figer + auto-filtre ───────────────────────────────────────────
    ws.auto_filter.ref = f"A1:G{max(len(rows)+1, 1)}"

    # ── Onglet récapitulatif ──────────────────────────────────────────
    ws2 = wb.create_sheet("Résumé")
    ws2.sheet_view.showGridLines = False
    from collections import Counter
    counts = Counter(r.action for r in rows)
    ws2.cell(1, 1, "Action").font = _hf(); ws2.cell(1, 1).fill = _fill(NAVY); ws2.cell(1, 1).alignment = _ctr()
    ws2.cell(1, 2, "Libellé").font = _hf(); ws2.cell(1, 2).fill = _fill(NAVY); ws2.cell(1, 2).alignment = _ctr()
    ws2.cell(1, 3, "Occurrences").font = _hf(); ws2.cell(1, 3).fill = _fill(NAVY); ws2.cell(1, 3).alignment = _ctr()
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 14
    for ri2, (act, cnt) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), 2):
        bg2 = "F8FAFC" if ri2 % 2 == 0 else WHITE
        for ci2, val2 in enumerate([act, _ACTION_LABELS.get(act, act), cnt], 1):
            cell2 = ws2.cell(ri2, ci2, val2)
            cell2.fill = _fill(bg2); cell2.border = _brd()
            cell2.alignment = _ctr() if ci2 == 3 else _lft()
            cell2.font = _nf(ci2 == 3)

    # ── Stream ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import date as date_type
    fname = f"journal_activite_{date_type.today().isoformat()}.xlsx"
    if action:
        fname = f"journal_{action}_{date_type.today().isoformat()}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/audit-log")
def get_audit_log(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    action: Optional[str] = Query(None),
    user_id: Optional[int] = Query(None),
    _admin: Utilisateur = Depends(_require_admin),
    db: Session = Depends(get_db),
):
    """Journal d'activité paginé avec filtres optionnels."""
    q = db.query(AuditLog).order_by(desc(AuditLog.created_at))
    if action:
        q = q.filter(AuditLog.action == action)
    if user_id:
        q = q.filter(
            (AuditLog.user_id == user_id) | (AuditLog.target_user_id == user_id)
        )
    total = q.count()
    rows  = q.offset((page - 1) * per_page).limit(per_page).all()

    items = []
    for r in rows:
        actor  = db.get(Utilisateur, r.user_id)        if r.user_id        else None
        target = db.get(Utilisateur, r.target_user_id) if r.target_user_id else None
        created = r.created_at.replace(tzinfo=timezone.utc) if r.created_at.tzinfo is None else r.created_at
        items.append({
            "id":           r.id,
            "action":       r.action,
            "action_label": _ACTION_LABELS.get(r.action, r.action),
            "actor":        actor.nom_complet or actor.email if actor else "Système",
            "actor_id":     r.user_id,
            "target":       target.nom_complet or target.email if target else None,
            "target_id":    r.target_user_id,
            "ip_address":   r.ip_address,
            "details":      json.loads(r.details) if r.details else None,
            "created_at":   created.isoformat(),
        })
    return {"total": total, "page": page, "per_page": per_page, "items": items}
