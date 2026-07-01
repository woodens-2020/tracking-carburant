"""
Routes API section Hôtel — préfixe /api/hotel
Chambres · Employés · Réservations (nuit / moment)
"""
from __future__ import annotations

from datetime import datetime, timezone, timedelta
from decimal import Decimal
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func
from sqlalchemy.orm import Session

from database import get_db
from models import HotelChambre, HotelEmploye, HotelReservation

router = APIRouter(prefix="/api/hotel", tags=["Hotel"])


def _uid(request: Request) -> int | None:
    u = getattr(request.state, "user", None)
    return u.id if u else None


def _d(v) -> Decimal:
    return Decimal(str(v)) if v is not None else Decimal("0")


# ══════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════

def _chambre_dict(c: HotelChambre, nb_actives: int = 0) -> dict:
    return {
        "id":           c.id,
        "numero":       c.numero,
        "type_chambre": c.type_chambre,
        "etage":        c.etage,
        "capacite":     c.capacite,
        "prix_nuit":    float(_d(c.prix_nuit)),
        "prix_moment":  float(_d(c.prix_moment)) if c.prix_moment else None,
        "statut":       c.statut,
        "description":  c.description or "",
        "actif":        c.actif,
        "nb_reservations_actives": nb_actives,
    }


def _employe_dict(e: HotelEmploye) -> dict:
    return {
        "id":            e.id,
        "nom":           e.nom,
        "prenom":        e.prenom,
        "nom_complet":   f"{e.prenom} {e.nom}",
        "poste":         e.poste,
        "telephone":     e.telephone or "",
        "email":         e.email or "",
        "date_embauche": str(e.date_embauche) if e.date_embauche else None,
        "salaire_base":  float(_d(e.salaire_base)) if e.salaire_base else None,
        "actif":         e.actif,
        "notes":         e.notes or "",
    }


def _res_dict(r: HotelReservation) -> dict:
    return {
        "id":                 r.id,
        "chambre_id":         r.chambre_id,
        "chambre_numero":     r.chambre.numero if r.chambre else str(r.chambre_id),
        "chambre_type":       r.chambre.type_chambre if r.chambre else "",
        "client_nom":         r.client_nom,
        "client_contact":     r.client_contact or "",
        "client_id_piece":    r.client_id_piece or "",
        "type_sejour":        r.type_sejour,
        "date_arrivee":       r.date_arrivee.isoformat(),
        "date_depart_prevue": r.date_depart_prevue.isoformat(),
        "date_depart_reel":   r.date_depart_reel.isoformat() if r.date_depart_reel else None,
        "nb_nuits":           r.nb_nuits,
        "nb_heures":          float(_d(r.nb_heures)) if r.nb_heures else None,
        "prix_unitaire":      float(_d(r.prix_unitaire)),
        "montant_total":      float(_d(r.montant_total)),
        "montant_paye":       float(_d(r.montant_paye)),
        "solde":              float(_d(r.solde)),
        "statut":             r.statut,
        "mode_paiement":      r.mode_paiement or "",
        "notes":              r.notes or "",
        "employe_id":         r.employe_id,
        "employe_nom":        (f"{r.employe.prenom} {r.employe.nom}") if r.employe else None,
        "created_at":         r.created_at.isoformat(),
    }


# ══════════════════════════════════════════════════════════════════
# CHAMBRES
# ══════════════════════════════════════════════════════════════════

class ChambreIn(BaseModel):
    numero:       str
    type_chambre: str = "SIMPLE"
    etage:        Optional[int]   = None
    capacite:     int             = 1
    prix_nuit:    float           = Field(gt=0)
    prix_moment:  Optional[float] = Field(default=None, gt=0)
    description:  Optional[str]   = None
    actif:        bool            = True


@router.get("/chambres")
def liste_chambres(actif: Optional[bool] = Query(default=None), db: Session = Depends(get_db)):
    q = db.query(HotelChambre)
    if actif is not None:
        q = q.filter(HotelChambre.actif == actif)
    chambres = q.order_by(HotelChambre.numero).all()
    actives_map = {
        r.chambre_id: r.count
        for r in db.query(
            HotelReservation.chambre_id,
            func.count(HotelReservation.id).label("count")
        ).filter(HotelReservation.statut == "EN_COURS")
         .group_by(HotelReservation.chambre_id)
         .all()
    }
    return [_chambre_dict(c, actives_map.get(c.id, 0)) for c in chambres]


@router.get("/chambres/disponibles")
def chambres_disponibles(db: Session = Depends(get_db)):
    chambres = (
        db.query(HotelChambre)
        .filter(HotelChambre.statut == "DISPONIBLE", HotelChambre.actif == True)
        .order_by(HotelChambre.numero)
        .all()
    )
    return [_chambre_dict(c) for c in chambres]


@router.post("/chambres", status_code=201)
def creer_chambre(data: ChambreIn, db: Session = Depends(get_db)):
    num = data.numero.strip().upper()
    if db.query(HotelChambre).filter_by(numero=num).first():
        raise HTTPException(409, f"Chambre « {num} » existe déjà.")
    if data.type_chambre not in ("SIMPLE", "DOUBLE", "SUITE", "VIP"):
        raise HTTPException(422, "type_chambre doit être SIMPLE, DOUBLE, SUITE ou VIP.")
    c = HotelChambre(
        numero       = num,
        type_chambre = data.type_chambre,
        etage        = data.etage,
        capacite     = data.capacite,
        prix_nuit    = Decimal(str(data.prix_nuit)),
        prix_moment  = Decimal(str(data.prix_moment)) if data.prix_moment else None,
        statut       = "DISPONIBLE",
        description  = data.description,
        actif        = data.actif,
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return _chambre_dict(c)


@router.put("/chambres/{chambre_id}")
def modifier_chambre(chambre_id: int, data: ChambreIn, db: Session = Depends(get_db)):
    c = db.query(HotelChambre).filter_by(id=chambre_id).first()
    if not c:
        raise HTTPException(404, "Chambre introuvable.")
    num = data.numero.strip().upper()
    doublon = db.query(HotelChambre).filter(
        HotelChambre.numero == num, HotelChambre.id != chambre_id
    ).first()
    if doublon:
        raise HTTPException(409, f"Chambre « {num} » existe déjà.")
    c.numero       = num
    c.type_chambre = data.type_chambre
    c.etage        = data.etage
    c.capacite     = data.capacite
    c.prix_nuit    = Decimal(str(data.prix_nuit))
    c.prix_moment  = Decimal(str(data.prix_moment)) if data.prix_moment else None
    c.description  = data.description
    c.actif        = data.actif
    db.commit()
    return _chambre_dict(c)


@router.patch("/chambres/{chambre_id}/statut")
def changer_statut_chambre(chambre_id: int, statut: str = Query(...), db: Session = Depends(get_db)):
    if statut not in ("DISPONIBLE", "OCCUPEE", "MAINTENANCE", "FERMEE"):
        raise HTTPException(422, "Statut invalide.")
    c = db.query(HotelChambre).filter_by(id=chambre_id).first()
    if not c:
        raise HTTPException(404, "Chambre introuvable.")
    c.statut = statut
    db.commit()
    return {"ok": True, "statut": c.statut}


@router.delete("/chambres/{chambre_id}", status_code=200)
def supprimer_chambre(chambre_id: int, db: Session = Depends(get_db)):
    from sqlalchemy.exc import IntegrityError
    c = db.query(HotelChambre).filter_by(id=chambre_id).first()
    if not c:
        raise HTTPException(404, "Chambre introuvable.")
    nb = db.query(HotelReservation).filter_by(chambre_id=chambre_id).count()
    if nb > 0:
        raise HTTPException(409, f"Impossible : {nb} réservation(s) liées à cette chambre.")
    try:
        db.delete(c)
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "Chambre utilisée — impossible de supprimer.")
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════
# EMPLOYÉS HÔTEL
# ══════════════════════════════════════════════════════════════════

class EmployeHotelIn(BaseModel):
    nom:           str
    prenom:        str
    poste:         str = "RECEPTIONNISTE"
    telephone:     Optional[str]  = None
    email:         Optional[str]  = None
    date_embauche: Optional[str]  = None
    salaire_base:  Optional[float] = None
    actif:         bool           = True
    notes:         Optional[str]  = None


@router.get("/employes")
def liste_employes(actif: Optional[bool] = Query(default=None), db: Session = Depends(get_db)):
    q = db.query(HotelEmploye)
    if actif is not None:
        q = q.filter(HotelEmploye.actif == actif)
    return [_employe_dict(e) for e in q.order_by(HotelEmploye.nom).all()]


@router.post("/employes", status_code=201)
def creer_employe(data: EmployeHotelIn, db: Session = Depends(get_db)):
    postes_valides = ("RECEPTIONNISTE", "FEMME_DE_CHAMBRE", "GERANT", "SECURITE", "AUTRE")
    if data.poste not in postes_valides:
        raise HTTPException(422, f"poste doit être parmi {postes_valides}.")
    from datetime import date as dt_type
    emb = None
    if data.date_embauche:
        try:
            emb = dt_type.fromisoformat(data.date_embauche)
        except ValueError:
            raise HTTPException(422, "date_embauche invalide (YYYY-MM-DD).")
    e = HotelEmploye(
        nom           = data.nom.strip(),
        prenom        = data.prenom.strip(),
        poste         = data.poste,
        telephone     = data.telephone,
        email         = data.email,
        date_embauche = emb,
        salaire_base  = Decimal(str(data.salaire_base)) if data.salaire_base else None,
        actif         = data.actif,
        notes         = data.notes,
    )
    db.add(e)
    db.commit()
    db.refresh(e)
    return _employe_dict(e)


@router.put("/employes/{employe_id}")
def modifier_employe(employe_id: int, data: EmployeHotelIn, db: Session = Depends(get_db)):
    e = db.query(HotelEmploye).filter_by(id=employe_id).first()
    if not e:
        raise HTTPException(404, "Employé introuvable.")
    from datetime import date as dt_type
    emb = None
    if data.date_embauche:
        try:
            emb = dt_type.fromisoformat(data.date_embauche)
        except ValueError:
            raise HTTPException(422, "date_embauche invalide.")
    e.nom           = data.nom.strip()
    e.prenom        = data.prenom.strip()
    e.poste         = data.poste
    e.telephone     = data.telephone
    e.email         = data.email
    e.date_embauche = emb
    e.salaire_base  = Decimal(str(data.salaire_base)) if data.salaire_base else None
    e.actif         = data.actif
    e.notes         = data.notes
    db.commit()
    return _employe_dict(e)


@router.delete("/employes/{employe_id}", status_code=200)
def supprimer_employe(employe_id: int, db: Session = Depends(get_db)):
    e = db.query(HotelEmploye).filter_by(id=employe_id).first()
    if not e:
        raise HTTPException(404, "Employé introuvable.")
    e.actif = False
    db.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════
# RÉSERVATIONS
# ══════════════════════════════════════════════════════════════════

class ReservationIn(BaseModel):
    chambre_id:      int
    client_nom:      str
    client_contact:  Optional[str] = None
    client_id_piece: str
    type_sejour:     str           = "NUIT"      # NUIT | MOMENT
    date_arrivee:    str                         # ISO datetime
    nb_nuits:        Optional[int]  = None       # si NUIT
    nb_heures:       Optional[float] = None      # si MOMENT
    montant_paye:    float          = 0
    mode_paiement:   Optional[str]  = None
    employe_id:      Optional[int]  = None
    notes:           Optional[str]  = None


class PaiementIn(BaseModel):
    montant:       float = Field(gt=0)
    mode_paiement: Optional[str] = None


@router.get("/reservations")
def liste_reservations(
    statut:     Optional[str] = Query(default=None),
    chambre_id: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
):
    q = db.query(HotelReservation).order_by(HotelReservation.date_arrivee.desc())
    if statut:
        q = q.filter(HotelReservation.statut == statut.upper())
    if chambre_id:
        q = q.filter(HotelReservation.chambre_id == chambre_id)
    return [_res_dict(r) for r in q.limit(300).all()]


@router.get("/reservations/en-cours")
def reservations_en_cours(db: Session = Depends(get_db)):
    res = (
        db.query(HotelReservation)
        .filter(HotelReservation.statut == "EN_COURS")
        .order_by(HotelReservation.date_arrivee)
        .all()
    )
    return [_res_dict(r) for r in res]


@router.get("/reservations/{res_id}")
def detail_reservation(res_id: int, db: Session = Depends(get_db)):
    r = db.query(HotelReservation).filter_by(id=res_id).first()
    if not r:
        raise HTTPException(404, "Réservation introuvable.")
    return _res_dict(r)


@router.post("/reservations", status_code=201)
def creer_reservation(data: ReservationIn, request: Request, db: Session = Depends(get_db)):
    chambre = db.query(HotelChambre).filter_by(id=data.chambre_id).first()
    if not chambre:
        raise HTTPException(404, "Chambre introuvable.")
    if chambre.statut != "DISPONIBLE":
        raise HTTPException(409, f"Chambre {chambre.numero} n'est pas disponible (statut : {chambre.statut}).")

    if not data.client_nom.strip():
        raise HTTPException(422, "Nom du client requis.")
    if not data.client_id_piece.strip():
        raise HTTPException(422, "NIF / Pièce d'identité requise.")
    if data.type_sejour not in ("NUIT", "MOMENT"):
        raise HTTPException(422, "type_sejour doit être NUIT ou MOMENT.")

    try:
        date_arr = datetime.fromisoformat(data.date_arrivee.replace("Z", "+00:00"))
        if date_arr.tzinfo is None:
            date_arr = date_arr.replace(tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(422, "date_arrivee invalide (ISO 8601).")

    if data.type_sejour == "NUIT":
        if not data.nb_nuits or data.nb_nuits < 1:
            raise HTTPException(422, "nb_nuits requis (≥ 1) pour séjour NUIT.")
        prix_unit = _d(chambre.prix_nuit)
        nb_nuits  = data.nb_nuits
        nb_heures = None
        date_dep  = date_arr + timedelta(days=nb_nuits)
        montant   = prix_unit * nb_nuits
    else:  # MOMENT
        if not data.nb_heures or data.nb_heures <= 0:
            raise HTTPException(422, "nb_heures requis (> 0) pour séjour MOMENT.")
        if not chambre.prix_moment:
            raise HTTPException(422, f"La chambre {chambre.numero} n'a pas de prix moment configuré.")
        prix_unit = _d(chambre.prix_moment)
        nb_nuits  = None
        nb_heures = Decimal(str(data.nb_heures))
        date_dep  = date_arr + timedelta(hours=float(nb_heures))
        montant   = prix_unit * nb_heures

    montant_paye = min(Decimal(str(data.montant_paye)), montant)
    solde        = montant - montant_paye

    r = HotelReservation(
        chambre_id         = chambre.id,
        client_nom         = data.client_nom.strip(),
        client_contact     = data.client_contact,
        client_id_piece    = data.client_id_piece,
        type_sejour        = data.type_sejour,
        date_arrivee       = date_arr,
        date_depart_prevue = date_dep,
        nb_nuits           = nb_nuits,
        nb_heures          = nb_heures,
        prix_unitaire      = prix_unit,
        montant_total      = montant,
        montant_paye       = montant_paye,
        solde              = solde,
        statut             = "EN_COURS",
        mode_paiement      = data.mode_paiement,
        employe_id         = data.employe_id,
        notes              = data.notes,
    )
    db.add(r)

    # Marquer la chambre comme occupée
    chambre.statut = "OCCUPEE"
    db.commit()
    db.refresh(r)
    return _res_dict(r)


@router.post("/reservations/{res_id}/paiement", status_code=200)
def ajouter_paiement(res_id: int, data: PaiementIn, db: Session = Depends(get_db)):
    r = db.query(HotelReservation).filter_by(id=res_id).first()
    if not r:
        raise HTTPException(404, "Réservation introuvable.")
    if r.statut != "EN_COURS":
        raise HTTPException(409, "Réservation non active.")
    montant = Decimal(str(data.montant))
    r.montant_paye += montant
    r.solde = max(Decimal("0"), r.montant_total - r.montant_paye)
    if data.mode_paiement:
        r.mode_paiement = data.mode_paiement
    db.commit()
    return _res_dict(r)


@router.post("/reservations/{res_id}/terminer", status_code=200)
def terminer_reservation(res_id: int, db: Session = Depends(get_db)):
    r = db.query(HotelReservation).filter_by(id=res_id).first()
    if not r:
        raise HTTPException(404, "Réservation introuvable.")
    if r.statut != "EN_COURS":
        raise HTTPException(409, "Réservation déjà terminée ou annulée.")
    r.statut          = "TERMINEE"
    r.date_depart_reel = datetime.now(tz=timezone.utc)
    # Remettre la chambre disponible
    if r.chambre:
        r.chambre.statut = "DISPONIBLE"
    db.commit()
    return _res_dict(r)


@router.post("/reservations/{res_id}/annuler", status_code=200)
def annuler_reservation(res_id: int, db: Session = Depends(get_db)):
    r = db.query(HotelReservation).filter_by(id=res_id).first()
    if not r:
        raise HTTPException(404, "Réservation introuvable.")
    if r.statut != "EN_COURS":
        raise HTTPException(409, "Réservation déjà terminée ou annulée.")
    r.statut = "ANNULEE"
    r.date_depart_reel = datetime.now(tz=timezone.utc)
    if r.chambre:
        r.chambre.statut = "DISPONIBLE"
    db.commit()
    return _res_dict(r)


# ══════════════════════════════════════════════════════════════════
# STATS / DASHBOARD
# ══════════════════════════════════════════════════════════════════

@router.get("/stats")
def stats_hotel(db: Session = Depends(get_db)):
    total_chambres   = db.query(HotelChambre).filter_by(actif=True).count()
    chambres_occup   = db.query(HotelChambre).filter_by(statut="OCCUPEE").count()
    chambres_dispo   = db.query(HotelChambre).filter_by(statut="DISPONIBLE", actif=True).count()
    sejours_actifs   = db.query(HotelReservation).filter_by(statut="EN_COURS").count()
    revenu_mois      = db.query(func.sum(HotelReservation.montant_paye)).filter(
        HotelReservation.statut.in_(["EN_COURS", "TERMINEE"]),
    ).scalar() or Decimal("0")
    solde_total      = db.query(func.sum(HotelReservation.solde)).filter(
        HotelReservation.statut == "EN_COURS"
    ).scalar() or Decimal("0")
    return {
        "total_chambres":  total_chambres,
        "chambres_occupees": chambres_occup,
        "chambres_dispo":  chambres_dispo,
        "taux_occupation": round(chambres_occup / total_chambres * 100, 1) if total_chambres else 0,
        "sejours_actifs":  sejours_actifs,
        "revenu_total":    float(revenu_mois),
        "solde_en_attente": float(solde_total),
    }


def _get_rapport_data(db: Session, date_debut: Optional[str], date_fin: Optional[str]) -> dict:
    from datetime import date as date_type
    from collections import defaultdict

    today = datetime.now(timezone.utc).date()
    try:
        d_debut = datetime.strptime(date_debut, "%Y-%m-%d").date() if date_debut else date_type(today.year, today.month, 1)
    except ValueError:
        d_debut = date_type(today.year, today.month, 1)
    try:
        d_fin = datetime.strptime(date_fin, "%Y-%m-%d").date() if date_fin else today
    except ValueError:
        d_fin = today

    dt_debut = datetime(d_debut.year, d_debut.month, d_debut.day, tzinfo=timezone.utc)
    dt_fin   = datetime(d_fin.year,   d_fin.month,   d_fin.day, 23, 59, 59, tzinfo=timezone.utc)

    reservations = (
        db.query(HotelReservation)
        .filter(HotelReservation.date_arrivee >= dt_debut,
                HotelReservation.date_arrivee <= dt_fin)
        .all()
    )

    moments = [r for r in reservations if r.type_sejour == "MOMENT"]
    nuits   = [r for r in reservations if r.type_sejour == "NUIT"]

    def _sum(lst, field):
        return float(sum(_d(getattr(r, field) or 0) for r in lst))

    par_chambre: dict = defaultdict(lambda: {"nb": 0, "revenu": 0.0, "moments": 0, "nuits": 0})
    for r in reservations:
        ch = r.chambre.numero if r.chambre else str(r.chambre_id)
        par_chambre[ch]["nb"]     += 1
        par_chambre[ch]["revenu"] += float(_d(r.montant_paye or 0))
        par_chambre[ch]["moments"] += 1 if r.type_sejour == "MOMENT" else 0
        par_chambre[ch]["nuits"]   += 1 if r.type_sejour == "NUIT"   else 0

    dt_today     = datetime(today.year, today.month, today.day, tzinfo=timezone.utc)
    dt_today_end = datetime(today.year, today.month, today.day, 23, 59, 59, tzinfo=timezone.utc)
    today_res = (
        db.query(HotelReservation)
        .filter(HotelReservation.date_arrivee >= dt_today,
                HotelReservation.date_arrivee <= dt_today_end)
        .all()
    )

    actifs_nb    = db.query(HotelReservation).filter_by(statut="EN_COURS").count()
    actifs_solde = float(
        db.query(func.sum(HotelReservation.solde))
          .filter_by(statut="EN_COURS").scalar() or 0
    )

    return {
        "periode": {"debut": str(d_debut), "fin": str(d_fin)},
        "kpis": {
            "nb_total":              len(reservations),
            "nb_moments":            len(moments),
            "nb_nuits":              len(nuits),
            "revenu_total":          _sum(reservations, "montant_paye"),
            "revenu_moments":        _sum(moments,      "montant_paye"),
            "revenu_nuits":          _sum(nuits,        "montant_paye"),
            "montant_total_facture": _sum(reservations, "montant_total"),
            "solde_impaye":          _sum(reservations, "solde"),
        },
        "aujourd_hui": {
            "nb":      len(today_res),
            "moments": sum(1 for r in today_res if r.type_sejour == "MOMENT"),
            "nuits":   sum(1 for r in today_res if r.type_sejour == "NUIT"),
            "revenu":  sum(float(_d(r.montant_paye or 0)) for r in today_res),
        },
        "actifs": {
            "nb":               actifs_nb,
            "solde_en_attente": actifs_solde,
        },
        "par_chambre": sorted(
            [{"numero": ch, **v} for ch, v in par_chambre.items()],
            key=lambda x: x["revenu"], reverse=True
        ),
    }


@router.get("/rapport")
def rapport_hotel(
    date_debut: Optional[str] = Query(default=None),
    date_fin:   Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    return _get_rapport_data(db, date_debut, date_fin)


@router.get("/rapport/pdf")
def rapport_hotel_pdf(
    date_debut: Optional[str] = Query(default=None),
    date_fin:   Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER

    data = _get_rapport_data(db, date_debut, date_fin)
    buf  = io.BytesIO()

    doc  = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    styles    = getSampleStyleSheet()
    s_title   = ParagraphStyle("title", parent=styles["Title"], fontSize=18, spaceAfter=4)
    s_sub     = ParagraphStyle("sub",   parent=styles["Normal"], fontSize=10, textColor=colors.grey, spaceAfter=10)
    s_section = ParagraphStyle("sec",   parent=styles["Heading2"], fontSize=11, spaceBefore=14, spaceAfter=6, textColor=colors.HexColor("#1e3a5f"))

    HDR_BG = colors.HexColor("#1e3a5f")
    ALT_BG = colors.HexColor("#f0f4f8")
    GRID_C = colors.HexColor("#cccccc")

    def _gdes(v):
        return f"G {float(v):,.2f}".replace(",", " ")

    def _tbl(header, rows, widths):
        data_t = [header] + rows
        t = Table(data_t, colWidths=widths)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1,  0), HDR_BG),
            ("TEXTCOLOR",     (0, 0), (-1,  0), colors.white),
            ("FONTNAME",      (0, 0), (-1,  0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 9),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, ALT_BG]),
            ("GRID",          (0, 0), (-1, -1), 0.25, GRID_C),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return t

    elems = []
    elems.append(Paragraph("Rapport Hôtel", s_title))
    elems.append(Paragraph(f"Période : {data['periode']['debut']}  →  {data['periode']['fin']}", s_sub))

    auj = data["aujourd_hui"]
    elems.append(Paragraph("Aujourd'hui", s_section))
    elems.append(_tbl(
        ["Arrivées", "Moments", "Nuits", "Revenu"],
        [[str(auj["nb"]), str(auj["moments"]), str(auj["nuits"]), _gdes(auj["revenu"])]],
        [3.5*cm, 3.5*cm, 3.5*cm, 5*cm],
    ))

    k = data["kpis"]
    elems.append(Paragraph(f"Période : {data['periode']['debut']} → {data['periode']['fin']}", s_section))
    elems.append(_tbl(
        ["Total", "Moments", "Nuits", "Revenu total", "Impayé"],
        [[str(k["nb_total"]), str(k["nb_moments"]), str(k["nb_nuits"]),
          _gdes(k["revenu_total"]), _gdes(k["solde_impaye"])]],
        [2.5*cm, 2.5*cm, 2.5*cm, 5*cm, 4*cm],
    ))

    elems.append(Paragraph("Revenus par type", s_section))
    elems.append(_tbl(
        ["Type", "Séjours", "Revenu"],
        [
            ["Moments", str(k["nb_moments"]), _gdes(k["revenu_moments"])],
            ["Nuits",   str(k["nb_nuits"]),   _gdes(k["revenu_nuits"])],
        ],
        [5*cm, 4*cm, 7*cm],
    ))

    if data["par_chambre"]:
        elems.append(Paragraph("Performance par chambre", s_section))
        elems.append(_tbl(
            ["Chambre", "Total", "Moments", "Nuits", "Revenu"],
            [[ch["numero"], str(ch["nb"]), str(ch["moments"]), str(ch["nuits"]), _gdes(ch["revenu"])]
             for ch in data["par_chambre"]],
            [3.5*cm, 2.5*cm, 3*cm, 3*cm, 5*cm],
        ))

    elems.append(Spacer(1, 0.5*cm))
    elems.append(Paragraph(
        f"Généré le {datetime.now(timezone.utc).strftime('%d/%m/%Y à %H:%M')} UTC",
        ParagraphStyle("footer", parent=styles["Normal"], fontSize=8, textColor=colors.grey),
    ))

    doc.build(elems)
    buf.seek(0)
    fname = f"rapport_hotel_{data['periode']['debut']}_{data['periode']['fin']}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/rapport/xlsx")
def rapport_hotel_xlsx(
    date_debut: Optional[str] = Query(default=None),
    date_fin:   Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = _get_rapport_data(db, date_debut, date_fin)

    HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
    ALT_FILL = PatternFill("solid", fgColor="F0F4F8")
    HDR_FONT = Font(bold=True, color="FFFFFF")
    BOLD     = Font(bold=True)
    _side    = Side(style="thin", color="BBBBBB")
    BORDER   = Border(left=_side, right=_side, top=_side, bottom=_side)

    def _gdes(v):
        return f"G {float(v):,.2f}".replace(",", " ")

    def _hdr(ws, row, cols):
        for ci, val in enumerate(cols, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center"); c.border = BORDER

    def _row(ws, row, vals, alt=False):
        fill = ALT_FILL if alt else PatternFill()
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = fill; c.border = BORDER
            c.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")

    wb = openpyxl.Workbook()

    # ── Feuille 1 : Résumé ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 20

    r = 1
    t = ws1.cell(r, 1, "Rapport Hôtel"); t.font = Font(bold=True, size=14)
    ws1.merge_cells(f"A{r}:B{r}"); r += 1
    ws1.cell(r, 1, f"Période : {data['periode']['debut']} → {data['periode']['fin']}")
    ws1.merge_cells(f"A{r}:B{r}"); r += 2

    ws1.cell(r, 1, "Aujourd'hui").font = BOLD; r += 1
    auj = data["aujourd_hui"]
    for lbl, val in [("Arrivées du jour", auj["nb"]), ("Moments", auj["moments"]),
                     ("Nuits", auj["nuits"]), ("Revenu du jour", _gdes(auj["revenu"]))]:
        ws1.cell(r, 1, lbl); ws1.cell(r, 2, val); r += 1

    r += 1
    ws1.cell(r, 1, "Période").font = BOLD; r += 1
    k = data["kpis"]
    for lbl, val in [
        ("Total séjours", k["nb_total"]), ("Moments", k["nb_moments"]), ("Nuits", k["nb_nuits"]),
        ("Revenu total",  _gdes(k["revenu_total"])),
        ("Rev. Moments",  _gdes(k["revenu_moments"])),
        ("Rev. Nuits",    _gdes(k["revenu_nuits"])),
        ("Montant facturé", _gdes(k["montant_total_facture"])),
        ("Impayé",          _gdes(k["solde_impaye"])),
    ]:
        ws1.cell(r, 1, lbl); ws1.cell(r, 2, val); r += 1

    r += 1
    ws1.cell(r, 1, "Actifs en cours").font = BOLD; r += 1
    ac = data["actifs"]
    ws1.cell(r, 1, "Clients en cours"); ws1.cell(r, 2, ac["nb"]); r += 1
    ws1.cell(r, 1, "Solde en attente"); ws1.cell(r, 2, _gdes(ac["solde_en_attente"])); r += 1

    # ── Feuille 2 : Par Chambre ───────────────────────────────────
    ws2 = wb.create_sheet("Par Chambre")
    for i, w in enumerate([14, 10, 12, 10, 20], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    _hdr(ws2, 1, ["Chambre", "Total", "Moments", "Nuits", "Revenu"])
    for idx, ch in enumerate(data["par_chambre"], 2):
        _row(ws2, idx,
             [ch["numero"], ch["nb"], ch["moments"], ch["nuits"], _gdes(ch["revenu"])],
             alt=(idx % 2 == 1))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"rapport_hotel_{data['periode']['debut']}_{data['periode']['fin']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
