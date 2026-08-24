"""
Routes de gestion des sessions de caisse.

Préfixe : /api/pos/caisse

Endpoints caissière :
  GET  /caissieres                         — liste des employés caissiers
  GET  /dashboard?caissier_id=N&date=...   — stats du jour pour une caissière
  GET  /sessions                           — liste des sessions (admin + caissière)
  GET  /sessions/{id}                      — détail d'une session
  POST /sessions/ouvrir                    — ouvrir / retrouver la session du jour
  POST /sessions/{id}/soumettre            — caissière soumet son rapport
  POST /sessions/{id}/valider              — admin valide la session
  GET  /sessions/{id}/rapport.pdf          — export PDF
  GET  /sessions/{id}/rapport.xlsx         — export XLSX
"""
from __future__ import annotations

import io
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from database import get_db
from models import BarVente, BarLigneVente, BarSessionCaisse, BarSessionEvaluation, Employe, Utilisateur
from tz_utils import HAITI_TZ, today_haiti, bounds_haiti

router = APIRouter(prefix="/api/pos/caisse", tags=["caisse"])


# ── helpers ──────────────────────────────────────────────────────────

def _uid(request: Request) -> Optional[int]:
    return getattr(request.state, "utilisateur_id", None)


def _require_pdg_ou_admin_pos(request: Request, db: Session = Depends(get_db)) -> Utilisateur:
    """Même logique que main.require_pdg_ou_admin — dupliquée ici pour
    éviter un import circulaire entre routers (voir pos_routes.py)."""
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(403, "Non autorisé")
    if user.role in ("pdg", "admin"):
        return user
    if user.role_id:
        u = db.get(Utilisateur, user.id)
        if u and u.role_obj and u.role_obj.permissions.get("admin", False):
            return u
    raise HTTPException(403, "Accès réservé au PDG et aux administrateurs")


def _session_ou_404(session_id: int, db: Session) -> BarSessionCaisse:
    s = db.query(BarSessionCaisse).filter_by(id=session_id).first()
    if not s:
        raise HTTPException(404, "Session introuvable")
    return s


def _ventes_session(session: BarSessionCaisse, db: Session):
    from sqlalchemy import func as _f
    today_start, today_end = bounds_haiti(session.date_session)
    return (
        db.query(BarVente)
        .filter(
            BarVente.caissier_id == session.caissier_id,
            BarVente.date_heure  >= today_start,
            BarVente.date_heure  <= today_end,
            BarVente.statut      != "ANNULEE",
        )
        .order_by(BarVente.date_heure)
        .all()
    )


def _cash_remboursements_collectes(employe: Optional[Employe], jour: date, db: Session) -> Decimal:
    """Cash physiquement recu ce jour-la par cette caissiere en reglement
    d'un credit — quelle que soit la date de la vente d'origine (un credit
    peut avoir ete vendu un autre jour). Le paiement d'un credit
    (PUT /ventes/{id}/payer ou POST /credits/{id}/remboursement) ne change
    jamais le mode_paiement ni la date_heure de la vente d'origine, donc ce
    cash ne peut pas etre retrouve via _ventes_session/mode_paiement — il
    est traque separement via bar_remboursements (date_remb + utilisateur_id
    de la personne qui a encaisse), rattache a la caissiere via son compte
    utilisateur lie (Employe.utilisateur_id)."""
    from models import BarRemboursement
    if not employe or not employe.utilisateur_id:
        return Decimal("0")
    jour_start, jour_end = bounds_haiti(jour)
    rembs = (
        db.query(BarRemboursement)
        .filter(
            BarRemboursement.utilisateur_id == employe.utilisateur_id,
            BarRemboursement.date_remb >= jour_start,
            BarRemboursement.date_remb <= jour_end,
        )
        .all()
    )
    return sum((Decimal(str(r.montant)) for r in rembs), Decimal("0"))


def _cash_ventes_decimal(ventes) -> Decimal:
    """Cash CASH/MIXTE d'une liste de ventes, en Decimal exact — une somme en
    flottant binaire ne peut pas représenter exactement la plupart des
    valeurs décimales (0,10 ; 0,20...) et peut dériver de quelques centimes
    une fois additionnée sur toutes les ventes d'une journée à fort volume.
    Utilisé à la fois par _stats_ventes() (affichage) et soumettre_session()
    (stockage de la réconciliation) pour qu'ils calculent identiquement."""
    return sum(
        (Decimal(str(v.montant_paye)) for v in ventes if v.mode_paiement in ("CASH", "MIXTE")),
        Decimal("0"),
    )


def _stats_ventes(ventes, employe: Optional[Employe] = None, jour: Optional[date] = None, db: Session = None):
    total      = sum((Decimal(str(v.montant_total)) for v in ventes), Decimal("0"))
    cash       = _cash_ventes_decimal(ventes)
    if employe is not None and jour is not None and db is not None:
        cash += _cash_remboursements_collectes(employe, jour, db)
    credit_tot = sum((Decimal(str(v.montant_total)) for v in ventes if v.mode_paiement == "CREDIT"), Decimal("0"))
    modes: dict = {}
    for v in ventes:
        modes[v.mode_paiement] = modes.get(v.mode_paiement, Decimal("0")) + Decimal(str(v.montant_total))

    produits: dict = {}
    for v in ventes:
        for l in v.lignes:
            nom = (l.produit.nom if l.produit else None) or (l.cuisine_plat.nom if l.cuisine_plat else "?")
            if nom not in produits:
                produits[nom] = {"nom": nom, "quantite": Decimal("0"), "total": Decimal("0")}
            produits[nom]["quantite"] += Decimal(str(l.quantite))
            produits[nom]["total"]    += Decimal(str(l.sous_total))

    top  = sorted(produits.values(), key=lambda x: x["total"], reverse=True)[:10]
    tous = sorted(produits.values(), key=lambda x: x["nom"])
    return {
        "nb_ventes":  len(ventes),
        "total":      float(total),
        "cash":       float(cash),
        "credit":     float(credit_tot),
        "par_mode":   {k: float(v) for k, v in modes.items()},
        "top_produits": [
            {"nom": p["nom"], "quantite": float(p["quantite"]), "total": float(p["total"])}
            for p in top
        ],
        # Liste complète (non plafonnée), triée par nom — pour le
        # récapitulatif de clôture de session que la caissière confirme
        # article par article, contrairement à top_produits (top 10 par CA).
        "tous_produits": [
            {"nom": p["nom"], "quantite": float(p["quantite"]), "total": float(p["total"])}
            for p in tous
        ],
    }


def _ecart_couleur(ecart: float, cash_attendu: float) -> str:
    """Code hex (sans #) selon la gravité de l'écart de caisse — mêmes seuils
    que le JS _posEcartNiveau() côté frontend, pour que PDF/XLSX/écran
    affichent toujours la même couleur pour le même écart."""
    base = max(abs(cash_attendu or 0), 1000)
    pct  = abs(ecart or 0) / base
    if pct <= 0.01:
        return "22C55E"
    if pct <= 0.03:
        return "F7A93B"
    return "F87171"


def _score_couleur(score: float | None) -> str:
    """Code hex (sans #) selon le score d'évaluation d'un rapport — plus le
    score est élevé, meilleur c'est (contrairement à _ecart_couleur)."""
    if score is None:
        return "888888"
    if score >= 90:
        return "22C55E"
    if score >= 70:
        return "F7A93B"
    return "F87171"


def _session_dict(s: BarSessionCaisse, stats: dict | None = None) -> dict:
    return {
        "id":            s.id,
        "caissier_id":   s.caissier_id,
        "caissier_nom":  (s.caissier.nom + " " + s.caissier.prenom) if s.caissier else None,
        "date_session":  str(s.date_session),
        "statut":        s.statut,
        "created_at":    s.created_at.isoformat() if s.created_at else None,
        "soumis_at":     s.soumis_at.isoformat() if s.soumis_at else None,
        "valide_at":     s.valide_at.isoformat() if s.valide_at else None,
        "valide_par":    (s.valide_par.nom + " " + s.valide_par.prenom) if s.valide_par else None,
        "notes_admin":   s.notes_admin,
        "cash_attendu_soumission": float(s.cash_attendu_soumission) if s.cash_attendu_soumission is not None else None,
        "montant_compte":          float(s.montant_compte) if s.montant_compte is not None else None,
        "ecart":                   float(s.ecart) if s.ecart is not None else None,
        "evaluation_statut": s.evaluation_statut,
        "score":             float(s.score) if s.score is not None else None,
        "evalue_par":        (s.evalue_par.nom + " " + s.evalue_par.prenom) if s.evalue_par else None,
        "evalue_le":         s.evalue_le.isoformat() if s.evalue_le else None,
        **(stats or {}),
    }


# ── endpoints ────────────────────────────────────────────────────────

@router.get("/caissieres")
def liste_caissieres(db: Session = Depends(get_db)):
    """Tous les employés actifs (peuvent être affectés comme caissier)."""
    employes = (
        db.query(Employe)
        .filter(Employe.actif == True)
        .order_by(Employe.nom, Employe.prenom)
        .all()
    )
    return [
        {
            "id":             e.id,
            "nom":            e.nom,
            "prenom":         e.prenom,
            "poste":          e.poste,
            "utilisateur_id": e.utilisateur_id,
        }
        for e in employes
    ]


@router.get("/dashboard")
def dashboard_caissiere(
    caissier_id: int = Query(...),
    date_sel:    str = Query(None, alias="date"),
    db: Session = Depends(get_db),
):
    """Stats du jour (ou date choisie) pour une caissière."""
    if date_sel:
        try:
            jour = date.fromisoformat(date_sel)
        except ValueError:
            raise HTTPException(422, "date invalide (YYYY-MM-DD)")
    else:
        jour = today_haiti()

    employe = db.query(Employe).filter_by(id=caissier_id).first()
    if not employe:
        raise HTTPException(404, "Caissier introuvable")

    dt_start, dt_end = bounds_haiti(jour)

    ventes = (
        db.query(BarVente)
        .filter(
            BarVente.caissier_id == caissier_id,
            BarVente.date_heure  >= dt_start,
            BarVente.date_heure  <= dt_end,
            BarVente.statut      != "ANNULEE",
        )
        .order_by(BarVente.date_heure)
        .all()
    )

    session = (
        db.query(BarSessionCaisse)
        .filter_by(caissier_id=caissier_id, date_session=jour)
        .first()
    )

    # Evolution horaire (pour graphe)
    evolution: list = []
    cumul = 0.0
    for v in ventes:
        cumul += float(v.montant_total)
        evolution.append({
            "heure":   v.date_heure.astimezone(HAITI_TZ).strftime("%H:%M"),
            "montant": float(v.montant_total),
            "cumul":   cumul,
            "ticket":  v.numero_ticket,
        })

    stats = _stats_ventes(ventes, employe, jour, db)
    return {
        "date":          str(jour),
        "caissier_id":   caissier_id,
        "caissier_nom":  employe.nom + " " + employe.prenom,
        "session_id":    session.id if session else None,
        "session_statut": session.statut if session else None,
        "evolution":     evolution,
        **stats,
    }


@router.get("/sessions")
def liste_sessions(
    caissier_id: Optional[int] = Query(None),
    statut:      Optional[str] = Query(None),
    date_debut:  Optional[str] = Query(None),
    date_fin:    Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(BarSessionCaisse)
    if caissier_id:
        q = q.filter(BarSessionCaisse.caissier_id == caissier_id)
    if statut:
        q = q.filter(BarSessionCaisse.statut == statut.upper())
    if date_debut:
        q = q.filter(BarSessionCaisse.date_session >= date.fromisoformat(date_debut))
    if date_fin:
        q = q.filter(BarSessionCaisse.date_session <= date.fromisoformat(date_fin))

    sessions = q.order_by(BarSessionCaisse.date_session.desc(), BarSessionCaisse.id.desc()).all()

    result = []
    for s in sessions:
        ventes = _ventes_session(s, db)
        stats  = _stats_ventes(ventes, s.caissier, s.date_session, db)
        result.append(_session_dict(s, stats))
    return result


@router.get("/sessions/{session_id}")
def detail_session(session_id: int, db: Session = Depends(get_db)):
    s      = _session_ou_404(session_id, db)
    ventes = _ventes_session(s, db)
    stats  = _stats_ventes(ventes, s.caissier, s.date_session, db)
    d      = _session_dict(s, stats)
    d["ventes"] = [
        {
            "id":            v.id,
            "numero_ticket": v.numero_ticket,
            "date_heure":    v.date_heure.isoformat(),
            "montant_total": float(v.montant_total),
            "montant_paye":  float(v.montant_paye),
            "mode_paiement": v.mode_paiement,
            "client_nom":    v.client_nom,
            "lignes": [
                {
                    "produit": (l.produit.nom if l.produit else None) or (l.cuisine_plat.nom if l.cuisine_plat else "?"),
                    "quantite":  float(l.quantite),
                    "prix_unit": float(l.prix_unitaire_applique),
                    "sous_total": float(l.sous_total),
                }
                for l in v.lignes
            ],
        }
        for v in ventes
    ]
    d["evaluations"] = {e.produit_nom: e.statut for e in s.evaluations}
    return d


class OuvrirIn(BaseModel):
    caissier_id: int


@router.post("/sessions/ouvrir")
def ouvrir_session(data: OuvrirIn, db: Session = Depends(get_db)):
    """Ouvre (ou retrouve) la session du jour pour une caissière."""
    employe = db.query(Employe).filter_by(id=data.caissier_id).first()
    if not employe:
        raise HTTPException(404, "Caissier introuvable")

    aujourd_hui = today_haiti()
    session = (
        db.query(BarSessionCaisse)
        .filter_by(caissier_id=data.caissier_id, date_session=aujourd_hui)
        .first()
    )
    if not session:
        session = BarSessionCaisse(
            caissier_id  = data.caissier_id,
            date_session = aujourd_hui,
            statut       = "EN_COURS",
        )
        db.add(session)
        db.commit()
        db.refresh(session)

    ventes = _ventes_session(session, db)
    stats  = _stats_ventes(ventes, session.caissier, session.date_session, db)
    return _session_dict(session, stats)


class SoumettreIn(BaseModel):
    montant_compte: float = Field(..., ge=0)
    notes: Optional[str] = None


@router.post("/sessions/{session_id}/soumettre")
def soumettre_session(session_id: int, body: SoumettreIn, db: Session = Depends(get_db)):
    """Soumet la session : passe en SOUMIS, fige le cash attendu et l'écart.
    Une session ne peut être soumise qu'une fois — au-delà, le snapshot doit
    rester figé pour que la réconciliation ait un sens."""
    s = _session_ou_404(session_id, db)
    if s.statut == "SOUMIS":
        raise HTTPException(409, "Session déjà soumise — contactez un administrateur pour toute correction.")
    if s.statut == "VALIDE":
        raise HTTPException(409, "Session déjà validée par l'admin.")

    ventes = _ventes_session(s, db)
    stats  = _stats_ventes(ventes, s.caissier, s.date_session, db)

    # Calculé en Decimal exact (jamais via stats["cash"], déjà passé par un
    # flottant pour l'affichage) : c'est la donnée de réconciliation stockée,
    # elle ne doit jamais dériver de quelques centimes sur un fort volume.
    cash_attendu   = _cash_ventes_decimal(ventes) + _cash_remboursements_collectes(s.caissier, s.date_session, db)
    montant_compte = Decimal(str(body.montant_compte))

    s.cash_attendu_soumission = cash_attendu
    s.montant_compte          = montant_compte
    s.ecart                   = montant_compte - cash_attendu
    s.statut                  = "SOUMIS"
    s.soumis_at               = datetime.now(tz=timezone.utc)
    if body.notes:
        s.notes_admin = body.notes
    db.commit()
    return _session_dict(s, stats)


class ValiderIn(BaseModel):
    notes: Optional[str] = None


@router.post("/sessions/{session_id}/valider")
def valider_session(session_id: int, body: ValiderIn = ValiderIn(), request: Request = None, db: Session = Depends(get_db)):
    """Admin valide la session — passe en VALIDE."""
    s = _session_ou_404(session_id, db)
    if s.statut == "EN_COURS":
        raise HTTPException(409, "La session doit d'abord être soumise par la caissière avant validation.")
    s.statut        = "VALIDE"
    s.valide_at     = datetime.now(tz=timezone.utc)
    s.valide_par_id = _uid(request) if request else None
    if body.notes:
        s.notes_admin = body.notes
    db.commit()
    ventes = _ventes_session(s, db)
    return _session_dict(s, _stats_ventes(ventes, s.caissier, s.date_session, db))


class EvaluationIn(BaseModel):
    evaluations: dict[str, str]   # {produit_nom: "CORRECT"|"NON_CORRECT"|"INTROUVABLE"}
    finaliser: bool = False


@router.post("/sessions/{session_id}/evaluer")
def evaluer_session(
    session_id: int,
    body: EvaluationIn,
    db: Session = Depends(get_db),
    _user: Utilisateur = Depends(_require_pdg_ou_admin_pos),
):
    """Évaluation produit par produit d'un rapport soumis (page Rapports
    Soumis). Sauvegarde progressive (finaliser=False) puis finalisation
    (finaliser=True) qui calcule le score et valide directement la session
    — évite un double clic redondant avec /valider."""
    s = _session_ou_404(session_id, db)
    valides = {"CORRECT", "NON_CORRECT", "INTROUVABLE"}
    for nom, statut in body.evaluations.items():
        if statut not in valides:
            raise HTTPException(422, f"Statut invalide pour '{nom}': {statut}")
        existant = db.query(BarSessionEvaluation).filter_by(session_id=session_id, produit_nom=nom).first()
        if existant:
            existant.statut = statut
            existant.evalue_par_id = _user.id
        else:
            db.add(BarSessionEvaluation(session_id=session_id, produit_nom=nom, statut=statut, evalue_par_id=_user.id))
    db.commit()

    if body.finaliser:
        ventes = _ventes_session(s, db)
        stats  = _stats_ventes(ventes, s.caissier, s.date_session, db)
        articles_attendus = {p["nom"] for p in stats["tous_produits"]}
        evals = db.query(BarSessionEvaluation).filter_by(session_id=session_id).all()
        evalues_noms = {e.produit_nom for e in evals}
        if articles_attendus - evalues_noms:
            raise HTTPException(422, "Tous les articles doivent être évalués avant de finaliser.")
        nb_correct = sum(1 for e in evals if e.statut == "CORRECT")
        score = round(100 * nb_correct / len(evals), 2) if evals else 0.0
        s.score              = score
        s.evaluation_statut  = "TERMINEE"
        s.evalue_par_id      = _user.id
        s.evalue_le          = datetime.now(tz=timezone.utc)
        # Finaliser l'évaluation valide directement le rapport.
        s.statut        = "VALIDE"
        s.valide_at      = datetime.now(tz=timezone.utc)
        s.valide_par_id  = _user.id
        db.commit()

    ventes = _ventes_session(s, db)
    return _session_dict(s, _stats_ventes(ventes, s.caissier, s.date_session, db))


# ── Exports PDF / XLSX ────────────────────────────────────────────────

def _build_rapport_data(session_id: int, db: Session):
    s      = _session_ou_404(session_id, db)
    ventes = _ventes_session(s, db)
    stats  = _stats_ventes(ventes, s.caissier, s.date_session, db)
    return s, ventes, stats


@router.get("/sessions/{session_id}/rapport.xlsx")
def export_xlsx(session_id: int, db: Session = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl non installé")

    s, ventes, stats = _build_rapport_data(session_id, db)
    caissier_nom     = (s.caissier.nom + " " + s.caissier.prenom) if s.caissier else "Inconnu"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport de caisse"

    # Palette
    ORANGE  = "E8893A"
    DARK    = "1A1A2E"
    LIGHT   = "F5F5F5"
    WHITE   = "FFFFFF"
    BORDER  = Side(style="thin", color="CCCCCC")
    thin    = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

    def hdr_cell(ws, row, col, val, bg=ORANGE, fg=WHITE, bold=True, center=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color=fg, size=11)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = thin
        if center:
            c.alignment = Alignment(horizontal="center", vertical="center")
        return c

    def data_cell(ws, row, col, val, bold=False, fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, size=10)
        c.border = thin
        c.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            c.number_format = fmt
        return c

    # Titre
    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1, value=f"Rapport de Caisse — {caissier_nom}")
    t.font = Font(bold=True, size=14, color=WHITE)
    t.fill = PatternFill("solid", fgColor=DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    d = ws.cell(row=2, column=1,
                value=f"Date : {s.date_session}  |  Statut : {s.statut}  |  Ventes : {stats['nb_ventes']}  |  Total : G {stats['total']:,.2f}")
    d.font = Font(size=10, color="555555")
    d.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    # Résumé financier
    r = 4
    hdr_cell(ws, r, 1, "Résumé financier", bg=DARK)
    ws.merge_cells(f"A{r}:B{r}")

    resume_rows = [
        ("Total encaissé (G)",   stats["total"]),
        ("Cash / Mixte (G)",     stats["cash"]),
        ("Crédit (G)",           stats["credit"]),
        ("Nombre de ventes",     stats["nb_ventes"]),
    ]
    if s.montant_compte is not None:
        resume_rows.append(("Montant compté (G)", float(s.montant_compte)))
        resume_rows.append(("Écart (G)",           float(s.ecart)))

    for label, val in resume_rows:
        r += 1
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=1).border = thin
        c = ws.cell(row=r, column=2, value=val)
        c.border = thin
        c.number_format = '#,##0.00' if isinstance(val, float) else '0'
        if label == "Écart (G)":
            c.font = Font(bold=True, color=_ecart_couleur(float(s.ecart), float(s.cash_attendu_soumission)))

    # Top produits
    r += 2
    hdr_cell(ws, r, 1, "Top produits", bg=DARK)
    ws.merge_cells(f"A{r}:C{r}")
    r += 1
    for lbl, col in [("Produit", 1), ("Qté", 2), ("Total G", 3)]:
        hdr_cell(ws, r, col, lbl)
    for prod in stats["top_produits"]:
        r += 1
        data_cell(ws, r, 1, prod["nom"])
        data_cell(ws, r, 2, prod["quantite"], fmt="0.##")
        data_cell(ws, r, 3, prod["total"],    fmt="#,##0.00")

    # Détail ventes
    r += 2
    headers = ["Ticket", "Heure", "Mode", "Client", "Produits", "Total G", "Payé G"]
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, r, ci, h)
    for v in ventes:
        r += 1
        produits_str = " / ".join(
            f"{l.produit.nom if l.produit else l.cuisine_plat.nom if l.cuisine_plat else '?'} x{float(l.quantite):.0f}"
            for l in v.lignes
        )
        data_cell(ws, r, 1, v.numero_ticket)
        data_cell(ws, r, 2, v.date_heure.astimezone(HAITI_TZ).strftime("%H:%M"))
        data_cell(ws, r, 3, v.mode_paiement)
        data_cell(ws, r, 4, v.client_nom or "")
        ws.cell(row=r, column=5, value=produits_str).border = thin
        data_cell(ws, r, 6, float(v.montant_total), fmt="#,##0.00")
        data_cell(ws, r, 7, float(v.montant_paye),  fmt="#,##0.00")

    # Évaluation du rapport (si terminée)
    if s.evaluation_statut == "TERMINEE":
        eval_couleur = {"CORRECT": "22C55E", "NON_CORRECT": "F7A93B", "INTROUVABLE": "F87171"}
        r += 2
        hdr_cell(ws, r, 1, "Évaluation du rapport", bg=DARK)
        ws.merge_cells(f"A{r}:C{r}")
        r += 1
        evalue_par_nom = (s.evalue_par.nom + " " + s.evalue_par.prenom) if s.evalue_par else "?"
        ws.cell(row=r, column=1, value="Score final").font = Font(bold=True)
        ws.cell(row=r, column=1).border = thin
        c = ws.cell(row=r, column=2, value=f"{float(s.score):.0f} %" if s.score is not None else "—")
        c.border = thin
        c.font = Font(bold=True, color=_score_couleur(float(s.score) if s.score is not None else None))
        r += 1
        ws.cell(row=r, column=1, value="Évalué par").font = Font(bold=True)
        ws.cell(row=r, column=1).border = thin
        ws.cell(row=r, column=2, value=evalue_par_nom).border = thin
        r += 2
        for lbl, col in [("Article", 1), ("Statut", 2)]:
            hdr_cell(ws, r, col, lbl)
        for e in sorted(s.evaluations, key=lambda e: e.produit_nom):
            r += 1
            data_cell(ws, r, 1, e.produit_nom)
            sc = data_cell(ws, r, 2, e.statut)
            sc.font = Font(bold=True, color=eval_couleur.get(e.statut, "555555"))

    # Largeurs colonnes
    for col, width in [(1,18),(2,9),(3,10),(4,18),(5,45),(6,14),(7,14)]:
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"rapport_caisse_{caissier_nom.replace(' ','_')}_{s.date_session}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/sessions/{session_id}/rapport.pdf")
def export_pdf(session_id: int, db: Session = Depends(get_db)):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        raise HTTPException(500, "reportlab non installé")

    s, ventes, stats = _build_rapport_data(session_id, db)
    caissier_nom     = (s.caissier.nom + " " + s.caissier.prenom) if s.caissier else "Inconnu"

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4,
                              leftMargin=1.8*cm, rightMargin=1.8*cm,
                              topMargin=1.8*cm, bottomMargin=1.8*cm)
    styles = getSampleStyleSheet()
    ORANGE = colors.HexColor("#E8893A")
    DARK   = colors.HexColor("#1A1A2E")

    title_style  = ParagraphStyle("title",  fontSize=16, textColor=DARK,   fontName="Helvetica-Bold", spaceAfter=4)
    sub_style    = ParagraphStyle("sub",    fontSize=10, textColor=colors.grey, spaceAfter=12)
    section_style = ParagraphStyle("sect",  fontSize=12, textColor=DARK,   fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)

    story = []
    story.append(Paragraph(f"Rapport de Caisse — {caissier_nom}", title_style))
    story.append(Paragraph(
        f"Date : <b>{s.date_session}</b> &nbsp;|&nbsp; Statut : <b>{s.statut}</b>",
        sub_style,
    ))
    story.append(HRFlowable(width="100%", thickness=1.5, color=ORANGE))
    story.append(Spacer(1, 10))

    # Résumé
    story.append(Paragraph("Résumé financier", section_style))
    resume_data = [
        ["Indicateur",         "Valeur"],
        ["Total encaissé",     f"G {stats['total']:,.2f}"],
        ["Cash / Mixte",       f"G {stats['cash']:,.2f}"],
        ["Crédit",             f"G {stats['credit']:,.2f}"],
        ["Nombre de ventes",   str(stats["nb_ventes"])],
    ]
    ecart_row_idx = None
    if s.montant_compte is not None:
        resume_data.append(["Montant compté", f"G {float(s.montant_compte):,.2f}"])
        ecart_row_idx = len(resume_data)
        signe = '+' if float(s.ecart) >= 0 else ''
        resume_data.append(["Écart", f"{signe}G {float(s.ecart):,.2f}"])
    for mode, val in stats["par_mode"].items():
        resume_data.append([f"Mode {mode}", f"G {val:,.2f}"])

    t_resume = Table(resume_data, colWidths=[8*cm, 5*cm])
    resume_style = [
        ("BACKGROUND",   (0,0),(1,0), DARK),
        ("TEXTCOLOR",    (0,0),(1,0), colors.white),
        ("FONTNAME",     (0,0),(1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0),(-1,-1), 9),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.HexColor("#F5F5F5"), colors.white]),
        ("GRID",         (0,0),(-1,-1), 0.4, colors.HexColor("#CCCCCC")),
        ("ALIGN",        (1,0),(1,-1), "RIGHT"),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("TOPPADDING",   (0,0),(-1,-1), 5),
    ]
    if ecart_row_idx is not None:
        ecart_couleur = colors.HexColor(f"#{_ecart_couleur(float(s.ecart), float(s.cash_attendu_soumission))}")
        resume_style.append(("TEXTCOLOR", (0,ecart_row_idx),(1,ecart_row_idx), ecart_couleur))
        resume_style.append(("FONTNAME",  (0,ecart_row_idx),(1,ecart_row_idx), "Helvetica-Bold"))
    t_resume.setStyle(TableStyle(resume_style))
    story.append(t_resume)
    story.append(Spacer(1, 10))

    # Top produits
    if stats["top_produits"]:
        story.append(Paragraph("Top produits", section_style))
        prod_data = [["Produit", "Qté", "Total G"]]
        for p in stats["top_produits"]:
            prod_data.append([p["nom"], f"{p['quantite']:.1f}", f"G {p['total']:,.2f}"])
        t_prod = Table(prod_data, colWidths=[9*cm, 3*cm, 4*cm])
        t_prod.setStyle(TableStyle([
            ("BACKGROUND",   (0,0),(-1,0), ORANGE),
            ("TEXTCOLOR",    (0,0),(-1,0), colors.white),
            ("FONTNAME",     (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0),(-1,-1), 9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.HexColor("#FFF8F0"), colors.white]),
            ("GRID",         (0,0),(-1,-1), 0.4, colors.HexColor("#CCCCCC")),
            ("ALIGN",        (1,0),(-1,-1), "CENTER"),
            ("BOTTOMPADDING",(0,0),(-1,-1), 5),
            ("TOPPADDING",   (0,0),(-1,-1), 5),
        ]))
        story.append(t_prod)
        story.append(Spacer(1, 10))

    # Détail ventes
    story.append(Paragraph("Détail des ventes", section_style))
    vente_data = [["Ticket", "Heure", "Mode", "Client", "Total G", "Payé G"]]
    for v in ventes:
        vente_data.append([
            v.numero_ticket,
            v.date_heure.astimezone(HAITI_TZ).strftime("%H:%M"),
            v.mode_paiement,
            (v.client_nom or "")[:20],
            f"G {float(v.montant_total):,.2f}",
            f"G {float(v.montant_paye):,.2f}",
        ])
    t_ventes = Table(vente_data, colWidths=[3.5*cm, 2*cm, 2.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
    t_ventes.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(-1,0), DARK),
        ("TEXTCOLOR",    (0,0),(-1,0), colors.white),
        ("FONTNAME",     (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0),(-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.HexColor("#F5F5F5"), colors.white]),
        ("GRID",         (0,0),(-1,-1), 0.3, colors.HexColor("#DDDDDD")),
        ("ALIGN",        (1,0),(-1,-1), "CENTER"),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ("TOPPADDING",   (0,0),(-1,-1), 4),
    ]))
    story.append(t_ventes)

    # Évaluation du rapport (si terminée)
    if s.evaluation_statut == "TERMINEE":
        eval_couleur = {
            "CORRECT":      colors.HexColor("#22C55E"),
            "NON_CORRECT":  colors.HexColor("#F7A93B"),
            "INTROUVABLE":  colors.HexColor("#F87171"),
        }
        story.append(Spacer(1, 14))
        story.append(Paragraph("Évaluation du rapport", section_style))
        evalue_par_nom = (s.evalue_par.nom + " " + s.evalue_par.prenom) if s.evalue_par else "?"
        score_txt = f"{float(s.score):.0f} %" if s.score is not None else "—"
        story.append(Paragraph(
            f"Score final : <b>{score_txt}</b> &nbsp;|&nbsp; Évalué par : <b>{evalue_par_nom}</b>",
            sub_style,
        ))
        evals_tries = sorted(s.evaluations, key=lambda e: e.produit_nom)
        eval_data = [["Article", "Statut"]] + [[e.produit_nom, e.statut] for e in evals_tries]
        t_eval = Table(eval_data, colWidths=[9*cm, 4*cm])
        eval_style = [
            ("BACKGROUND",   (0,0),(-1,0), DARK),
            ("TEXTCOLOR",    (0,0),(-1,0), colors.white),
            ("FONTNAME",     (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0),(-1,-1), 9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.HexColor("#F5F5F5"), colors.white]),
            ("GRID",         (0,0),(-1,-1), 0.4, colors.HexColor("#CCCCCC")),
            ("ALIGN",        (1,0),(-1,-1), "CENTER"),
            ("BOTTOMPADDING",(0,0),(-1,-1), 5),
            ("TOPPADDING",   (0,0),(-1,-1), 5),
        ]
        for i, e in enumerate(evals_tries, start=1):
            eval_style.append(("TEXTCOLOR", (1,i),(1,i), eval_couleur.get(e.statut, colors.grey)))
            eval_style.append(("FONTNAME",  (1,i),(1,i), "Helvetica-Bold"))
        t_eval.setStyle(TableStyle(eval_style))
        story.append(t_eval)

    # Footer
    story.append(Spacer(1, 16))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    story.append(Paragraph(
        f"Généré le {datetime.now(tz=timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC — Konekta · Bon Prix",
        ParagraphStyle("footer", fontSize=7, textColor=colors.grey, alignment=TA_CENTER, spaceBefore=4),
    ))

    doc.build(story)
    buf.seek(0)
    filename = f"rapport_caisse_{caissier_nom.replace(' ','_')}_{s.date_session}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
